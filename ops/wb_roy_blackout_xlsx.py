#!/usr/bin/env python3
# поток: mkt
"""ops/wb_roy_blackout_xlsx.py — список ⚫ «вывод из рекламы» для ручного удаления в ЛК.

Метода убрать номенклатуру из кампании у ВБ в API НЕТ (проверено 10.08.2026): ставку мы
опускаем на пол сами (`wb_roy_apply --blackout`), а саму карточку из кампании удаляет человек
в личном кабинете. Этот отчёт — рабочий лист для такого удаления: сгруппирован ПО КАМПАНИЯМ,
потому что в ЛК ходят кампания за кампанией, а не SKU за SKU.

⚫ означает «маржа ниже KPI 25 % ещё до рекламы» → допустимый ДРР = маржа − 25 ≤ 0.
Из продажи товар НЕ убираем: органика по нему остаётся.

Запуск:
  ./venv/bin/python -m ops.wb_roy_blackout_xlsx 2026-08-16
"""
import csv
import sys
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

BASE_DIR = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
from core import db  # noqa: E402

ACC = 'wb_acc1'
HEAD = ['кампания', 'ID кампании', 'артикул', 'nm_id', 'название', 'ставка ₽',
        'маржа %', 'расход ₽/нед', 'выручка ₽/нед', 'заказы', 'что делать']


def main(tag):
    rep = BASE_DIR / 'docs' / 'reports'
    src = rep / f'mkt_roy_profile_{tag}.csv'
    cards = {r['nm_id']: (r['vendor_code'] or '', r['title'] or '')
             for r in db.query("select nm_id, vendor_code, title from wb_cards where account=%s", (ACC,))}
    camp = {str(r['advert_id']): r['name'] or ''
            for r in db.query("select distinct on (advert_id) advert_id, name from wb_ads "
                              "where account=%s order by advert_id, period desc", (ACC,))}

    rows = []
    with open(src, encoding='utf-8-sig') as fh:
        for r in csv.DictReader(fh, delimiter=';'):
            if not r['цвет'].startswith('⚫'):
                continue
            nm = int(r['nm_id'])
            art, title = cards.get(nm, ('', ''))
            rows.append([camp.get(r['advert_id'], ''), int(r['advert_id']), art, nm, title,
                         float(r['ставка_₽']), float(r['маржа_live_%'] or 0),
                         float(r['расход_₽']), float(r['выручка_ВСЯ_₽']),
                         int(float(r['заказы_всего'] or 0)),
                         'удалить из кампании'])
    rows.sort(key=lambda x: (x[0], -x[7]))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Вывод из рекламы'
    ws.append(HEAD)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDDDDD')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for r in rows:
        ws.append(r)
    for i, w in enumerate([34, 12, 16, 12, 52, 9, 9, 12, 13, 8, 20], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    out = rep / f'mkt_wb_blackout_{tag}.xlsx'
    wb.save(out)
    csv_out = rep / f'mkt_wb_blackout_{tag}.csv'
    with open(csv_out, 'w', encoding='utf-8-sig', newline='') as fh:
        w = csv.writer(fh, delimiter=';')
        w.writerow(HEAD)
        w.writerows(rows)
    print(f"{len(rows)} SKU в {len({r[1] for r in rows})} кампаниях")
    print(f"расход {sum(r[7] for r in rows):.0f} ₽/нед · выручка {sum(r[8] for r in rows):.0f} ₽/нед")
    print(out)
    print(csv_out)


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else '2026-08-16')
