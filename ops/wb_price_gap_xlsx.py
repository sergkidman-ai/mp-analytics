#!/usr/bin/env python3
# поток: mkt
"""ops/wb_price_gap_xlsx.py — разрывы цены против выдачи ВБ, лист для ручного разбора.

Вход — json съёма выдачи (`mkt_wb_serp_y50_<дата>.json`, поля nm/query/qsrc/pos/card_price/
med10/our_fb/med10_fb/our_qty/rivals_cheaper). Считаем отношение нашей цены к медиане топ-10
по тому же запросу и выкладываем всё, что дороже в 1.5 раза и больше.

Зачем руками: разрыв означает ЛИБО что мы промахнулись с ценой, ЛИБО что сосед по запросу
торгует другим товаром (совместимый аналог против оригинала). Различить может только человек,
поэтому в файле рядом лежат запрос, позиция, отзывы и остаток — всё, что нужно для взгляда.

Строки с пометкой «из названия» в столбце «источник запроса» сравнивались по запросу,
собранному из заголовка карточки, а не из Джема — сравнение там слабее, это надо держать в уме.

Запуск:
  ./venv/bin/python -m ops.wb_price_gap_xlsx docs/reports/mkt_wb_serp_y50_2026-08-17.json
"""
import csv
import json
import sys
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

BASE_DIR = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
from core import db  # noqa: E402

ACC = 'wb_acc1'
MIN_GAP = 1.5
HEAD = ['разрыв', 'артикул', 'nm_id', 'название', 'запрос', 'источник запроса', 'наша позиция',
        'наша цена ₽', 'медиана топ-10 ₽', 'отзывов у нас', 'отзывов у соседей',
        'соседей дешевле', 'остаток', 'вывод глазами']
RED = PatternFill('solid', fgColor='F8D7DA')


def main(path):
    rows_in = json.load(open(path, encoding='utf-8'))
    cards = {r['nm_id']: (r['vendor_code'] or '', r['title'] or '')
             for r in db.query("select nm_id, vendor_code, title from wb_cards where account=%s", (ACC,))}

    out = []
    comparable = 0
    for r in rows_in:
        if not r.get('med10') or not r.get('card_price'):
            continue
        comparable += 1
        gap = r['card_price'] / r['med10']
        if gap < MIN_GAP:
            continue
        art, title = cards.get(r['nm'], ('', ''))
        out.append([round(gap, 2), art, r['nm'], title, r['query'], r['qsrc'],
                    r['pos'] if r.get('pos') is not None else 'нет в топ-100',
                    round(r['card_price']), round(r['med10']), r.get('our_fb') or 0,
                    r.get('med10_fb') or 0, r.get('rivals_cheaper') or 0, r.get('our_qty') or 0, ''])
    out.sort(key=lambda x: -x[0])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Разрывы цены'
    ws.append(HEAD)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDDDDD')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for r in out:
        ws.append(r)
        if r[0] >= 3:
            for c in ws[ws.max_row]:
                c.fill = RED
    for i, w in enumerate([9, 16, 12, 46, 34, 15, 14, 12, 15, 13, 15, 15, 10, 30], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    tag = pathlib.Path(path).stem.split('_')[-1]
    rep = BASE_DIR / 'docs' / 'reports'
    xl = rep / f'mkt_wb_price_gap_{tag}.xlsx'
    wb.save(xl)
    cs = rep / f'mkt_wb_price_gap_{tag}.csv'
    with open(cs, 'w', encoding='utf-8-sig', newline='') as fh:
        w = csv.writer(fh, delimiter=';')
        w.writerow(HEAD)
        w.writerows(out)
    print(f"сравнимых {comparable} из {len(rows_in)} · в файле {len(out)} (разрыв ≥{MIN_GAP}x), "
          f"из них ≥3x — {sum(1 for r in out if r[0] >= 3)} (подсвечены красным)")
    print(xl)
    print(cs)


if __name__ == '__main__':
    main(sys.argv[1])
