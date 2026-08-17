#!/usr/bin/env python3
# поток: mkt
"""ops/wb_roy_add_xlsx.py — список «кого завести в рекламу» в Excel.

Собирает два CSV недельного Роя (`mkt_roy_add_core_*.csv` — ядро на трёх месяцах,
`mkt_roy_add_*.csv` — недельный срез) в одну книгу с двумя листами и подтягивает
название/артикул карточки из wb_cards, чтобы список читался без БД.

Ядро на первом листе намеренно: недельный срез — это по одному заказу на позицию,
решение по нему принимать нельзя, он лежит рядом для полноты картины.

Запуск:
  ./venv/bin/python -m ops.wb_roy_add_xlsx 2026-08-16
"""
import csv
import sys
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
from core import db  # noqa: E402

ACC = 'wb_acc1'
SHEETS = [('Ядро (3 мес)', 'mkt_roy_add_core_{}.csv'),
          ('За неделю', 'mkt_roy_add_{}.csv')]


def _cards():
    return {r['nm_id']: (r['vendor_code'] or '', r['title'] or '')
            for r in db.query("select nm_id, vendor_code, title from wb_cards where account=%s",
                              (ACC,))}


def main(tag):
    rep = BASE_DIR / 'docs' / 'reports'
    cards = _cards()
    wb = Workbook()
    wb.remove(wb.active)
    for name, pat in SHEETS:
        src = rep / pat.format(tag)
        if not src.exists():
            continue
        with open(src, encoding='utf-8-sig') as fh:
            rows = list(csv.reader(fh, delimiter=';'))
        head, body = rows[0], rows[1:]
        ws = wb.create_sheet(name)
        ws.append(head[:1] + ['артикул', 'название'] + head[1:])
        for r in body:
            nm = int(r[0])
            vc, ti = cards.get(nm, ('', ''))
            ws.append([nm, vc, ti] + [float(x.replace(',', '.')) if x.replace(',', '.')
                                      .replace('.', '', 1).replace('-', '', 1).isdigit() else x
                                      for x in r[1:]])
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for i, _ in enumerate(ws[1], 1):
            w = max((len(str(ws.cell(row=r, column=i).value or '')) for r in range(1, ws.max_row + 1)),
                    default=10)
            ws.column_dimensions[get_column_letter(i)].width = min(46, max(9, w + 2))
    out = rep / f"mkt_roy_add_{tag}.xlsx"
    wb.save(out)
    print(f"{out}  ({', '.join(ws.title + ': ' + str(ws.max_row - 1) for ws in wb)})")


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else '2026-08-16')
