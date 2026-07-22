# -*- coding: utf-8 -*-
"""
invoice_to_po.py — единый конвейер: файл счёта поставщика (XLS/XLSX/PDF или ссылка Я.Диска)
                   → черновик «Заказ поставщику» в МойСклад.

Фаза 1. Создаём ТОЛЬКО purchaseorder, applicable=false (черновик). Счёт/УПД не трогаем.

Запуск:
  python invoice_to_po.py <файл|ссылка Я.Диска> [--create] [--suffix ТЕСТ]
  По умолчанию — DRY-RUN (парсит, матчит, считает, показывает payload, НО не постит).
  --create  — реально создать черновик (с проверкой уникальности номера).
  --suffix S — добавить S к номеру заказа (для тестов, чтобы не ловить коллизии).

Собранные правила (все согласованы с заказчиком, см. память ms-purchase-order-automation-recon):
  • Покупатель (ИНН из счёта) → организация + склад.
  • Поставщик (ИНН из счёта) → контрагент-agent; для Солюшнс принт всегда карточка «МСК».
  • Товары по нативному полю article; per-поставщик стратегия извлечения артикула.
  • Группы поставщиков (supplier_groups): товар должен принадлежать ГРУППЕ; иначе — предупреждение в «Комментарий».
  • Один article у товаров разных групп → берём товар своей группы (pick_in_group), факт — в «Комментарий».
  • Артикул не найден → строку пропускаем, показываем пользователю, итог подгоняем под (счёт − пропуски).
  • Доставка → услуга «Доставка заказа».
  • НДС 22% включ. по умолчанию; поставщик с профилем novat=True — «Без НДС» (vatEnabled=false, vat=0).
  • Цены руб→коп (×100); «Ожидание»=inTransit=quantity.
  • moment = дата счёта + текущее время МСК; deliveryPlannedMoment = +1 рабочий день (workcal, праздники, 6-дневка).
  • Итог заказа подгоняем под сумму счёта (правка цены последней товарной строки).
  • name заказа = номер счёта; перед --create проверяем уникальность, при коллизии — СТОП.
  • state = «Закрыт»; applicable=false.
"""
import os, sys, re, json, gzip, time, subprocess, urllib.request, urllib.parse, urllib.error
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ms import get, post, put     # noqa
import workcal                    # noqa
import supplier_groups as SG      # noqa

MSU = "https://api.moysklad.ru/api/remap/1.2"
MSK = ZoneInfo("Europe/Moscow")
STATE_CLOSED = "1c1d62af-46e5-11ee-0a80-01d30002d662"
SIX_INN = {"7806486149", "7725744338"}          # субботы рабочие
AGENT_OVERRIDE = {"7806486149": "9263b7de-97bf-11f0-0a80-015900374bc3"}  # Солюшнс принт → карточка «МСК»

# ── Покупатели: ИНН из счёта → организация (по ИНН) + склад (по имени) ──────────
BUYERS = {
    "7807355364": {"org_inn": "7807355364", "store": "звездный"},  # ЦИФРОВОЙ КВАДРАТ → Звездный
    "7811803918": {"org_inn": "7811803918", "store": "дисквер"},   # ДИСКВЭР → Дисквэр
}
# ── Поставщики: ИНН из счёта → профиль парсинга/извлечения артикула ─────────────
# article: column | sp | name_last | name_regex(pattern) ; six берётся из SIX_INN.
SUPPLIERS = {
    "7806486149": {"name": "Солюшнс принт МСК", "article": "sp"},
    "7730244274": {"name": "Одиссей",           "article": "column"},
    "9717092410": {"name": "Тонерстор",         "article": "column", "pdf": "tonerstor"},
    "7718978470": {"name": "Блоссом",           "article": "column"},
    "7725744338": {"name": "Тонеропттторг",     "article": "column"},
    "9731107362": {"name": "Феррет",            "article": "name_regex",
                   "pattern": r"\b(?:CS|GG|CR)-[0-9A-Za-z]+(?:[/-][0-9A-Za-z]+)*", "pdf": "ferret"},
    "7736123276": {"name": "Позитив",           "article": "name_last"},
    "7840480595": {"name": "Колортек",          "article": "column"},
    "7722341813": {"name": "КВК Трейд",         "article": "column"},
    "9718075418": {"name": "Картридж Трейд (Блоссом)", "article": "name_regex",
                   "pattern": r"\bBS-[0-9A-Za-z]+(?:-[0-9A-Za-z]+)*", "novat": True,
                   "year_suffix_on_collision": True},  # сбрасывает нумерацию по годам → развести суффиксом года
    "7719482878": {"name": "КПД",                "article": "column"},
}
WD = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
MONTHS = {"января":1,"февраля":2,"марта":3,"апреля":4,"мая":5,"июня":6,"июля":7,
          "августа":8,"сентября":9,"октября":10,"ноября":11,"декабря":12}


def meta(ent, i, t=None):
    return {"meta": {"href": f"{MSU}/entity/{ent}/{i}", "type": t or ent, "mediaType": "application/json"}}


def get_r(path, tries=6):
    """get() с ретраем и backoff на 429 (rate limit МС)."""
    for a in range(tries):
        try:
            return get(path)
        except urllib.error.HTTPError as e:
            if e.code == 429 and a < tries - 1:
                time.sleep(1.5 * (a + 1)); continue
            raise


# ═══════════════════════ 1. Загрузка и чтение файла ═══════════════════════════
def fetch(src):
    """Локальный путь или публичная ссылка Я.Диска → локальный файл (bytes на диске)."""
    if os.path.exists(src):
        return src
    if "disk.yandex" in src or "yadi.sk" in src:
        api = "https://cloud-api.yandex.net/v1/disk/public/resources/download?public_key=" + urllib.parse.quote(src)
        with urllib.request.urlopen(api, timeout=60) as r:
            href = json.loads(r.read())["href"]
        dst = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_inbox_invoice")
        with urllib.request.urlopen(href, timeout=120) as r:
            data = r.read()
        # определить расширение по сигнатуре
        ext = _sig_ext(data[:8])
        dst += ext
        open(dst, "wb").write(data)
        return dst
    raise SystemExit(f"Не файл и не ссылка Я.Диска: {src}")


def _sig_ext(head):
    if head[:4] == b"%PDF": return ".pdf"
    if head[:2] == b"PK":   return ".xlsx"
    if head[:4] == b"\xD0\xCF\x11\xE0": return ".xls"
    return ".bin"


def _fix_sharedstrings(path):
    """Пересобрать xlsx, переименовав xl/SharedStrings.xml → xl/sharedStrings.xml (регистр).
    Некоторые не-Excel генераторы пишут заглавную S, а openpyxl читает строго строчную."""
    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(path) as z, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as o:
        for it in z.infolist():
            data = z.read(it.filename)
            name = it.filename
            if name.lower() == "xl/sharedstrings.xml":
                name = "xl/sharedStrings.xml"
                data = data.replace(b"SharedStrings.xml", b"sharedStrings.xml")
            o.writestr(name, data)
    fixed = path + ".fixed.xlsx"
    open(fixed, "wb").write(buf.getvalue())
    return fixed


def _read_xlsx(path):
    """xlsx → rows[][]; при заглавном SharedStrings.xml (KeyError) чиним и перечитываем."""
    import openpyxl
    try:
        ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]
    except KeyError:
        if open(path, "rb").read(2) != b"PK":
            raise
        ws = openpyxl.load_workbook(_fix_sharedstrings(path), data_only=True).worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def read_grid(path):
    """→ (kind, payload): ('table', rows[][]) для xls/xlsx или ('pdf', text) для pdf."""
    head = open(path, "rb").read(8)
    if head[:4] == b"%PDF":
        txt = subprocess.run(["pdftotext", "-layout", "-enc", "UTF-8", path, "-"],
                             capture_output=True, text=True).stdout
        return "pdf", txt
    if head[:2] == b"PK":
        return "table", _read_xlsx(path)
    if head[:4] == b"\xD0\xCF\x11\xE0":
        import xlrd
        try:
            sh = xlrd.open_workbook(path).sheet_by_index(0)
            rows = []
            for r in range(sh.nrows):
                rows.append([(int(v) if isinstance(v, float) and v == int(v) else v) for v in sh.row_values(r)])
            return "table", rows
        except Exception:
            # некоторые генераторы (напр. SolutionPrint/s-print.ru) пишут CDF-контейнер,
            # который xlrd не может разобрать («Expected BOF record») — calamine (Rust) читает такие
            return "table", _read_xls_calamine(path)
    raise SystemExit("Неизвестный формат файла")


def _read_xls_calamine(path):
    from python_calamine import CalamineWorkbook
    sheet = CalamineWorkbook.from_path(path).get_sheet_by_index(0)
    rows = []
    for row in sheet.to_python():
        rows.append([(int(v) if isinstance(v, float) and v == int(v) else v) for v in row])
    return rows


def grid_text(kind, payload):
    if kind == "pdf":
        return payload
    return "\n".join(" ".join(str(v) for v in row if v not in ("", None)) for row in payload)


# ═══════════════════════ 2. Шапка счёта ═══════════════════════════════════════
def parse_header(text):
    """→ dict(number, inv_date, total, buyer_inn, supplier_inn)."""
    h = {}
    # номер + дата
    m = re.search(r"(?:СЧ[ЁЕ]Т|Счет|Счёт)[^\n№]*№\s*(\S+)\s+от\s+([^\n]+)", text)
    if not m:
        raise SystemExit("Не найдена строка «Счёт № … от …»")
    h["number"] = m.group(1).strip()
    h["inv_date"] = _parse_date(m.group(2))
    # ИНН покупателя/поставщика — по известным множествам
    h["buyer_inn"] = next((i for i in BUYERS if re.search(rf"(?<!\d){i}(?!\d)", text)), None)
    if not h["buyer_inn"]:  # фолбэк по имени (в шаблоне SP ИНН покупателя не печатается)
        low = text.lower()
        for kw, inn in (("цифровой квадрат", "7807355364"), ("дисквэр", "7811803918"), ("дисквер", "7811803918")):
            if kw in low:
                h["buyer_inn"] = inn; break
    h["supplier_inn"] = next((i for i in SUPPLIERS if re.search(rf"(?<!\d){i}(?!\d)", text)), None)
    # итог по НДС/оплате (в порядке приоритета)
    h["total"] = _parse_total(text)
    return h


def _parse_date(s):
    s = s.strip()
    m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", s)
    if m:
        d, mo, y = map(int, m.groups())
        if y < 100: y += 2000
        return date(y, mo, d)
    m = re.match(r"(\d{1,2})\s+([А-Яа-я]+)\s+(\d{4})", s)
    if m and m.group(2).lower() in MONTHS:
        return date(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1)))
    raise SystemExit(f"Не распознана дата счёта: {s!r}")


def _parse_total(text):
    pats = [r"Итого\s*с\s*НДС", r"Всего\s*к\s*оплате", r"Итого\s*к\s*оплате", r"Итого,?\s*руб", r"Итого"]
    for p in pats:
        for line in text.splitlines():
            if re.search(p, line, re.I):
                nums = re.findall(r"(\d[\d ]*[.,]\d{2})", line)
                if nums:
                    return float(nums[-1].replace(" ", "").replace(",", "."))
    return None


# ═══════════════════════ 3. Позиции счёта ═════════════════════════════════════
QTY_L = {"кол-во", "количество"}; ART_L = {"артикул", "код"}
def _norm(v): return str(v).strip().lower() if v not in ("", None) else ""


def parse_table(rows):
    """Заголовок-ориентированный разбор xls/xlsx → список позиций."""
    hdr = None
    for r, row in enumerate(rows):
        low = [_norm(v) for v in row]
        has_num = any(x == "№" or x.startswith("№") for x in low)
        has_price = any(x == "цена" or x.startswith("цена") for x in low)
        if any(x in QTY_L for x in low) and has_price and has_num:
            hdr = r; break
    if hdr is None:
        raise SystemExit("Не найдена строка-заголовок таблицы (№ / Кол-во / Цена / Сумма)")
    cols = {}
    for c, v in enumerate(rows[hdr]):
        n = _norm(v)
        # «сумма»-ветка исключает «сумма ндс» (это НДС, не итог по строке) — итог ищем и по «стоимость…»
        # (напр. «Стоимость товаров… всего с налогом»), т.к. не все поставщики называют колонку «Сумма»
        if (n == "№" or n.startswith("№")) and "num" not in cols: cols["num"] = c
        elif n in ART_L and "art" not in cols: cols["art"] = c
        elif (n.startswith("товар") or n.startswith("наименован")) and "name" not in cols: cols["name"] = c
        elif n in QTY_L and "qty" not in cols: cols["qty"] = c
        elif (n == "цена" or n.startswith("цена")) and "price" not in cols: cols["price"] = c
        elif (("сумма" in n and "ндс" not in n) or n.startswith("стоимость")) and "sum" not in cols: cols["sum"] = c
    items = []
    for row in rows[hdr + 1:]:
        joined = " ".join(_norm(v) for v in row)
        if any(k in joined for k in ("итого", "всего к оплате", "всего наименован")):
            break
        num = _cell(row, cols.get("num"))
        if not _is_int(num):
            continue
        name = str(_cell(row, cols.get("name")) or "").strip()
        art = str(_cell(row, cols.get("art")) or "").strip() if "art" in cols else ""
        qty = _num(_cell(row, cols.get("qty")))
        price = _num(_cell(row, cols.get("price")))
        summ = _num(_cell(row, cols.get("sum")))
        if qty is None or price is None:
            continue
        # пропускаем строку-легенду («1 2 3 4 …») и прочий мусор: у товара в названии есть буквы
        if not re.search(r"[A-Za-zА-Яа-я]", name):
            continue
        items.append({"num": int(float(num)), "art_raw": art, "name": name,
                      "qty": qty, "price": price, "sum": summ})
    return items


def parse_pdf_tonerstor(text):
    items = []
    for line in text.splitlines():
        m = re.match(r"^\s*(\d+)\s+(.*?)\s+(\d+)\s+шт\s+([\d ]*\d,\d{2})\s+([\d ]*\d,\d{2})\s*$", line)
        if not m:
            continue
        num, mid, qty, price, summ = m.groups()
        tok = mid.strip().split()[0] if mid.strip() else ""
        art = tok if re.match(r"^[A-Z][A-Z0-9]{3,}$", tok) else ""
        items.append({"num": int(num), "art_raw": art, "name": mid.strip(),
                      "qty": float(qty), "price": _num(price), "sum": _num(summ)})
    return items


def parse_pdf_ferret(text):
    # CS/GG — Cactus, CR — CopyRite (Феррет продаёт обе линейки в одном счёте);
    # артикул может быть многосоставным через дефис и/или слэш (CS-M21-250C342, CS-PGI2400BK/C/M/Y) — берём целиком
    arts = re.findall(r"\b(?:CS|GG|CR)-[0-9A-Za-z]+(?:[/-][0-9A-Za-z]+)*", text)
    rows = []
    # код номенклатуры (2-я колонка) бывает буквенно-цифровым (напр. S24199174609), не только \d+
    for m in re.finditer(r"^\s*(\d+)\s+([0-9A-Za-z]+)\s+.*?(\d+,\d{2})\s+(\d+)\s+шт\s+(\d+,\d{2})\s*$", text, re.M):
        num, kod, price, qty, summ = m.groups()
        rows.append((int(num), float(qty), _num(price), _num(summ)))
    if len(arts) != len(rows):
        raise SystemExit(f"Феррет: рассинхрон артикулов ({len(arts)}) и строк ({len(rows)}) — стоп")
    return [{"num": n, "art_raw": a, "name": a, "qty": q, "price": p, "sum": s}
            for (n, q, p, s), a in zip(rows, arts)]


def _cell(row, c):
    return row[c] if (c is not None and c < len(row)) else None
def _is_int(v):
    try: return float(str(v).replace(",", ".")) == int(float(str(v).replace(",", "."))) and str(v).strip() != ""
    except Exception: return False
def _num(v):
    if v in ("", None): return None
    try: return float(str(v).replace(" ", "").replace(",", "."))
    except Exception: return None


# ═══════════════════════ 4. Извлечение артикула ═══════════════════════════════
def resolve_article(item, prof):
    strat = prof["article"]
    raw = item["art_raw"]; name = item["name"]
    if strat == "sp":        return f"SP{raw}_MSK"
    if strat == "name_last": return name.split()[-1] if name else ""
    if strat == "name_regex":
        m = re.search(prof["pattern"], name)
        return m.group(0) if m else ""
    return raw  # column


def is_delivery(item):
    return "доставк" in (item["name"] or "").lower()


# ═══════════════════════ 5. Разрешение ссылок МС ══════════════════════════════
def resolve_buyer(buyer_inn):
    cfg = BUYERS[buyer_inn]
    org = next(o for o in get("/entity/organization")["rows"] if o.get("inn") == cfg["org_inn"])
    store = next(s for s in get("/entity/store")["rows"] if _norm(s["name"]) == cfg["store"])
    return org, store


def resolve_agent(supplier_inn):
    if supplier_inn in AGENT_OVERRIDE:
        return AGENT_OVERRIDE[supplier_inn]
    rows = get(f"/entity/counterparty?filter=inn={supplier_inn}")["rows"]
    if not rows:
        raise SystemExit(f"Контрагент по ИНН {supplier_inn} не найден")
    return rows[0]["id"]


def resolve_delivery_service():
    svc = get(f"/entity/service?search={urllib.parse.quote('Доставка заказа')}&limit=5")["rows"]
    return next(s for s in svc if s["name"].strip() == "Доставка заказа")["id"]


# ═══════════════════════ 6. Матчинг товаров (задачи 2, 3) ═════════════════════
def match_products(items, prof, group):
    """→ (positions_meta, matched_info, skipped, warnings)."""
    positions, matched_info, skipped, warns = [], [], [], []
    def lsum(it):
        return it["sum"] if it.get("sum") is not None else round(it["qty"] * it["price"], 2)
    for it in items:
        if is_delivery(it):
            positions.append({"_deliv": True, "qty": it["qty"], "price": it["price"], "line_sum": lsum(it)})
            continue
        art = resolve_article(it, prof)
        if not art:
            skipped.append({**it, "art": "", "reason": "не извлечён артикул"}); continue
        cands = get_r(f"/entity/product?filter=article={urllib.parse.quote(art)}&limit=10&expand=supplier").get("rows", [])
        if not cands and "/" in art:
            # в каталоге артикул мультицветного набора иногда обрезан после первого доп. цвета
            # (напр. счёт: CS-I-PG510-CL511M/C/Y, карточка: CS-I-PG510-CL511M/C) — пробуем короче
            parts = art.split("/")
            for k in range(len(parts) - 1, 0, -1):
                trial = "/".join(parts[:k])
                cands = get_r(f"/entity/product?filter=article={urllib.parse.quote(trial)}&limit=10&expand=supplier").get("rows", [])
                if cands:
                    art = trial; break
        if not cands:
            skipped.append({**it, "art": art, "reason": "нет в МС (архив?)"}); continue
        chosen, ambiguous = SG.pick_in_group(cands, group)
        sid = SG._sup_id(chosen)
        supname = (chosen.get("supplier") or {}).get("name")
        positions.append({"_deliv": False, "qty": it["qty"], "price": it["price"],
                          "product_id": chosen["id"], "line_sum": lsum(it)})
        matched_info.append({"article": art, "name": chosen.get("name"),
                             "supplier_id": sid, "supplier_name": supname,
                             "ambiguous": ambiguous})
        if ambiguous:
            warns.append(f"ℹ {art}: артикул встречается у нескольких товаров — выбран товар группы «{group}»")
    warns += SG.check_positions(group, matched_info)
    return positions, matched_info, skipped, warns


# ═══════════════════════ 7. Округление под сумму счёта ════════════════════════
def adjust_total(positions, target_rub):
    if target_rub is None:
        return
    target = round(target_rub * 100)
    cur = sum(round(p["price"] * 100) * p["qty"] for p in positions)
    # правим последнюю ТОВАРНУЮ позицию
    prod_idx = [i for i, p in enumerate(positions) if not p["_deliv"]]
    if not prod_idx:
        return
    li = prod_idx[-1]
    last = positions[li]
    others = cur - round(last["price"] * 100) * last["qty"]
    new_last_kop = round((target - others) / last["qty"])
    positions[li] = {**last, "_price_kop": new_last_kop}


def price_kop(p):
    return p.get("_price_kop", round(p["price"] * 100))


# ═══════════════════════ 8. Сборка и создание ═════════════════════════════════
def build_payload(hdr, org, store, agent_id, positions, deliv_id, name, warns, skipped):
    now = datetime.now(MSK)
    inv = hdr["inv_date"]
    six = hdr["supplier_inn"] in SIX_INN
    novat = SUPPLIERS.get(hdr["supplier_inn"], {}).get("novat", False)  # поставщик без НДС
    pl = workcal.plan_date(inv, six)
    ms_positions = []
    for p in positions:
        pos = {"quantity": p["qty"], "price": price_kop(p),
               "vat": 0 if novat else 22, "vatEnabled": not novat, "inTransit": p["qty"]}
        if p["_deliv"]:
            pos["assortment"] = meta("service", deliv_id, "service")
        else:
            pos["assortment"] = meta("product", p["product_id"])
        ms_positions.append(pos)
    # ВАЖНО: тревожное (пропуски + предупреждения) — ПЕРВЫМ, чтобы в списке заказов
    # сразу было видно, что в «Комментарии» висит предупреждение.
    alerts = []
    for s in skipped:
        alerts.append(f"НЕ СМАТЧЕНО, вписать вручную: #{s['num']} арт {s.get('art') or '—'} "
                      f"{str(s['name'])[:50]} — {s.get('sum')}₽ ({s['reason']})")
    alerts += warns
    base = f"Автозагрузка из счёта поставщика № {hdr['number']} от {inv:%d.%m.%y}."
    if alerts:
        desc = "⚠ ПРОВЕРЬ:\n" + "\n".join(alerts) + "\n\n" + base
    else:
        desc = base
    payload = {
        "name": name,
        "organization": meta("organization", org["id"]),
        "agent": meta("counterparty", agent_id, "counterparty"),
        "store": meta("store", store["id"]),
        "moment": f"{inv:%Y-%m-%d} {now:%H:%M:%S}",
        "deliveryPlannedMoment": f"{pl:%Y-%m-%d} {now:%H:%M:%S}",
        "applicable": False,
        "vatEnabled": not novat, "vatIncluded": not novat,
        "state": meta("purchaseorder/metadata/states", STATE_CLOSED, "state"),
        "description": desc,
        "positions": ms_positions,
    }
    return payload, pl, six


# ═══════════════════════ PROCESS (вызывается из CLI и из бота) ════════════════
def process(src, create=True, suffix=""):
    """Полный пайплайн. Возвращает dict-результат (никогда не бросает — ошибки в res['error'])."""
    res = {"ok": False, "error": None, "src": src}
    try:
        path = fetch(src)
        kind, grid = read_grid(path)
        text = grid_text(kind, grid)
        hdr = parse_header(text)
        if not hdr["supplier_inn"]:
            res["error"] = "Поставщик не распознан: нет ИНН из реестра в счёте. Добавь профиль поставщика."
            return res
        if not hdr["buyer_inn"]:
            res["error"] = "Покупатель не распознан: нет ИНН Дисквэр/Цифровой квадрат в счёте."
            return res
        prof = SUPPLIERS[hdr["supplier_inn"]]

        if kind == "pdf":
            parser = {"tonerstor": parse_pdf_tonerstor, "ferret": parse_pdf_ferret}.get(prof.get("pdf"))
            if parser is None:
                res["error"] = (f"PDF-счёт от «{prof['name']}» не поддержан (нет PDF-парсера для этого "
                                f"поставщика). Пришлите Excel (xls/xlsx).")
                return res
            items = parser(text)
        else:
            items = parse_table(grid)
        if not items:
            res["error"] = "Не распознано ни одной позиции в счёте."
            return res

        org, store = resolve_buyer(hdr["buyer_inn"])
        agent_id = resolve_agent(hdr["supplier_inn"])
        group = SG.group_of_counterparty(agent_id)
        deliv_id = resolve_delivery_service()

        positions, matched, skipped, warns = match_products(items, prof, group)
        target = round(sum(p["line_sum"] for p in positions), 2) if positions else None
        adjust_total(positions, target)
        all_items_sum = round(sum((it.get("sum") if it.get("sum") is not None else it["qty"] * it["price"]) for it in items), 2)

        name = (hdr["number"] + suffix) if suffix else hdr["number"]
        po, plan_dt, six = build_payload(hdr, org, store, agent_id, positions, deliv_id, name, warns, skipped)
        json.dump(po, open(os.path.join(os.path.dirname(path), "_last_payload.json"), "w"),
                  ensure_ascii=False, indent=1)

        res.update({
            "ok": True, "number": hdr["number"], "name": name,
            "supplier": prof["name"], "supplier_inn": hdr["supplier_inn"], "group": group, "six": six,
            "buyer": org["name"], "store": store["name"],
            "inv_date": f"{hdr['inv_date']:%d.%m.%Y}", "kind": kind,
            "n_items": len(items), "n_matched": len(matched),
            "n_deliv": sum(1 for p in positions if p["_deliv"]), "n_skipped": len(skipped),
            "moment": po["moment"], "plan": f"{plan_dt:%d.%m.%Y}", "plan_wd": WD[plan_dt.weekday()],
            "ordersum": round(sum(price_kop(p) * p["qty"] for p in positions) / 100, 2),
            "target": target, "all_items_sum": all_items_sum, "hdr_total": hdr["total"],
            "hdr_mismatch": (hdr["total"] is not None and abs(hdr["total"] - all_items_sum) > 0.02),
            "skipped": [{"num": s["num"], "art": s.get("art") or "—", "name": s["name"][:60],
                         "sum": s.get("sum"), "reason": s["reason"]} for s in skipped],
            "warns": warns, "created": False, "stop": False, "post_error": None,
        })

        if not create:
            res["dry_run"] = True
            return res

        # ── Коллизия номера заказа (номера счетов у поставщиков сбрасываются по годам) ──
        # Правило для ВСЕХ поставщиков: если «Заказ поставщику» с таким номером уже есть и он
        # ДРУГОГО года — СТАРОМУ дописываем его год (6262 → 6262-2025), а новый заводим под чистым
        # номером (чтобы входящая УПД нашла его по номеру). Тот же год = повтор/дубль → не создаём.
        ex = [o for o in get(f"/entity/purchaseorder?filter=name={urllib.parse.quote(name)}&limit=100")["rows"]
              if str(o.get("name", "")).strip() == name]
        if ex and not suffix:
            new_year = hdr["inv_date"].year
            same_year = [o for o in ex if str(o.get("moment", ""))[:4] == str(new_year)]
            if same_year:
                res["stop"] = True
                res["stop_msg"] = (f"Заказ «{name}» за {new_year} г. уже существует "
                                   f"(id {same_year[0]['id'][:8]}, {same_year[0]['moment'][:10]}). Не создаю — дубль.")
                return res
            renamed = []
            for o in ex:                                    # старые заказы прошлых лет — переименовать с их годом
                oy = str(o.get("moment", ""))[:4] or "----"
                cand = f"{name}-{oy}"
                busy = [b for b in get(f"/entity/purchaseorder?filter=name={urllib.parse.quote(cand)}&limit=1")["rows"]
                        if str(b.get("name", "")).strip() == cand]
                if busy:
                    res["stop"] = True
                    res["stop_msg"] = (f"Заказ «{name}» занят старым (id {o['id'][:8]}, {oy}), "
                                       f"а «{cand}» уже существует. Разведи вручную.")
                    return res
                st_r, _ = put(f"/entity/purchaseorder/{o['id']}", {"name": cand})
                if st_r not in (200, 201):
                    res["stop"] = True
                    res["stop_msg"] = f"Не смог переименовать старый заказ «{name}» → «{cand}»: HTTP {st_r}."
                    return res
                renamed.append(cand)
            res["renamed_old"] = renamed
            res["auto_suffix_note"] = (f"Номер «{name}» был занят заказом(ами) прошлых лет — "
                                       f"переименовал старый в {', '.join(renamed)}, новый завёл под «{name}».")
            ex = []
        if ex:
            res["stop"] = True
            res["stop_msg"] = f"Заказ «{name}» уже существует (id {ex[0]['id'][:8]}, {ex[0]['moment']}). Не создаю."
            return res
        st, resp = post("/entity/purchaseorder", po)
        if st in (200, 201):
            res.update({"created": True, "order_id": resp.get("id"),
                        "order_sum": resp.get("sum", 0) / 100, "order_vatsum": resp.get("vatSum", 0) / 100,
                        "order_pos": resp.get("positions", {}).get("meta", {}).get("size"),
                        "order_url": "https://online.moysklad.ru/app/#purchaseorder/edit?id=" + resp.get("id")})
        else:
            res["post_error"] = f"HTTP {st}: " + json.dumps(resp, ensure_ascii=False)[:400]
        return res
    except SystemExit as e:
        res["error"] = str(e)
        return res
    except Exception as e:
        res["error"] = f"{type(e).__name__}: {e}"
        return res


def format_report(res):
    """Единый текстовый отчёт (для CLI и Telegram)."""
    if not res["ok"]:
        return "❌ Ошибка: " + (res["error"] or "неизвестно")
    L = []
    L.append(f"📄 Счёт № {res['number']} от {res['inv_date']}")
    L.append(f"Поставщик: {res['supplier']} · группа «{res['group']}»" + (" · 6-дн" if res["six"] else ""))
    L.append(f"Покупатель: {res['buyer']} → склад «{res['store']}»")
    L.append(f"Позиций: {res['n_matched']} тов." + (f" + {res['n_deliv']} дост." if res['n_deliv'] else "")
             + (f" · пропущено {res['n_skipped']}" if res['n_skipped'] else ""))
    L.append(f"Сумма: {res['ordersum']:.2f} ₽ · приёмка {res['plan']} ({res['plan_wd']})")
    if res["hdr_mismatch"]:
        L.append(f"⚠ шапка счёта {res['hdr_total']} ≠ сумма строк {res['all_items_sum']}")
    if res["skipped"]:
        L.append("⚠ Пропущены (вписать вручную):")
        for s in res["skipped"]:
            L.append(f"   #{s['num']} {s['art']}: {s['name']} — {s['sum']}₽ ({s['reason']})")
    if res["warns"]:
        L.append("ℹ Предупреждения (в «Комментарий»):")
        for w in res["warns"]:
            L.append("   " + w)
    if res.get("auto_suffix_note"):
        L.append("ℹ " + res["auto_suffix_note"])
    if res.get("dry_run"):
        L.append("\n[DRY-RUN] заказ не создан.")
    elif res.get("stop"):
        L.append("\n⛔ " + res["stop_msg"])
    elif res.get("created"):
        L.append(f"\n✅ Черновик создан: {res['name']} · {res['order_sum']:.2f} ₽ (НДС {res['order_vatsum']:.2f})")
        L.append(res["order_url"])
    elif res.get("post_error"):
        L.append("\n❌ Не создан: " + res["post_error"])
    return "\n".join(L)


def main():
    args = sys.argv[1:]
    if not args:
        raise SystemExit("Usage: invoice_to_po.py <файл|ссылка> [--create] [--suffix S]")
    create = "--create" in args
    suffix = args[args.index("--suffix") + 1] if "--suffix" in args else ""
    res = process(args[0], create=create, suffix=suffix)
    print(format_report(res))
    if create:                                    # ручной CLI-прогон тоже пишем в общий журнал /report
        import proc_log
        proc_log.log_event("invoice", "cli", os.path.basename(args[0]), "cli", res)


if __name__ == "__main__":
    main()
