"""collectors/yandex_services.py — импорт «Отчёта о стоимости услуг маркетплейса» Яндекса.

Ручная выгрузка из ЛК (Финансы → Стоимость услуг), один лист на вид услуги.
Нужна, потому что API продвижения отдаёт лишь ~70 дней вглубь: реклама за янв–апр
(буст продаж/показов, Полки, товарные баннеры), а также баллы за отзыв и подписка
достаются только отсюда. Май–июнь по бусту сходятся с API до рубля (сверено).

Пишем построчно в raw_yandex_services (сырьё), свёртка по месяцам — в collect() витрины
(reports через services_monthly). Идемпотентно: полный снапшот, перезаливаем целиком.

Запуск:  ./venv/bin/python collectors/yandex_services.py [путь.xlsx]
По умолчанию — incoming/marketplace_services_financial_month.xlsx.
"""
import os
import io
import sys
import csv
import time
import zipfile
import hashlib
import calendar
import datetime
import pathlib

import openpyxl
import requests
from psycopg2.extras import Json
from dotenv import load_dotenv

BASE_DIR = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
from core import db  # noqa: E402

load_dotenv(BASE_DIR / ".env")
API = "https://api.partner.market.yandex.ru"
ACCOUNT = "ya_acc1"
DEFAULT_FILE = BASE_DIR / "incoming" / "marketplace_services_financial_month.xlsx"

# Единый отчёт Партнёр-API (united-marketplace-services) отдаёт по CSV на вид услуги.
# ИСТОЧНИК ИСТИНЫ для расходных категорий витрины — сверено с ЛК до копейки (январь 2026):
#   placement → commission (комиссия за размещение, колонка TOTAL_AMOUNT — не SERVICE_PRICE!);
#   delivery + express_delivery → logistics (SERVICE_PRICE);
#   payment_transfer + payment_accepting → acquiring/эквайринг (SERVICE_PRICE);
#   order_processing + storage_of_returns → misc/«прочее» (SERVICE_PRICE);
#   boost → boost_sales; cpm-boost/product-banners → boost_shows; Полки → shelf (реклама раздельно);
#   loyalty_and_reviews → reviews; business_subscription → subscription.
# Раньше расходные CSV выбрасывались (учитывались через stats/orders commissions[]); теперь берём
# ИХ отсюда, а в stats/orders эти категории больше НЕ источаем (иначе задвоение). См. yandex_monthly.
CSV_CAT = {
    "boost.csv": "boost_sales",
    "cpm-boost.csv": "boost_shows",
    "product-banners.csv": "boost_shows",
    "placement.csv": "commission",
    "delivery.csv": "logistics",
    "express_delivery.csv": "logistics",
    "payment_transfer.csv": "acquiring",
    "payment_accepting.csv": "acquiring",
    "order_processing.csv": "misc",
    "storage_of_returns.csv": "misc",
    "loyalty_and_reviews.csv": "reviews",
    "business_subscription.csv": "subscription",
}
DATE_COLS = ("SERVICE_DATE", "SERVICE_DATE_TIME")
BONUS_COLS = ("BONUS_PAID", "PAYMENT_WITH_BONUSES")
# Реальная оплата ₽ лежит в РАЗНЫХ колонках у разных услуг (сверено с ручной выгрузкой ЛК):
#   буст продаж → PREPAID+POSTPAID; буст показов/баннеры → PAYMENT;
#   отзывы/подписка/Полки → SERVICE_PRICE (фолбэк AMOUNT_WITHOUT_BONUSES/TOTAL_AMOUNT).
SERVICE_PRICE_FALLBACK = ("SERVICE_PRICE", "AMOUNT_WITHOUT_BONUSES", "TOTAL_AMOUNT")


def _fnum(v):
    if v in (None, ""):
        return 0.0
    try:
        return float(str(v).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except ValueError:
        return 0.0


def _csv_cat(fname, service_name):
    if "полк" in (service_name or "").lower():   # Полки — где бы ни лежали (shelf.csv и пр.)
        return "shelf"
    return CSV_CAT.get(fname)                     # None → строка пропускается


def _csv_cost(fname, rec):
    """Реальная оплата ₽ строки — колонка зависит от вида услуги (сверено с ЛК)."""
    if fname == "boost.csv":                       # буст продаж
        return _fnum(rec.get("PREPAID")) + _fnum(rec.get("POSTPAID"))
    if fname in ("cpm-boost.csv", "product-banners.csv"):  # буст показов / баннеры
        return _fnum(rec.get("PAYMENT"))
    if fname == "placement.csv":                   # комиссия за размещение — реальные деньги в TOTAL_AMOUNT
        return _fnum(rec.get("TOTAL_AMOUNT"))
    # delivery/express/payment_*/order_processing/storage_of_returns/отзывы/подписка → SERVICE_PRICE
    return next((_fnum(rec[c]) for c in SERVICE_PRICE_FALLBACK if _fnum(rec.get(c))), 0.0)

# Правило разбора на каждый лист услуги:
#   category   — свёрнутая категория (ad|subscription|reviews);
#   cost_cols  — заголовки колонок, чью сумму берём как реальную оплату в ₽;
#   bonus_col  — колонка «оплата бонусами» (не наши деньги), если есть;
#   date_col   — колонка даты оказания услуги.
DATE = "Дата оказания услуги"
DATE_DT = "Дата и время оказания услуги"   # часть листов зовёт дату так
# Расходные листы «Отчёта о стоимости услуг» (ручная выгрузка ЛК за старые месяцы, где
# API недоступен). Категории — те же, что источает API (_csv_cat): commission/logistics/
# acquiring/misc. Комиссия (лист «Размещение») — колонка «Стоимость услуги (AX = …)»:
# суффикс формулы меняется, поэтому матчим по префиксу «Стоимость услуги (AX» (маркер *).
# Проверено на январе до копейки: комиссия 219766.79, логистика 85751.91,
# эквайринг 12351.15, прочее 600.
SHEETS = {
    "Буст продаж, оплата за показы":  ("boost_shows",  ["Оплата, ₽"],                 "Оплата бонусами", DATE),
    "Буст продаж, оплата за продажи": ("boost_sales",  ["Предоплата, ₽", "Постоплата, ₽"], "Оплата бонусами", DATE),
    "Полки":                          ("shelf",        ["Оплата, ₽"],                 "Оплата бонусами", DATE),
    "Товарные баннеры":               ("boost_shows",  ["Оплата, ₽"],                 "Оплата бонусами", DATE),
    "Программа лояльности и отзывы":   ("reviews",      ["Стоимость услуги, ₽"], "Оплата бонусами", DATE),
    "Подписки":                       ("subscription", ["Стоимость услуги, ₽"],       None,              DATE),
    "Размещение товаров и услуг":      ("commission",   ["Стоимость услуги (AX*"],     None,              DATE_DT),
    "Доставка покупателю":            ("logistics",    ["Стоимость услуги, ₽"],       None,              DATE_DT),
    "Экспресс-доставка покупателю":    ("logistics",    ["Стоимость услуги"],          None,              DATE_DT),
    "Приём платежа":                  ("acquiring",    ["Стоимость услуги, ₽"],       None,              DATE_DT),
    "Перевод платежа":                ("acquiring",    ["Стоимость услуги, ₽"],       None,              DATE_DT),
    "Обработка заказов в СЦ или ПВЗ":  ("misc",         ["Стоимость услуги, ₽"],       None,              DATE),
    "Хранение невыкупов и возвратов":  ("misc",         ["Стоимость услуги, ₽"],       None,              DATE),
}
ORDER_HDRS = ("Номер заказа или отгрузки",)
SKU_HDRS = ("Ваш SKU",)


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def _ym(v):
    if hasattr(v, "year"):
        return f"{v.year}-{v.month:02d}"
    if isinstance(v, str) and len(v) >= 7 and v[4] == "-":
        return v[:7]
    return None


def _svc_date(v):
    if hasattr(v, "year"):
        return v.date() if hasattr(v, "date") else v
    if isinstance(v, str) and len(v) >= 10 and v[4] == "-":
        return v[:10]
    return None


def _header_row(ws):
    """Шапка не всегда во 2-й строке (у части листов сверху текст-примечание)."""
    for r in range(1, 8):
        names = [(str(ws.cell(r, c).value).strip() if ws.cell(r, c).value is not None else "")
                 for c in range(1, ws.max_column + 1)]
        if DATE in names or any(n.startswith("Оплата, ₽") for n in names) \
           or "Предоплата, ₽" in names \
           or any(n.startswith("Стоимость услуги") for n in names) \
           or any("оказания услуги" in n for n in names):
            return r, {n: i + 1 for i, n in enumerate(names) if n}
    return None, {}


def parse(path):
    """Возвращает список нормализованных строк услуг (dict)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for sheet, (category, cost_cols, bonus_col, date_col) in SHEETS.items():
        if sheet not in wb.sheetnames:
            print(f"  [services] лист «{sheet}» не найден — пропуск", flush=True)
            continue
        ws = wb[sheet]
        hr, h = _header_row(ws)
        if hr is None:
            print(f"  [services] «{sheet}»: шапка не найдена — пропуск", flush=True)
            continue
        # колонки стоимости: точное имя, либо префикс (маркер '*' в конце — для «AX = …»)
        cc = []
        for x in cost_cols:
            if x.endswith("*"):
                cc += [h[k] for k in h if k.startswith(x[:-1])]
            elif x in h:
                cc.append(h[x])
        cc = list(dict.fromkeys(cc))
        bc = h.get(bonus_col) if bonus_col else None
        # дата оказания услуги на части листов зовётся «Дата и время оказания услуги»
        dc = h.get(date_col) or next((h[k] for k in h if "оказания услуги" in k), None)
        oc = next((h[x] for x in ORDER_HDRS if x in h), None)
        sc = next((h[x] for x in SKU_HDRS if x in h), None)
        n = 0
        for r in range(hr + 1, ws.max_row + 1):
            cost = sum(_num(ws.cell(r, c).value) for c in cc)
            bonus = _num(ws.cell(r, bc).value) if bc else 0.0
            if cost == 0 and bonus == 0:
                continue
            dval = ws.cell(r, dc).value if dc else None
            ym = _ym(dval)
            if not ym:
                continue
            order_id = str(ws.cell(r, oc).value) if oc and ws.cell(r, oc).value is not None else None
            sku = str(ws.cell(r, sc).value) if sc and ws.cell(r, sc).value is not None else None
            key = f"{sheet}|{ym}|{order_id}|{sku}|{_svc_date(dval)}|{cost:.4f}|{bonus:.4f}|{n}"
            rows.append({
                "account": ACCOUNT, "service": sheet, "category": category,
                "ym": ym, "svc_date": _svc_date(dval),
                "order_id": order_id, "sku": sku,
                "cost": round(cost, 2), "bonus": round(bonus, 2), "source": "file",
                "row_hash": hashlib.md5(("file|" + key).encode()).hexdigest(),
                "payload": Json({"service": sheet, "ym": ym, "order_id": order_id,
                                 "sku": sku, "cost": round(cost, 2), "bonus": round(bonus, 2)}),
            })
            n += 1
        print(f"  [services] «{sheet}»: {n} строк, Σ оплата "
              f"{sum(x['cost'] for x in rows if x['service'] == sheet):,.0f} ₽", flush=True)
    return rows


def _replace_months(rows, account):
    """Помесячная идемпотентность: месяцем владеет тот, кто записал последним.
    Чистим только затрагиваемые месяцы (любой источник) и вставляем новые строки —
    так API (свежие месяцы) и файл (старые) не затирают друг друга."""
    if not rows:
        return 0
    months = sorted({r["ym"] for r in rows})
    db.execute("DELETE FROM raw_yandex_services WHERE account=%s AND ym = ANY(%s)",
               (account, months))
    db.upsert("raw_yandex_services", rows, conflict_cols=["account", "row_hash"])
    return len(rows), months


def import_file(path=DEFAULT_FILE, account=ACCOUNT):
    """Ручная выгрузка ЛК → raw_yandex_services (source='file'), помесячно идемпотентно."""
    path = pathlib.Path(path)
    if not path.exists():
        print(f"  [services] файл не найден: {path} — пропуск", flush=True)
        return 0
    rows = parse(path)
    if not rows:
        return 0
    n, months = _replace_months(rows, account)
    print(f"  [services/file] залито {n} строк, месяцы {months[0]}..{months[-1]}", flush=True)
    return n


def _services_report(date_from, date_to, timeout=240):
    """Единый отчёт о стоимости услуг → {имя_csv: [dict-строки]} (асинхронно, ZIP из CSV)."""
    key = os.getenv("YANDEX_API_KEY_ACC1")
    biz = int(os.getenv("YANDEX_BUSINESS_ID_ACC1"))
    H = {"Api-Key": key, "Content-Type": "application/json"}
    body = {"businessId": biz,
            "dateTimeFrom": f"{date_from}T00:00:00+03:00",
            "dateTimeTo": f"{date_to}T23:59:59+03:00"}
    r = requests.post(f"{API}/reports/united-marketplace-services/generate",
                      headers=H, params={"format": "CSV"}, json=body, timeout=60)
    r.raise_for_status()
    rid = r.json()["result"]["reportId"]
    t0 = time.time()
    while time.time() - t0 < timeout:
        i = requests.get(f"{API}/reports/info/{rid}", headers=H, timeout=30).json().get("result", {})
        st = i.get("status")
        if st == "DONE":
            f = requests.get(i["file"], timeout=180)
            out = {}
            with zipfile.ZipFile(io.BytesIO(f.content)) as z:
                for name in z.namelist():
                    if name.endswith(".csv"):
                        out[name] = list(csv.DictReader(
                            io.TextIOWrapper(z.open(name), encoding="utf-8")))
            return out
        if st == "FAILED":
            raise RuntimeError(f"united-services FAILED: {i.get('subStatus')}")
        time.sleep(3)
    raise RuntimeError(f"united-services: таймаут {timeout}с")


def collect_api(months=None, account=ACCOUNT):
    """Автосбор рекламы/подписки/отзывов из единого отчёта Партнёр-API за месяцы `months`
    ('YYYY-MM-01'). По умолчанию — текущий и прошлый месяц (глубина API ~70 дней).
    Пишет в raw_yandex_services (source='api'), помесячно идемпотентно."""
    today = datetime.date.today()
    if not months:
        cur = today.replace(day=1)
        prev = (cur - datetime.timedelta(days=1)).replace(day=1)
        months = [prev.isoformat(), cur.isoformat()]
    d_from = min(months)[:10]
    y, m = int(max(months)[:4]), int(max(months)[5:7])
    d_to = min(f"{max(months)[:7]}-{calendar.monthrange(y, m)[1]:02d}", today.isoformat())
    try:
        csvs = _services_report(d_from, d_to)
    except Exception as e:
        print(f"  [services/api] отчёт не получен: {e}", flush=True)
        return 0
    rows, skipped = [], {}
    for fname, recs in csvs.items():
        for i, rec in enumerate(recs):
            cat = _csv_cat(fname, rec.get("SERVICE_NAME"))
            if cat is None:
                skipped[fname] = skipped.get(fname, 0) + 1
                continue
            cost = _csv_cost(fname, rec)
            if cost == 0:
                continue
            dval = next((rec[c] for c in DATE_COLS if rec.get(c)), None)
            ym = _ym(dval)
            if not ym:
                continue
            bonus = next((_fnum(rec[c]) for c in BONUS_COLS if c in rec), 0.0)
            order_id = rec.get("ORDER_ID") or None
            sku = rec.get("SHOP_SKU") or None
            key = f"api|{fname}|{ym}|{order_id}|{sku}|{_svc_date(dval)}|{cost:.4f}|{i}"
            rows.append({
                "account": account, "service": fname.replace(".csv", ""), "category": cat,
                "ym": ym, "svc_date": _svc_date(dval), "order_id": order_id, "sku": sku,
                "cost": round(cost, 2), "bonus": round(bonus, 2), "source": "api",
                "row_hash": hashlib.md5(key.encode()).hexdigest(),
                "payload": Json({"file": fname, "ym": ym, "order_id": order_id, "sku": sku,
                                 "service_name": rec.get("SERVICE_NAME"),
                                 "cost": round(cost, 2), "bonus": round(bonus, 2)}),
            })
    if not rows:
        print(f"  [services/api] строк нет за {d_from}..{d_to} (глубина API?)", flush=True)
        return 0
    n, got = _replace_months(rows, account)
    from collections import defaultdict
    by = defaultdict(float)
    for r in rows:
        by[(r["ym"], r["category"])] += r["cost"]
    print(f"  [services/api] залито {n} строк, месяцы {got[0]}..{got[-1]}", flush=True)
    for (ym, cat), v in sorted(by.items()):
        print(f"    {ym} {cat}: {v:,.0f} ₽", flush=True)
    if skipped:
        print(f"    (пропущены как уже-в-stats: {dict(skipped)})", flush=True)
    return n


AD_CATS = ("boost_sales", "boost_shows", "shelf")  # компоненты «рекламы» (реклама = их сумма)


def services_monthly(account=ACCOUNT):
    """Свёртка по месяцам: {ym: {категория: ₽, '<кат>_bonus': ₽, 'ad': Σ рекламных}}.
    Категории: commission, logistics, acquiring, misc, boost_sales, boost_shows, shelf,
    subscription, reviews. «ad» — агрегат рекламных (boost_sales+boost_shows+shelf) для совместимости."""
    out = {}
    for r in db.query("""
            SELECT ym, category, sum(cost)::float cost, sum(bonus)::float bonus
            FROM raw_yandex_services WHERE account=%s GROUP BY ym, category""", (account,)):
        d = out.setdefault(r["ym"], {})
        d[r["category"]] = r["cost"]
        d[r["category"] + "_bonus"] = r["bonus"]
    for ym, d in out.items():
        d["ad"] = sum(d.get(c, 0.0) for c in AD_CATS)
        d["ad_bonus"] = sum(d.get(c + "_bonus", 0.0) for c in AD_CATS)
    return out


def main():
    """Ночной шаг: если файл в incoming/ — заливаем его; всегда добираем свежие месяцы из API."""
    import_file()
    collect_api()


if __name__ == "__main__":
    arg = sys.argv[1] if len(sys.argv) > 1 else ""
    if arg == "api":
        collect_api(sys.argv[2:] or None)
    elif arg and arg != "file":
        import_file(arg)          # путь к xlsx
    else:
        main()
    print("\nСвёртка по месяцам (комиссия | логистика | эквайринг | прочее | реклама | подписка | отзывы):")
    for ym, d in sorted(services_monthly().items()):
        print(f"  {ym}: комис {d.get('commission',0):>9,.0f} | логист {d.get('logistics',0):>8,.0f} "
              f"| эквайр {d.get('acquiring',0):>7,.0f} | прочее {d.get('misc',0):>6,.0f} "
              f"| реклама {d.get('ad',0):>9,.0f} (бустП {d.get('boost_sales',0):>7,.0f}/"
              f"бустПок {d.get('boost_shows',0):>6,.0f}/полки {d.get('shelf',0):>6,.0f}) "
              f"| подписка {d.get('subscription',0):>6,.0f} | отзывы {d.get('reviews',0):>6,.0f}")
