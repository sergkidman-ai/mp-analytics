"""collectors/supplier_dims.py — габариты поставщиков → таблица supplier_dims.

Ручные выгрузки прайсов/каталогов (incoming/size.zip → incoming/gab/): у каждого поставщика
свой формат единиц. Приводим объём к литрам, вес к кг, габариты к см. Ключ связки с нашим
каталогом — артикул поставщика (= products.article), сверено 1:1 на образцах.

Единицы у источников:
  Изи      — «Размеры мм» одной ячейкой «Д*Ш*В», вес кг;
  Cactus   — Длина/Высота/Ширина (ИУ) в мм по колонкам, вес брутто кг;
  Сакура   — Длина/Ширина/Высота в МЕТРАХ + объём м3, вес кг (листы Лазерная/Струйная/Матричная);
  Профилайн— только Вес и Объём (м3), без ДхШхВ, без штрихкода.

Запуск:  ./venv/bin/python collectors/supplier_dims.py
"""
import re
import sys
import pathlib
import warnings

import openpyxl
import xlrd

warnings.filterwarnings("ignore")
BASE_DIR = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
from core import db  # noqa: E402

GAB = BASE_DIR / "incoming" / "gab"


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# Разумный диапазон для штучного картриджа/расходника (литры). За его пределами —
# перепутанные единицы в источнике, объёму не верим (null).
VOL_LO, VOL_HI = 0.02, 60.0


def _cm(v):
    """Валидная габаритная величина в см (0 < x <= 150), иначе None."""
    return v if (v and 0 < v <= 150) else None


def _sane_vol(v):
    return round(v, 3) if (v and VOL_LO <= v <= VOL_HI) else None


def _vol_l(l_cm, w_cm, h_cm):
    l, w, h = _cm(l_cm), _cm(w_cm), _cm(h_cm)
    if l and w and h:
        return _sane_vol(l * w * h / 1000.0)   # см³ → литры
    return None


def _rows_xlsx(path, sheet=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in ([wb[sheet]] if sheet else wb.worksheets):
        for row in ws.iter_rows(values_only=True):
            yield row


def parse_izi(path):
    out = []
    for i, r in enumerate(_rows_xlsx(path, "Данные по картриджам")):
        if i == 0:
            continue
        art = r[0]
        if not art:
            continue
        dims = str(r[17]) if len(r) > 17 and r[17] else ""
        m = re.findall(r"\d+(?:[.,]\d+)?", dims)
        l = w = h = None
        if len(m) >= 3:
            l, w, h = [(_num(x) or 0) / 10.0 for x in m[:3]]   # мм → см
        out.append({"supplier": "изи", "article": str(art).strip(),
                    "barcode": str(r[7]).strip() if len(r) > 7 and r[7] else None,
                    "length_cm": _cm(l), "width_cm": _cm(w), "height_cm": _cm(h),
                    "weight_kg": _num(r[9]) if len(r) > 9 else None,
                    "volume_l": _vol_l(l, w, h),
                    "title": str(r[11]) if len(r) > 11 and r[11] else None})
    return out


def parse_cactus(path):
    out = []
    for i, r in enumerate(_rows_xlsx(path, "Лист1")):
        if i == 0:
            continue
        art = r[3] if len(r) > 3 else None       # PartNo (= products.article)
        if not art:
            continue
        l = (_num(r[10]) or 0) / 10.0 if len(r) > 10 and _num(r[10]) else None  # мм → см
        h = (_num(r[11]) or 0) / 10.0 if len(r) > 11 and _num(r[11]) else None
        w = (_num(r[12]) or 0) / 10.0 if len(r) > 12 and _num(r[12]) else None
        vol = _vol_l(l, w, h)
        if vol is None and len(r) > 7 and _num(r[7]):
            vol = _sane_vol(_num(r[7]) * 1000.0)  # «Объём единицы» м³ → л
        out.append({"supplier": "cactus", "article": str(art).strip(),
                    "barcode": str(r[6]).strip() if len(r) > 6 and r[6] else None,
                    "length_cm": _cm(l), "width_cm": _cm(w), "height_cm": _cm(h),
                    "weight_kg": _num(r[8]) if len(r) > 8 else None,
                    "volume_l": vol,
                    "title": str(r[5]) if len(r) > 5 and r[5] else None})
    return out


def parse_sakura(path):
    out = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:                       # Лазерная / Струйная / Матричная
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or len(r) < 30:
                continue
            art = r[4]
            if not art:
                continue
            l = (_num(r[25]) or 0) * 100 if _num(r[25]) else None   # м → см
            w = (_num(r[26]) or 0) * 100 if _num(r[26]) else None
            h = (_num(r[27]) or 0) * 100 if _num(r[27]) else None
            # у Сакуры колонки габаритов местами в перепутанных единицах — объём м³
            # надёжнее, берём его первым, дальше фолбэк на габариты.
            vol = _sane_vol(_num(r[28]) * 1000.0) if _num(r[28]) else None
            if vol is None:
                vol = _vol_l(l, w, h)
            out.append({"supplier": "sakura", "article": str(art).strip(),
                        "barcode": str(r[21]).strip() if r[21] else None,
                        "length_cm": _cm(l), "width_cm": _cm(w), "height_cm": _cm(h),
                        "weight_kg": _num(r[29]),
                        "volume_l": vol,
                        "title": str(r[8]) if r[8] else None})
    return out


def parse_profiline(path):
    out = []
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    guid = re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-", re.I)
    for r in range(sh.nrows):
        c0 = str(sh.cell_value(r, 0))
        art = sh.cell_value(r, 4) if sh.ncols > 4 else None
        if not guid.match(c0) or not art:      # строки-категории пропускаем
            continue
        vol_m3 = _num(sh.cell_value(r, 10)) if sh.ncols > 10 else None
        wt_g = _num(sh.cell_value(r, 9)) if sh.ncols > 9 else None   # вес в граммах
        out.append({"supplier": "profiline", "article": str(art).strip(), "barcode": None,
                    "length_cm": None, "width_cm": None, "height_cm": None,
                    "weight_kg": round(wt_g / 1000.0, 3) if wt_g else None,
                    "volume_l": _sane_vol(vol_m3 * 1000.0) if vol_m3 else None,
                    "title": str(sh.cell_value(r, 6)) if sh.ncols > 6 else None})
    return out


PARSERS = {
    "изи": (parse_izi, "Изи.xlsx"),
    "cactus": (parse_cactus, "Вся расходка Китай РФ Cactus GG PR.xlsx"),
    "sakura": (parse_sakura, "САКУРА АБДУЛ 05.06.xlsx"),
    "profiline": (parse_profiline, "Профилайн.xls"),
}


def build():
    total = 0
    for supplier, (fn, fname) in PARSERS.items():
        path = GAB / fname
        if not path.exists():
            print(f"  [supplier_dims] нет файла {fname} — пропуск", flush=True)
            continue
        rows = fn(path)
        # дедуп по (supplier, article) — оставляем первую непустую запись
        seen, dedup = set(), []
        for r in rows:
            k = (r["supplier"], r["article"])
            if k in seen:
                continue
            seen.add(k)
            r["src_file"] = fname
            dedup.append(r)
        db.execute("DELETE FROM supplier_dims WHERE supplier=%s", (supplier,))
        db.upsert("supplier_dims", dedup, conflict_cols=["supplier", "article"])
        withvol = sum(1 for r in dedup if r["volume_l"])
        total += len(dedup)
        print(f"  [supplier_dims] {supplier}: {len(dedup)} артикулов, с объёмом {withvol}", flush=True)
    print(f"  [supplier_dims] всего {total}", flush=True)
    return total


if __name__ == "__main__":
    build()
