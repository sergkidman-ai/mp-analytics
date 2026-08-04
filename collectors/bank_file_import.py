# поток: inv
"""collectors/bank_file_import.py — ручной импорт выписки из ФАЙЛА в `bank_txn`.

Зачем: у Озон Банка (счета есть у обеих фирм) API нет вообще, выписку можно только
выгрузить руками из личного кабинета. Контур разметки опер. расходов при этом должен
остаться единым: те же таблицы, те же правила по контрагенту, тот же экран «Выписки».
Поэтому файл разбирается здесь в те же нормализованные записи, что отдают
`alfa_statement.normalize` / `sber_statement.normalize`, и уходит в общий
`bank_txn_store.store()`.

Форматы:
  * **1С** (`1CClientBankExchange`, .txt) — основной и самый полный: есть ИНН, счета,
    БИК, номер и дата документа, назначение. Кодировка cp1251 или utf-8 — определяется.
  * **CSV / XLSX** — фолбэк по заголовкам колонок (банки называют их по-разному,
    поэтому распознавание — по ключевым словам). Если обязательную колонку не нашли,
    падаем с перечнем реальных заголовков файла, а не молча импортируем мусор.

Направление (DEBIT/CREDIT) в файле явно не указано: считаем по нашему счёту —
если плательщик мы, это расход. Когда счёт в файле не указан, фолбэк — по ИНН нашей
организации.

Идемпотентность: uuid банк не даёт, поэтому натуральный ключ — хеш реквизитов
(`bank_txn_store._nk`): счёт|дата|сумма|номер|назначение. Повторная загрузка того же
файла (и файла с перекрытием периода) новых строк не создаёт.

CLI:
    ./venv/bin/python collectors/bank_file_import.py <файл> --bank ozon \
        --org 7807355364 [--account 40702810...] [--since 2026-01-01] [--dry]
"""
import argparse
import csv
import io
import hashlib
import pathlib
import re
import sys

HERE = pathlib.Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent))

from collectors import bank_txn_store                       # noqa: E402

SINCE_DEFAULT = "2026-01-01"        # выписку ведём с начала года (решение Сергея 03.08.2026)


# ── общее ────────────────────────────────────────────────────────────────────
def decode(data):
    """Байты файла → текст. Русские банки отдают 1С-файл в cp1251, реже в utf-8."""
    if isinstance(data, str):
        return data
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("cp1251", errors="replace")


def _date(s):
    """'15.01.2026' | '2026-01-15' | '15.01.2026 12:33' → '2026-01-15'. Иначе None."""
    s = str(s or "").strip()
    if not s:
        return None
    m = re.match(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else None


def _amount(s):
    """'1 000,50' | '1000.50' | '-1 000,50' → float. Пусто/мусор → None."""
    if s is None or isinstance(s, (int, float)):
        return float(s) if s not in (None, "") else None
    t = re.sub(r"[^\d,.\-]", "", str(s)).replace(",", ".")
    if t.count(".") > 1:                       # '1.000.50' — точки как разделители тысяч
        head, _, tail = t.rpartition(".")
        t = head.replace(".", "") + "." + tail
    try:
        return float(t) if t not in ("", "-", ".") else None
    except ValueError:
        return None


def _digits(s):
    return re.sub(r"\D", "", str(s or "")) or None


def _row(bank, account, direction, day, amount, purpose, doc_no=None, doc_date=None,
         cp_name=None, cp_inn=None, cp_kpp=None, cp_acc=None, cp_bic=None, raw=None):
    """Нормализованная запись — ключи совпадают с normalize() Альфы и Сбера."""
    return {
        "bank": bank, "account": account, "uuid": None, "transaction_id": None,
        "direction": direction, "amount": amount, "currency": "RUB",
        "operation_date": day, "document_date": doc_date or day,
        "document_number": doc_no, "purpose": purpose,
        "counterparty_name": cp_name, "counterparty_inn": cp_inn, "counterparty_kpp": cp_kpp,
        "counterparty_account": cp_acc, "counterparty_bic": cp_bic,
        "_raw": raw,
    }


# ── формат 1С ────────────────────────────────────────────────────────────────
def is_1c(text):
    return text.lstrip().startswith("1CClientBankExchange")


def parse_1c(text, bank, org_inn, account=None):
    """1С-обмен → (список записей, счёт из шапки).

    Наш счёт берём из `СекцияРасчСчет`/`РасчСчет` шапки; параметр `account` его
    перебивает (нужно, когда в файле счёт не указан)."""
    lines = [ln.strip("\r\n") for ln in text.splitlines()]
    head_acc, docs, cur = None, [], None
    for ln in lines:
        key, _, val = ln.partition("=")
        key, val = key.strip(), val.strip()
        if key.startswith("СекцияДокумент"):
            cur = {"_kind": val}
            continue
        if key == "КонецДокумента":
            if cur:
                docs.append(cur)
            cur = None
            continue
        if cur is not None:
            cur[key] = val
        elif key == "РасчСчет" and val and not head_acc:
            head_acc = val

    our = account or head_acc
    our_d = _digits(our)
    ops = []
    for d in docs:
        payer_acc = d.get("ПлательщикСчет") or d.get("ПлательщикРасчСчет")
        payee_acc = d.get("ПолучательСчет") or d.get("ПолучательРасчСчет")
        if our_d and _digits(payer_acc) == our_d:
            direction = "DEBIT"
        elif our_d and _digits(payee_acc) == our_d:
            direction = "CREDIT"
        else:                       # счёт не совпал ни с чем — решаем по ИНН нашей фирмы
            direction = "DEBIT" if _digits(d.get("ПлательщикИНН")) == org_inn else "CREDIT"
        if direction == "DEBIT":
            cp_name = d.get("Получатель1") or d.get("Получатель")
            cp_inn, cp_kpp = d.get("ПолучательИНН"), d.get("ПолучательКПП")
            cp_acc, cp_bic = payee_acc, d.get("ПолучательБИК")
            day = _date(d.get("ДатаСписано")) or _date(d.get("Дата"))
        else:
            cp_name = d.get("Плательщик1") or d.get("Плательщик")
            cp_inn, cp_kpp = d.get("ПлательщикИНН"), d.get("ПлательщикКПП")
            cp_acc, cp_bic = payer_acc, d.get("ПлательщикБИК")
            day = _date(d.get("ДатаПоступило")) or _date(d.get("Дата"))
        amount = _amount(d.get("Сумма"))
        if not day or amount is None:
            continue
        # «ИНН 7807355364 ООО Ромашка» в поле имени — вычищаем префикс, ИНН уже отдельно
        cp_name = re.sub(r"^ИНН\s*\d{10,12}\s*", "", (cp_name or "").strip()) or None
        ops.append(_row(bank, our, direction, day, amount, d.get("НазначениеПлатежа"),
                        doc_no=d.get("Номер"), doc_date=_date(d.get("Дата")),
                        cp_name=cp_name, cp_inn=_digits(cp_inn), cp_kpp=_digits(cp_kpp),
                        cp_acc=cp_acc, cp_bic=_digits(cp_bic), raw=d))
    return ops, our


# ── формат таблицы (CSV / XLSX) ──────────────────────────────────────────────
# Заголовки у банков разные — ищем по ключевым словам в нижнем регистре.
COLS = {
    "date":     ("дата операции", "дата проводки", "дата документа", "дата"),
    "debit":    ("расход", "списание", "дебет", "сумма по дебету", "сумма списания"),
    "credit":   ("приход", "поступление", "кредит", "сумма по кредиту", "сумма зачисления"),
    "amount":   ("сумма операции", "сумма в валюте счета", "сумма"),
    "purpose":  ("назначение платежа", "назначение", "описание", "комментарий"),
    "cp_name":  ("контрагент", "наименование контрагента", "получатель", "плательщик",
                 "корреспондент"),
    "cp_inn":   ("инн контрагента", "инн"),
    "cp_acc":   ("счет контрагента", "счёт контрагента", "счет получателя", "счёт получателя"),
    "cp_bic":   ("бик",),
    "doc_no":   ("номер документа", "№ документа", "номер", "№"),
}


def _match_header(cells):
    """Строка заголовков → {роль: индекс}. Берём первое совпадение по ключевому слову."""
    got = {}
    low = [str(c or "").strip().lower() for c in cells]
    for role, keys in COLS.items():
        for k in keys:
            for i, c in enumerate(low):
                if c and k in c and i not in got.values():
                    got[role] = i
                    break
            if role in got:
                break
    return got


def _table_rows(data, filename):
    """Файл → список строк-списков (XLSX через openpyxl, CSV через sniffer)."""
    if str(filename).lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return [list(r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
    text = decode(data)
    sample = text[:4000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";" if sample.count(";") > sample.count(",") else ","
    return [r for r in csv.reader(io.StringIO(text), dialect)]


def parse_table(data, filename, bank, org_inn, account=None):
    """CSV/XLSX выписка → записи. Заголовок ищем в первых 15 строках."""
    rows = _table_rows(data, filename)
    head, hdr_i = None, None
    for i, r in enumerate(rows[:15]):
        got = _match_header(r)
        if "date" in got and ("amount" in got or "debit" in got or "credit" in got):
            head, hdr_i = got, i
            break
    if not head:
        seen = "; ".join(str(c) for c in (rows[0] if rows else [])[:12])
        raise ValueError("не нашёл в файле колонки даты и суммы. Заголовки первой строки: "
                         + (seen or "<пусто>"))

    def cell(r, role):
        i = head.get(role)
        return r[i] if i is not None and i < len(r) else None

    ops = []
    for r in rows[hdr_i + 1:]:
        day = _date(cell(r, "date"))
        if not day:
            continue
        deb, cre = _amount(cell(r, "debit")), _amount(cell(r, "credit"))
        if deb or cre:                        # раздельные колонки прихода и расхода
            direction = "DEBIT" if deb else "CREDIT"
            amount = abs(deb or cre)
        else:                                 # одна колонка суммы: минус = расход
            amount = _amount(cell(r, "amount"))
            if amount is None:
                continue
            direction, amount = ("DEBIT" if amount < 0 else "CREDIT"), abs(amount)
        ops.append(_row(bank, account, direction, day, amount, cell(r, "purpose"),
                        doc_no=(str(cell(r, "doc_no")).strip() if cell(r, "doc_no") else None),
                        cp_name=(str(cell(r, "cp_name")).strip() or None
                                 if cell(r, "cp_name") else None),
                        cp_inn=_digits(cell(r, "cp_inn")), cp_acc=cell(r, "cp_acc"),
                        cp_bic=_digits(cell(r, "cp_bic")),
                        raw={"row": [str(c) for c in r]}))
    return ops, account


# ── импорт ───────────────────────────────────────────────────────────────────
def parse(data, filename, bank, org_inn, account=None):
    """Файл любого поддержанного формата → (записи, наш счёт)."""
    if isinstance(data, str):
        data = data.encode("utf-8")
    if str(filename).lower().endswith((".xlsx", ".xlsm")):
        return parse_table(data, filename, bank, org_inn, account)
    text = decode(data)
    if is_1c(text):
        return parse_1c(text, bank, org_inn, account)
    return parse_table(data, filename, bank, org_inn, account)


def import_bytes(data, filename, bank, org_inn, account=None, since=None, dry=False):
    """Разобрать файл и положить в `bank_txn` (+ разметка по правилам).

    → stats из `bank_txn_store.store` плюс parsed / period / account / dry.
    dry=True — только разбор, в БД ничего не пишем (проверка формата перед загрузкой)."""
    ops, acc = parse(data, filename, bank, org_inn, account)
    days = sorted(o["operation_date"] for o in ops if o.get("operation_date"))
    info = {"parsed": len(ops), "period": [days[0], days[-1]] if days else None,
            "account": acc, "file": str(filename), "dry": bool(dry),
            "debit": round(sum(o["amount"] for o in ops if o["direction"] == "DEBIT"), 2),
            "credit": round(sum(o["amount"] for o in ops if o["direction"] == "CREDIT"), 2)}
    if dry or not ops:
        info.update({"seen": len(ops), "stored": 0, "dup": 0, "before_cutoff": 0, "ruled": 0})
        return info
    raws = [o.get("_raw") for o in ops]
    # `_raw` — служебный ключ этого модуля (сырьё строки файла), в bank_txn он уезжает
    # отдельной колонкой через raws; в самой записи он не нужен.
    clean = [{k: v for k, v in o.items() if k != "_raw"} | {"account": acc} for o in ops]
    stats = bank_txn_store.store(clean, bank, org_inn, account=acc,
                                 since=since or SINCE_DEFAULT, raws=raws)
    info.update(stats)
    return info


def import_file(path, bank, org_inn, account=None, since=None, dry=False):
    p = pathlib.Path(path)
    return import_bytes(p.read_bytes(), p.name, bank, org_inn, account, since, dry)


def main(argv=None):
    ap = argparse.ArgumentParser(description="Импорт выписки из файла в bank_txn")
    ap.add_argument("file")
    ap.add_argument("--bank", default="ozon", help="метка банка в БД (ozon, tinkoff, …)")
    ap.add_argument("--org", required=True, help="ИНН нашей организации")
    ap.add_argument("--account", default=None, help="наш счёт, если его нет в файле")
    ap.add_argument("--since", default=SINCE_DEFAULT)
    ap.add_argument("--dry", action="store_true", help="только разбор, без записи в БД")
    a = ap.parse_args(argv)
    r = import_file(a.file, a.bank, a.org, a.account, a.since, a.dry)
    print(f"{r['file']}: разобрано {r['parsed']}"
          + (f", период {r['period'][0]}…{r['period'][1]}" if r["period"] else "")
          + f", расход {r['debit']:.2f} ₽, приход {r['credit']:.2f} ₽")
    print(f"в БД новых {r['stored']}, уже было {r['dup']}, до отсечки {r['before_cutoff']}, "
          f"размечено правилом {r['ruled']}" + (" [DRY]" if r["dry"] else ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
