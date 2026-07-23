# поток: rev
"""collectors/wb_clearance.py — загрузка файла «Распродажа стока ВБ» → wb_clearance.

Файлы приходят через dropbox_bot в /opt/mp-analytics/dropbox/ (по одному на юрлицо):
  «…Распродажа_стока_ВБ_ЦК.xlsx»       → wb_acc1 (Цифровой квадрат)
  «…Распродажа_стока_ВБ_Дисквер.xlsx»  → wb_acc2 (Дисквэр)

Колонки (лист 1): Бренд | Категория | Артикул WB | Артикул продавца | Последний баркод |
  Остатки WB | Остатки продавца | Текущая цена | Новая скидка | Цена со скидкой.
Ключ джойна к остаткам — «Артикул WB» = nm_id.

Идемпотентность: файл = полный срез списка аккаунта → перед загрузкой DELETE по account, затем upsert.

Запуск: ./venv/bin/python collectors/wb_clearance.py [путь.xlsx]
В run_daily — main() берёт самые свежие файлы обоих юрлиц из dropbox.
"""
import os
import sys
import glob
import pathlib

import openpyxl

BASE_DIR = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
from core import db  # noqa: E402

DROPBOX = BASE_DIR / "dropbox"

# заголовок файла → колонка таблицы
COLMAP = {
    "Бренд": "brand",
    "Категория": "category",
    "Артикул WB": "nm_id",
    "Артикул продавца": "vendor_code",
    "Последний баркод": "barcode",
    "Остатки WB": "uploaded_wb_stock",
    "Остатки продавца": "seller_stock",
    "Текущая цена": "orig_price",
    "Новая скидка": "discount_pct",
    "Цена со скидкой": "clearance_price",
}
_NUM = {"uploaded_wb_stock", "seller_stock", "orig_price", "discount_pct", "clearance_price"}


def _account_from_name(fname):
    low = fname.lower()
    if "дисквер" in low or "дисквэр" in low or "dsquare" in low or "диксвер" in low:
        return "wb_acc2"
    if "цк" in low or "цифров" in low:
        return "wb_acc1"
    return None


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def parse(path):
    """xlsx → (account, [row dict, …]). account по имени файла."""
    account = _account_from_name(os.path.basename(path))
    if account is None:
        raise RuntimeError(f"не понял юрлицо по имени файла: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    idx = {h: i for i, h in enumerate(header)}
    missing = [h for h in ("Артикул WB", "Цена со скидкой") if h not in idx]
    if missing:
        wb.close()
        raise RuntimeError(f"нет обязательных колонок {missing}; заголовок: {header}")
    rows = []
    for r in it:
        nm_raw = r[idx["Артикул WB"]]
        if nm_raw is None or str(nm_raw).strip() == "":
            continue
        try:
            nm = int(str(nm_raw).strip())
        except ValueError:
            continue
        rec = {"account": account, "nm_id": nm, "source_file": os.path.basename(path)}
        for head, col in COLMAP.items():
            if col in ("nm_id",) or head not in idx:
                continue
            v = r[idx[head]]
            rec[col] = _num(v) if col in _NUM else (str(v).strip() if v is not None else None)
        rows.append(rec)
    wb.close()
    return account, rows


def load_file(path):
    account, rows = parse(path)
    db.execute("DELETE FROM wb_clearance WHERE account=%s", (account,))
    if rows:
        db.upsert("wb_clearance", rows, conflict_cols=["account", "nm_id"])
    print(f"[wb_clearance] {account}: загружено {len(rows)} позиций из {os.path.basename(path)}", flush=True)
    return len(rows)


def _latest_per_account():
    """Самый свежий файл распродажи для каждого юрлица из dropbox."""
    files = glob.glob(str(DROPBOX / "*аспродажа*ВБ*.xlsx"))
    best = {}
    for f in sorted(files, key=os.path.getmtime):  # свежие перезаписывают
        acc = _account_from_name(os.path.basename(f))
        if acc:
            best[acc] = f
    return best


def main(path=None):
    if path:
        return load_file(path)
    files = _latest_per_account()
    if not files:
        print("[wb_clearance] файлов распродажи в dropbox не найдено", flush=True)
        return 0
    total = 0
    for acc, f in files.items():
        total += load_file(f)
    return total


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else None)
