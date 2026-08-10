# поток: prc
# -*- coding: utf-8 -*-
"""
Разбор прайса поставщика в плоский список строк.

Колонки ищутся по ЗАГОЛОВКУ из профиля, а не по номеру: поставщик может вставить колонку,
и загрузка не должна от этого поехать. Если хоть один заголовок не найден — стоп с внятной
ошибкой, а не тихий разбор половины файла.

Форматы: xlsx (openpyxl) и старый xls (xlrd) — Одиссей и Сакура шлют именно xls.
Шапка бывает многострочной (Одиссей: строки 6-8, «Дил USD» / «Включает НДС» / «Цена»),
поэтому заголовки ищутся в ОКНЕ из profile.header_span подряд идущих строк, а данные
начинаются после последней строки окна.
"""
import io

from .profiles import norm


class PriceFormatError(RuntimeError):
    pass


class Sheet:
    """Единый вид на лист xlsx/xls: индексы с нуля, value(r, c)."""

    def __init__(self, nrows, ncols, getter, close=None):
        self.nrows, self.ncols = nrows, ncols
        self._get, self._close = getter, close

    def value(self, row, col):
        try:
            v = self._get(row, col)
        except IndexError:
            return None
        return None if v == "" else v

    def close(self):
        if self._close:
            self._close()


def open_sheet(content, sheet_name=None):
    """bytes файла -> Sheet. Формат определяем по сигнатуре, а не по имени вложения."""
    if content[:2] == b"PK":                      # zip = xlsx/xlsm
        import openpyxl
        book = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = book[sheet_name] if sheet_name and sheet_name in book.sheetnames else book[book.sheetnames[0]]
        return Sheet(ws.max_row, ws.max_column,
                     lambda r, c: ws.cell(r + 1, c + 1).value, book.close)
    if content[:4] == b"\xd0\xcf\x11\xe0":        # OLE2 = xls
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        names = book.sheet_names()
        ws = book.sheet_by_name(sheet_name) if sheet_name and sheet_name in names else book.sheet_by_index(0)
        return Sheet(ws.nrows, ws.ncols, ws.cell_value, book.release_resources)
    raise PriceFormatError("неизвестный формат файла прайса (ожидали xlsx или xls)")


def _find_header(sheet, wanted, scan_rows, span):
    """(строка данных, {логическое имя -> индекс колонки}).

    Заголовки собираются из окна в span строк: у части поставщиков шапка двух-трёхэтажная.
    """
    targets = {key: norm(title) for key, title in wanted.items()}
    limit = min(scan_rows, sheet.nrows)
    for top in range(limit):
        found = {}
        for row in range(top, min(top + span, sheet.nrows)):
            for col in range(sheet.ncols):
                cell = norm(sheet.value(row, col))
                for key, title in targets.items():
                    if cell == title and key not in found:
                        found[key] = col
        if len(found) == len(targets):
            return top + span, found
    raise PriceFormatError(
        f"шапка не найдена в первых {scan_rows} строках (окно {span}); "
        "ищем колонки: " + ", ".join(sorted(wanted.values()))
    )


def parse(content, profile):
    """bytes прайса -> (список строк, номер строки данных, 1-based).

    Строка: dict(row, article, name, stock_raw, qty, price_raw). qty=None — формулировка
    остатка не разобрана, такую строку загрузчик не грузит и выносит в отчёт.
    Строки без артикула пропускаем молча: это разделители разделов и итоги (Одиссей).
    """
    sheet = open_sheet(content, profile.sheet)
    try:
        wanted = {**profile.columns, **profile.category}
        first_row, cols = _find_header(sheet, wanted, profile.header_scan_rows,
                                       profile.header_span)
        rows = []
        for idx in range(first_row, sheet.nrows):
            article = sheet.value(idx, cols["article"])
            if article is None or str(article).strip() == "":
                continue
            if isinstance(article, float) and article.is_integer():
                article = int(article)
            stock_raw = sheet.value(idx, cols["stock"])
            row = {
                "row": idx + 1,
                "article": str(article).strip(),
                "name": str(sheet.value(idx, cols["name"]) or "").strip(),
                "stock_raw": None if stock_raw is None else str(stock_raw).strip(),
                "qty": profile.qty(stock_raw),
                "price_raw": sheet.value(idx, cols["price"]),
            }
            for key in profile.category:            # категории только для фильтра и отчёта
                row[key] = str(sheet.value(idx, cols[key]) or "").strip()
            rows.append(row)
    finally:
        sheet.close()
    if not rows:
        raise PriceFormatError("в прайсе нет ни одной строки с артикулом")
    return rows, first_row + 1
