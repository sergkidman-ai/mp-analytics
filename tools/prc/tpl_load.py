# поток: prc — заливка карточек из шаблона человека в МС (проба по умолчанию)
import sys, json, argparse
sys.path.insert(0, "/opt/mp-analytics")
from dotenv import load_dotenv; load_dotenv("/opt/mp-analytics/.env")
import openpyxl
from core import ms_api
from prices.ms_import import create_in_ms, wb_name

ap = argparse.ArgumentParser()
ap.add_argument("src")
ap.add_argument("--apply", action="store_true")
ap.add_argument("--only", nargs="*")
ap.add_argument("--drop", nargs="*", default=[], help="артикулы, которые не грузим")
ap.add_argument("--over", default="{}", help='JSON: {"код": {"поле": "значение"}}')
a = ap.parse_args()
OVER = json.loads(a.over)

ws = openpyxl.load_workbook(a.src, data_only=True).active
hdr = [c.value for c in ws[1]]
recs = []
for r in range(2, ws.max_row + 1):
    rec = dict(zip(hdr, [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]))
    if not any(v not in (None, "") for v in rec.values()):
        continue
    for k in ("Код", "Внешний код", "Артикул", "Доп. поле: Код поставщика", "Доп. поле: Связь"):
        if rec.get(k) is not None:
            rec[k] = str(rec[k]).strip()      # числовой артикул из xlsx приезжает int → МС 400/2016
    if rec["Артикул"] in a.drop:
        continue
    rec.update(OVER.get(rec["Код"], {}))
    rec["Доп. поле: Название WB"], _ = wb_name(rec["Доп. поле: Название WB"])
    recs.append(rec)

recs = [r for r in recs if not a.only or r["Код"] in a.only]
for r in recs:
    print(f"  {r['Код']:9} вн.{r['Внешний код']:5} {str(r['Артикул']):24} вес {r['Вес']} | {r['Штрихкод Code128']} | "
          f"связь {r['Доп. поле: Связь']} | {r['Закупочная цена']} ₽ | WB: {r['Доп. поле: Название WB']}")
by_sup = {}
for r in recs:
    by_sup.setdefault(r["Поставщик"], []).append(r)
for sup, group in by_sup.items():
    ids = [x["id"] for x in ms_api.get("/entity/counterparty", {"filter": f"name={sup}", "limit": 5}).get("rows", [])
           if x["name"] == sup]
    assert len(ids) == 1, f"контрагент «{sup}»: найдено {len(ids)}"
    print(f"\n{sup} — {len(group)}")
    create_in_ms(group, ids[0], dry=not a.apply)
