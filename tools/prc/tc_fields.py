# поток: prc
"""tools/prc/tc_fields.py — характеристики из каталога TheCartridge в доп. поля МойСклада.

Пять полей карточки МС заполняем из первоисточника — каталога ТК (`prc_tc_model`, слепок
ночного таймера prc-tc-catalog). Один внешний код = одна модель товара, поэтому у всей родни
одного кода значения обязаны совпадать: берём их из каталога, а не из названий карточек
(названия заводили руками, по-разному и с ошибками).

    Модель                 ← `title` ТК                     («PGI-450PGBK»)
    Доп. название модели   ← `additional_title` ТК          (OEM-код, «6499B001»)
    Цвет                   ← `ink_colors` ТК строчными      («черный», «набор cmyk»)
    Чип                    ← `chip` ТК                      («с чипом» / «с чипом без счётчика» /
                                                             «без чипа» / «не указан» — пусто в ТК
                                                             НЕ оставляем пустым в МС)
    Ресурс (поставщика)    ← число из НАЗВАНИЯ карточки МС  (это название пришло из прайса,
                                                             то есть ресурс так, как его заявил
                                                             поставщик; у ТК он свой и мельче
                                                             варьируется — сюда его НЕ пишем)

Решения Сергея 19.08.2026:
  * цвет и чип ПЕРЕЗАПИСЫВАЕМ по ТК, но если название поставщика прямо противоречит каталогу —
    поле не трогаем вовсе, карточка уходит в список на проверку человеком
    (`prc_tc_fields_{color,chip}_conflicts_<дата>.xlsx`, уходит в бот PRC);
  * карточки, чьего внешнего кода нет в каталоге ТК, не трогаем НИ ПО ОДНОМУ полю;
  * ресурс пишем числом; поставщик не указал — оставляем пусто (не выдумываем и не берём ТК);
  * поле «Название WB» этот инструмент не трогает (у него отдельная задача).

Ловушка формата: «400 к.» в названии — это 400 КОПИЙ, а не 400 тысяч. Общий разбор
`prices/features.resource` читает «к» как «тысяч» и на таких названиях завышает в 1000 раз
(55 живых карточек). Здесь свой разбор: «к.»/«коп»/«стр» — штуки, голое «k»/«к» — тысячи.
Матчер на общем разборе не трогаем — это его отдельный вопрос.

    ./venv/bin/python tools/prc/tc_fields.py                 # отчёт (живой опрос МС)
    ./venv/bin/python tools/prc/tc_fields.py --from-db       # отчёт по слепку raw_moysklad_product
    ./venv/bin/python tools/prc/tc_fields.py --apply [--limit N] [--only 0011,0054] [--fields Чип]

Свод правил по полям, архивации и созданию карточки — `docs/MS_CARD_FIELDS.md`.
"""
import re
import sys
import csv
import json
import time
import pathlib
import argparse
import collections
from datetime import date

BASE_DIR = pathlib.Path(__file__).resolve().parents[2]
sys.path.insert(0, str(BASE_DIR))

from core import db, ms_api                     # noqa: E402
from prices import features as F                # noqa: E402

PAUSE = 0.3          # пауза между пачками записи

# Доп. поля товара МС: имя → (id, тип). Имена полей в МС меняли, переиспользуя id
# («Список принтеров» → «Чип», «Ресурс печати» → «Ресурс (поставщика)»), поэтому ключ здесь
# по id, а не по имени: у 322 карточек в поле «Чип» сейчас лежит старый список принтеров,
# и его надо затереть, а не принять за значение.
ATTRS = {
    "Модель":               ("7d0494e8-70a6-11ed-0a80-065e005bf7dd", "string"),
    "Доп. название модели": ("8e479906-9bfe-11f1-0a80-06a70000d704", "string"),
    "Цвет":                 ("1121b62c-6013-11ed-0a80-0bf6000b8be0", "string"),
    "Чип":                  ("a8832122-6012-11ed-0a80-0225000bddc1", "string"),
    "Ресурс (поставщика)":  ("33207840-6019-11ed-0a80-0ed5000d3b5e", "string"),
}
ATTR_BY_ID = {v[0]: k for k, v in ATTRS.items()}

# Слова для поля берём из общего словаря разбора (prices/features.CHIP_NAMES), чтобы
# карточка и матчер называли одно и то же одинаково.
CHIP_TEXT = F.CHIP_NAMES

# Написания чипа у поставщиков разные, и часть из них общий разбор
# (`prices/features.chip`) читает наоборот. Проверено по каталогу ТК на живых карточках:
#   «б/ч»                       — 66 карточек, разбор молчит, у ТК 58 из них «без чипа»;
#   «необходим/требуется чип»   — 12 карточек, разбор читает «с чипом», у ТК 11 «без чипа»
#                                 (речь о чипе, который надо переставить со старого картриджа);
#   «w/chip» у Blossom          — «без чипа» (сказал Сергей 19.08.2026; в живых названиях
#                                 пока не встретилось, держим на будущее — не путать с «w/o chip»).
# Правим ЛОКАЛЬНО: общий разбор общий с матчингом прайсов, менять его в этой задаче не лезем.
EXTRA_NOCHIP = re.compile(r"\bб/ч\b|(?:необходим|требует\w*)\s+чип|чип\s+отдельн|\bw/chip\b", re.I)


def name_chip(name):
    """Чип по названию поставщика, с поправкой на его написания."""
    if EXTRA_NOCHIP.search(str(name or "")):
        return "nochip"
    return F.chip(name)


# Чип не пишем в двух случаях (решения Сергея 19.08.2026):
#   «противоречие»    — оба источника сказали, и сказали разное (в т.ч. «с чипом» против
#                       «с чипом без счётчика»: тип чипа тоже важен);
#   «в ТК нет данных» — каталог молчит, а поставщик в названии чип указал. Написать сюда
#                       «не указан» значило бы затереть то, что поставщик знает, поэтому
#                       карточка идёт на проверку и пополнение каталога ТК.
def chip_hold(tc_chip, name_chip_code):
    if not name_chip_code:
        return None
    if not tc_chip:
        return "в ТК нет данных"
    return "противоречие" if tc_chip != name_chip_code else None

# Ресурс из названия. Порядок альтернатив важен: «к.» (копий) проверяется РАНЬШЕ голого «к»
# (тысяч), иначе «400 к.» превратится в 400 000.
RES_RE = re.compile(
    r"(?<![A-Za-zА-Яа-я\d\-/])(\d+(?:\s\d{3})*(?:[.,]\d+)?)\s*"
    r"(к\.|коп|стр|pages|pag|k\b|к\b)", re.I)


def supplier_resource(name):
    """Ресурс, как его заявил поставщик в названии карточки. Число страниц или None.

    Берём максимальное из найденного: рядом с ресурсом в названии стоят объём тонера и число
    цветов, ресурс — самое большое число с единицей. Меньше 100 отбрасываем: там уже не ресурс,
    а обрывки кода. Ведущий ноль — признак кода («069K»), не ресурса.
    """
    best = None
    for m in RES_RE.finditer(str(name or "")):
        raw = m.group(1).replace(" ", "").replace(",", ".")
        if re.match(r"0\d", raw):
            continue
        try:
            value = float(raw)
        except ValueError:
            continue
        if m.group(2).lower() in ("k", "к") and value < 1000:   # «11K» = 11 000
            value *= 1000
        if value >= 100 and (best is None or value > best):
            best = value
    return int(best) if best else None


def catalog():
    """Внешний код → характеристики модели ТК (только живые модели каталога)."""
    out = {}
    for r in db.query("""select external_code, title, additional_title, color, chip, raw
                         from prc_tc_model where gone_at is null"""):
        raw = r["raw"]
        if isinstance(raw, str):
            raw = json.loads(raw)
        out[r["external_code"]] = {
            "title": (r["title"] or "").strip().lstrip("№").strip(),
            "add": (r["additional_title"] or "").strip().lstrip("№").strip() or None,
            "color": r["color"],
            "inks": raw.get("ink_colors") or [],
            "chip": r["chip"],
        }
    return out


def cards(from_db, limit=None, only=None):
    """Живые карточки МС: (id, code, name, externalCode, attributes).

    По умолчанию опрашиваем МС вживую: писать мы будем ПОЛНЫЙ список атрибутов карточки,
    и вчерашний слепок затёр бы правки, сделанные человеком за сутки.

    Архивные карточки не трогаем ни по одному полю и не выносим на проверку (решение
    Сергея 19.08.2026): фильтр `archived=false` стоит явно, а не на умолчании МС.
    """
    if from_db:
        rows = db.query("""select payload from raw_moysklad_product
                           where coalesce((payload->>'archived')::bool,false)=false""")
        src = (r["payload"] for r in rows)
    else:
        src = ms_api.iter_rows("/entity/product", params={"filter": "archived=false"}, page=1000)
    n = 0
    for c in src:
        if c.get("archived"):
            continue
        ec = (c.get("externalCode") or "").strip()
        if only and ec not in only:
            continue
        yield c
        n += 1
        if limit and n >= limit:
            return


def current(card):
    """Текущие значения наших пяти полей: имя поля → строка (по id, не по имени в карточке)."""
    out = {}
    for a in card.get("attributes") or []:
        aid = (a.get("meta") or {}).get("href", "").rsplit("/", 1)[-1]
        name = ATTR_BY_ID.get(aid)
        if name:
            v = a.get("value")
            out[name] = "" if v is None else str(v).strip()
    return out


# Карточки, где расхождение цвета человек проверил и признал ложным (цвет ТК верный):
# белый список из docs/prc_color_approved.csv, ключ — код карточки МС.
APPROVED_COLOR_FILE = BASE_DIR / "docs" / "prc_color_approved.csv"


def approved_color():
    if not APPROVED_COLOR_FILE.exists():
        return set()
    with APPROVED_COLOR_FILE.open(encoding="utf-8") as f:
        return {r["код"].strip() for r in csv.DictReader(f, delimiter=";") if r.get("код")}


APPROVED = approved_color()


def plan_card(card, tc):
    """Что записать в карточку. → (новые значения, {поле: (значение ТК, значение из названия)}).

    Спор источников не решаем молча: где каталог ТК противоречит названию поставщика,
    поле не трогаем вовсе и отдаём карточку человеку на проверку (решение Сергея 19.08.2026 —
    сначала выясняем и правим первоисточник, потом заполняем).
    """
    new, conflicts = {}, {}
    name = card.get("name") or ""
    new["Модель"] = tc["title"]
    if tc["add"]:
        new["Доп. название модели"] = tc["add"]

    name_color = F.color(name)
    if (tc["color"] and name_color and name_color != tc["color"]
            and (card.get("code") or "").strip() not in APPROVED):
        conflicts["Цвет"] = (", ".join(tc["inks"]), name_color, "противоречие")
    elif tc["inks"]:
        new["Цвет"] = ", ".join(str(i).strip().lower() for i in tc["inks"] if str(i).strip())

    nchip = name_chip(name)
    hold = chip_hold(tc["chip"], nchip)
    if hold:
        conflicts["Чип"] = (CHIP_TEXT[tc["chip"]], CHIP_TEXT[nchip], hold)
    else:
        new["Чип"] = CHIP_TEXT[tc["chip"]]

    res = supplier_resource(name)
    if res:
        new["Ресурс (поставщика)"] = str(res)
    return new, conflicts


def collect(from_db, fields, limit=None, only=None):
    todo, conflicts, skip = [], [], collections.Counter()
    tc_all = catalog()
    for card in cards(from_db, limit, only):
        ec = (card.get("externalCode") or "").strip()
        tc = tc_all.get(ec)
        if not tc:
            skip["внешнего кода нет в каталоге ТК — не трогаем"] += 1
            continue
        new, conf = plan_card(card, tc)
        cur = current(card)
        for field, (tc_val, name_val, kind) in conf.items():
            conflicts.append({"поле": field, "тип расхождения": kind,
                              "код": card.get("code"), "внешний код": ec,
                              "наименование МС": card.get("name"),
                              "значение ТК": tc_val, "в названии поставщика": name_val,
                              "написание в названии": " | ".join(
                                  f.strip() for f in CHIP_FRAG.findall(card.get("name") or "")
                              ) if field == "Чип" else "",
                              "сейчас в поле МС": cur.get(field, "")})
            skip[f"«{field}»: {kind} — поле не трогаем"] += 1
        new = {k: v for k, v in new.items() if k in fields}
        diff = {k: v for k, v in new.items() if cur.get(k, "") != v}
        if not diff:
            skip["уже верно"] += 1
            continue
        todo.append({"ms_id": card["id"], "code": card.get("code"), "ec": ec,
                     "name": card.get("name"), "cur": cur, "new": diff,
                     "attributes": card.get("attributes") or []})
    return todo, conflicts, skip


def write_report(todo, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["код", "внешний код", "наименование МС", "поле", "было", "станет"])
        for r in sorted(todo, key=lambda x: (x["ec"], x["code"] or "")):
            for f, v in r["new"].items():
                w.writerow([r["code"], r["ec"], r["name"], f, r["cur"].get(f, ""), v])
    return path


CONF_COLS = ["поле", "тип расхождения", "код", "внешний код", "наименование МС", "значение ТК",
             "в названии поставщика", "написание в названии", "сейчас в поле МС"]

# Кусок названия, из которого мы прочли чип: именно его человек и разбирает.
CHIP_FRAG = re.compile(r"[^,;()]{0,25}(?:чип|chip|безлимит|unlimited|б/ч)[^,;()]{0,20}", re.I)


def write_conflicts(rows, path):
    """Список на проверку человеку: CSV рядом и XLSX (его и уносим в бот PRC)."""
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, delimiter=";", fieldnames=CONF_COLS)
        w.writeheader()
        w.writerows(rows)
    return path


def write_xlsx(rows, path, title):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(CONF_COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in CONF_COLS])
    for i, width in enumerate((10, 18, 12, 12, 70, 26, 22, 34, 22), start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return path


def backup(todo, path):
    """Текущие значения наших полей по всем карточкам, которые собираемся тронуть."""
    with open(path, "w", encoding="utf-8") as fh:
        json.dump([{"ms_id": r["ms_id"], "code": r["code"], "cur": r["cur"]} for r in todo],
                  fh, ensure_ascii=False)
    return path


def apply(todo, dry=True, log=print):
    """Запись пачками по 100 (bulk POST /entity/product обновляет карточку по meta).

    Шлём ПОЛНЫЙ список атрибутов карточки со своими значениями: так соседние доп. поля
    не зависят от того, мерджит МС массив атрибутов или заменяет его целиком.
    """
    done = 0
    for start in range(0, len(todo), 100):
        chunk = todo[start:start + 100]
        body = []
        for r in chunk:
            ids = {ATTRS[f][0] for f in r["new"]}
            attrs = [a for a in r["attributes"]
                     if (a.get("meta") or {}).get("href", "").rsplit("/", 1)[-1] not in ids]
            for f, v in r["new"].items():
                aid, atype = ATTRS[f]
                attrs.append({"meta": {"href": f"{ms_api.BASE}/entity/product/metadata/attributes/{aid}",
                                       "type": "attributemetadata",
                                       "mediaType": "application/json"},
                              "type": atype, "value": v})
            body.append({"meta": {"href": f"{ms_api.BASE}/entity/product/{r['ms_id']}",
                                  "type": "product", "mediaType": "application/json"},
                         "attributes": attrs})
        if dry:
            log(f"[проба] пачка {start // 100 + 1}: {len(body)} карточек")
            continue
        ms_api.post("/entity/product", body)
        done += len(body)
        if done % 1000 == 0 or start + 100 >= len(todo):
            log(f"[запись] {done} из {len(todo)}")
        if start + 100 < len(todo):
            time.sleep(PAUSE)
    return done


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="писать в МойСклад (иначе только отчёт)")
    ap.add_argument("--from-db", action="store_true", help="карточки из слепка, не из МС")
    ap.add_argument("--limit", type=int, help="ограничить число карточек")
    ap.add_argument("--only", help="только эти внешние коды через запятую")
    ap.add_argument("--fields", help="только эти поля через запятую (по умолчанию все пять)")
    args = ap.parse_args()

    fields = set(ATTRS)
    if args.fields:
        fields = {f.strip() for f in args.fields.split(",")}
        bad = fields - set(ATTRS)
        if bad:
            raise SystemExit(f"неизвестные поля: {', '.join(sorted(bad))}")
    only = {c.strip() for c in args.only.split(",")} if args.only else None

    todo, conflicts, skip = collect(args.from_db, fields, args.limit, only)

    day = date.today().isoformat()
    rep = BASE_DIR / "docs" / "reports" / f"prc_tc_fields_{day}.csv"
    bak_dir = BASE_DIR / "backups" / f"prc_tc_fields_{day}"
    bak_dir.mkdir(parents=True, exist_ok=True)
    write_report(todo, rep)
    backup(todo, bak_dir / "before.json")

    outs = []
    for field, tag in (("Цвет", "color"), ("Чип", "chip")):
        rows = [r for r in conflicts if r["поле"] == field]
        base = BASE_DIR / "docs" / "reports" / f"prc_tc_fields_{tag}_conflicts_{day}"
        write_conflicts(rows, base.with_suffix(".csv"))
        write_xlsx(rows, base.with_suffix(".xlsx"), f"конфликты {field}")
        outs.append((field, len(rows), base.with_suffix(".xlsx")))

    per_field = collections.Counter(f for r in todo for f in r["new"])
    print(f"карточек к правке: {len(todo)}")
    for f, n in per_field.most_common():
        print(f"  {f:22} {n}")
    for k, n in skip.most_common():
        print(f"  пропуск: {k:38} {n}")
    print(f"отчёт: {rep}")
    for field, n, path in outs:
        print(f"конфликты «{field}» ({n}): {path}")
    print(f"бэкап текущих значений: {bak_dir / 'before.json'}")

    if args.apply:
        done = apply(todo, dry=False)
        print(f"записано карточек: {done}")
    else:
        print("режим отчёта; запись — с --apply")


if __name__ == "__main__":
    main()
