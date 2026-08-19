# поток: prc — старую карточку-дубль в архив с очисткой Артикул/Код/Внешний код.
#
# Когда это дубль (решение Сергея 19.08.2026): артикул поставщика в МС отличается от файла ТОЛЬКО
# разделителями (дефис/пробел/скобка) или опечаткой в цифрах. Отличие в доп. БУКВАХ — это другой
# товар, старую карточку не трогаем. Архивируем ТОЛЬКО при нулевом остатке на КАЖДОМ складе:
# гейт жёсткий, при остатке падаем. Прежние значения — в бэкап-JSON, откат построчно.
import sys, json, argparse
sys.path.insert(0, "/opt/mp-analytics")
from dotenv import load_dotenv; load_dotenv("/opt/mp-analytics/.env")
import openpyxl
from core import ms_api

ap = argparse.ArgumentParser()
ap.add_argument("src", help="*_fixed.xlsx — из него берём ВЕРНЫЙ артикул, чтобы не задеть новую карточку")
ap.add_argument("codes", nargs="+")
ap.add_argument("--backup", required=True)
ap.add_argument("--apply", action="store_true")
a = ap.parse_args()

ws = openpyxl.load_workbook(a.src, data_only=True).active
col = {c.value: i + 1 for i, c in enumerate(ws[1])}
new_art = {str(ws.cell(r, col["Код"]).value).strip(): str(ws.cell(r, col["Артикул"]).value).strip()
           for r in range(2, ws.max_row + 1) if ws.cell(r, col["Код"]).value}

backup, bodies = [], []
for code in a.codes:
    rows = [p for p in ms_api.get("/entity/product", {"filter": f"code={code}", "limit": 20}).get("rows", [])
            if (p.get("article") or "") != new_art.get(code)]
    assert len(rows) == 1, f"{code}: найдено {len(rows)} карточек"
    p = rows[0]
    st = ms_api.get("/report/stock/bystore", {"filter": f"product={p['meta']['href']}", "limit": 50}).get("rows", [])
    left = sum(float(s.get("stock") or 0) for row in st for s in row.get("stockByStore", []))
    assert left == 0, f"{code}: остаток {left} — не архивируем"
    backup.append({"id": p["id"], "code": p.get("code"), "article": p.get("article"),
                   "externalCode": p.get("externalCode"), "archived": p.get("archived"), "name": p.get("name")})
    bodies.append({"meta": p["meta"], "archived": True, "article": "", "code": "", "externalCode": ""})
    print(f"  {code} → архив | арт «{p.get('article')}» вместо «{new_art.get(code)}» | {(p.get('name') or '')[:55]}")

json.dump(backup, open(a.backup, "w"), ensure_ascii=False, indent=1)
if not a.apply:
    print(f"\n[проба] к архивации {len(bodies)}; бэкап {a.backup}")
    sys.exit()
done = ms_api.post("/entity/product", bodies)
print(f"\nархивировано {len(done)} (бэкап {a.backup})")
