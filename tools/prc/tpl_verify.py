# поток: prc — сверка созданных карточек чтением из МС с исходным шаблоном
import sys
sys.path.insert(0, "/opt/mp-analytics")
from dotenv import load_dotenv; load_dotenv("/opt/mp-analytics/.env")
import openpyxl
from core import ms_api
from prices.ms_import import wb_name, ATTRS

src, codes = sys.argv[1], sys.argv[2:]
ws = openpyxl.load_workbook(src, data_only=True).active
hdr = [c.value for c in ws[1]]
bad = 0
for r in range(2, ws.max_row + 1):
    rec = dict(zip(hdr, [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]))
    code = str(rec.get("Код") or "")
    if codes and code not in codes:
        continue
    rows = ms_api.get("/entity/product", {"filter": f"code={code}", "limit": 2, "expand": "productFolder,uom,country"}).get("rows", [])
    if len(rows) != 1:
        print(f"  {code}: карточек в МС {len(rows)}"); bad += 1; continue
    p = rows[0]
    at = {a["name"]: a.get("value") for a in p.get("attributes", [])}
    bc = [b["code128"] for b in p.get("barcodes", []) if "code128" in b]
    want = {
        "externalCode": str(rec["Внешний код"]), "article": str(rec["Артикул"]),
        "name": rec["Наименование"], "weight": float(rec["Вес"]),
        "buyPrice": round(float(rec["Закупочная цена"]) * 100),
        "vat": 22, "uom": "шт", "country": "Китай",
        "code128": [str(rec["Штрихкод Code128"])] if rec.get("Штрихкод Code128") else [],
        "Код поставщика": str(rec["Доп. поле: Код поставщика"]),
        "Название WB": wb_name(rec["Доп. поле: Название WB"])[0],
        "Гарантия/ Срок службы": 365,
    }
    got = {
        "externalCode": p.get("externalCode"), "article": p.get("article"), "name": p.get("name"),
        "weight": p.get("weight"), "buyPrice": p.get("buyPrice", {}).get("value"),
        "vat": p.get("vat"), "uom": p.get("uom", {}).get("name"), "country": p.get("country", {}).get("name"),
        "code128": bc,
        "Код поставщика": at.get("Код поставщика"), "Название WB": at.get("Название WB"),
        "Гарантия/ Срок службы": at.get("Гарантия/ Срок службы"),
    }
    diff = {k: (want[k], got[k]) for k in want if want[k] != got[k]}
    folder = (p.get("productFolder") or {}).get("name")
    if diff:
        bad += len(diff)
        print(f"  {code}: {diff}")
    else:
        print(f"  {code} ок — {folder} — {p['id'][:8]}")
print(f"\nрасхождений: {bad}")
