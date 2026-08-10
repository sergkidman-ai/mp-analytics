# поток: prc — проверка шаблона загрузки карточек в МС (файл человека из дропбокса)
import sys, re, collections
sys.path.insert(0, "/opt/mp-analytics")
from dotenv import load_dotenv; load_dotenv("/opt/mp-analytics/.env")
import openpyxl
from core import ms_api
from prices.ms_import import family, wb_name, query
from prices import features as F
from prices.catalog import compare, FEATURE_NAMES
from prices.novelty import kind

SIGNS = ("model_ok", "kind_ok", "brand_ok", "color_ok", "resource_ok", "chip_ok")


def signs_of(name, article, wb):
    """Признаки для сравнения. Название МС + «Название WB»: чип пишут только во втором."""
    text = f"{name} | {wb or ''}"
    item = {"name": text, "article": str(article or "")}
    item["kind"] = kind(text)
    item.update(F.parse(text, item["article"]))
    return item

SRC = sys.argv[1]
OUT = sys.argv[2]

ws = openpyxl.load_workbook(SRC, data_only=True).active
hdr = [c.value for c in ws[1]]
rows = []
for r in range(2, ws.max_row + 1):
    rec = dict(zip(hdr, [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]))
    if any(v not in (None, "") for v in rec.values()):
        rec["Код"] = str(rec["Код"] or "").strip()
        rec["Внешний код"] = str(rec["Внешний код"] or "").strip()
        rec["Артикул"] = str(rec["Артикул"] or "").strip()
        rows.append(rec)

def ms_by(field, values):
    out = {}
    for v in values:
        if not v:
            continue
        res = ms_api.get("/entity/product", {"filter": f"{field}={v}", "limit": 100})
        out[v] = res.get("rows", [])
    return out

ecs   = {r["Внешний код"] for r in rows}
by_ec   = ms_by("externalCode", ecs)
by_code = ms_by("code", {r["Код"] for r in rows})
by_art  = ms_by("article", {r["Артикул"] for r in rows})
kin     = family(ecs)

last = query("SELECT max(external_code::int) m FROM ms_product WHERE external_code ~ '^[0-9]{4}$'")[0]["m"]
while ms_api.get("/entity/product", {"filter": f"externalCode={last+1}", "limit": 1}).get("rows"):
    last += 1

# суффикс кода → бренд/поставщик живых карточек
sfx = {r["Код"][len(r["Внешний код"]):] for r in rows if r["Код"].startswith(r["Внешний код"])}
sfx_owner = {}
for s in sfx:
    rs = query("""SELECT left(name, 45) n FROM ms_product
                   WHERE NOT archived AND code ~ ('^[0-9]+' || %s || '$') LIMIT 200""", (s,))
    sfx_owner[s] = collections.Counter(re.sub(r"^(Тонер-|Драм-|Фото)?[Кк]артридж\w*\s*", "", x["n"])[:22]
                                       for x in rs).most_common(3)

out = [f"# Проверка шаблона МС — {SRC.split('/')[-1]}\n",
       f"Строк: {len(rows)}. Поставщики: {sorted({str(r['Поставщик']) for r in rows})}. "
       f"Последний занятый внешний код: **{last}**\n"]
problems = []
for r in rows:
    code, ec, art = r["Код"], r["Внешний код"], r["Артикул"]
    tag = code or art
    out.append(f"\n## {code or '(без кода)'} / вн.код {ec or '—'} / {art}\n- {r['Наименование']}")
    live_ec = by_ec.get(ec, [])
    out.append(f"- Внешний код в МС: {len(live_ec)} карточек: " +
               ", ".join(f"{c.get('code')}({c.get('article')})" for c in live_ec[:14]))
    if not ec:
        problems.append(f"{tag}: внешнего кода нет в файле (последний занятый {last})")
    elif not live_ec:
        problems.append(f"{tag}: внешний код {ec} в МС свободен — это НОВЫЙ код (правило 25: вверх от {last})")
    if by_code.get(code):
        problems.append(f"{tag}: код карточки занят — «{by_code[code][0].get('name')[:60]}»")
    if by_art.get(art):
        problems.append(f"{tag}: артикул {art} уже на карточке(ах) " +
                        ", ".join(f"{c.get('code')}/вн.{c.get('externalCode')}" for c in by_art[art]))
    fam = kin.get(ec, [])
    paths = sorted({c["path"] for c in fam if c["path"]})
    b128  = sorted({c["code128"] for c in fam if c["code128"]})
    fresh = max(fam, key=lambda c: c["created"]) if fam else None
    out.append(f"- Родня: {len(fam)}; группы {paths}; Code128 {b128}; "
               f"свежайшая {fresh['code'] if fresh else '—'} "
               f"({fresh['created'].strftime('%d.%m.%Y') if fresh else '—'}) вес {fresh['weight'] if fresh else '—'}")
    out.append(f"- WB родни: {sorted({c['wb'] for c in fam if c['wb']})[:4]}")
    if paths and r["Группы"] not in paths:
        problems.append(f"{tag}: группа «{r['Группы']}» не у родни {paths}")
    if b128 and str(r["Штрихкод Code128"] or "") not in b128:
        problems.append(f"{tag}: Code128 «{r['Штрихкод Code128']}» ≠ Code128 родни {b128}")
    if fresh and fresh["weight"] and float(r["Вес"] or 0) != float(fresh["weight"]):
        problems.append(f"{tag}: вес {r['Вес']} ≠ вес свежайшей родни {fresh['code']} = {fresh['weight']}")
    for k, v in {"Единица измерения": "шт", "НДС": 22, "Страна": "Китай",
                 "Доп. поле: Гарантия/ Срок службы": 365}.items():
        if str(r.get(k)).strip() != str(v):
            problems.append(f"{tag}: {k} = «{r.get(k)}», ожидалось «{v}»")
    if str(r.get("Доп. поле: Код поставщика") or "").strip() != art:
        problems.append(f"{tag}: «Код поставщика» = «{r.get('Доп. поле: Код поставщика')}» ≠ артикул {art}")
    fixed, note = wb_name(r["Доп. поле: Название WB"])
    if fixed != str(r["Доп. поле: Название WB"] or "").strip():
        problems.append(f"{tag}: «Название WB» → «{fixed}» (в файле «{r['Доп. поле: Название WB']}»)")
    if note:
        problems.append(f"{tag}: {note}")
    if not str(r["Описание"] or "").strip():
        problems.append(f"{tag}: пустое описание")

    # шесть признаков (правило 30) — наша строка против КАЖДОЙ карточки родни
    mine = signs_of(r["Наименование"], art, r.get("Доп. поле: Название WB"))
    tally = {s: {True: 0, False: 0, None: 0} for s in SIGNS}
    clash = []
    for c in live_ec:
        wb = {a["name"]: a.get("value") for a in c.get("attributes", [])}.get("Название WB")
        flags = compare(mine, signs_of(c.get("name") or "", c.get("article"), wb))
        for s in SIGNS:
            tally[s][flags[s]] += 1
        bad = [FEATURE_NAMES[s] for s in SIGNS if flags[s] is False]
        if bad:
            clash.append(f"{c.get('code')} — {', '.join(bad)} — {(c.get('name') or '')[:70]}")
    if live_ec:
        out.append("- 6 признаков (да/нет/молчит по родне): " +
                   " · ".join(f"{FEATURE_NAMES[s]} {tally[s][True]}/{tally[s][False]}/{tally[s][None]}" for s in SIGNS))
        for c in clash:
            out.append(f"  - расходится: {c}")
        for s in SIGNS:
            # признак противоречит ВСЕЙ родне — это не разнобой названий, это другой товар
            if tally[s][False] and not tally[s][True]:
                problems.append(f"{tag}: {FEATURE_NAMES[s]} противоречит ВСЕЙ родне ({tally[s][False]} карточек)")
            elif tally[s][False]:
                problems.append(f"{tag}: {FEATURE_NAMES[s]} расходится с {tally[s][False]} из {len(live_ec)} родни — глазами")
        # чип: обоюдное молчание = совпадение (решение 10.08). Просим заполнить только там,
        # где родня про чип ГОВОРИТ, а в шаблоне пусто — это и есть «вводим чип в шаблон».
        if mine["chip"] is None and tally["chip_ok"][None]:
            said = sorted({F.CHIP_NAMES[c] for c in
                           (signs_of(x.get("name") or "", x.get("article"),
                                     {a["name"]: a.get("value") for a in x.get("attributes", [])}.get("Название WB"))["chip"]
                            for x in live_ec) if c})
            problems.append(f"{tag}: чип в шаблоне не указан, у родни — {', '.join(said)}; заполнить")

out.append("\n## Суффиксы кода — кто их носит в МС\n")
for s, top in sfx_owner.items():
    out.append(f"- `{s}` → " + (", ".join(f"«{n}» ×{c}" for n, c in top) or "НЕТ таких карточек"))
out.append("\n## Все поля строк\n")
for r in rows:
    out.append(f"\n### {r['Код'] or '(без кода)'} — {r['Наименование']}")
    out += [f"- **{k}**: {r[k]}" for k in hdr]
out.append("\n## Замечания\n")
out += [f"- {p}" for p in problems] or ["- нет"]
open(OUT, "w", encoding="utf-8").write("\n".join(out))
print(f"строк {len(rows)}, замечаний {len(problems)}, последний занятый вн.код {last}")
print("файл:", OUT)
for p in problems:
    print(" *", p)
