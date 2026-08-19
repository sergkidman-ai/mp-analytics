# поток: prc — из шаблона человека собрать `*_fixed.xlsx`: поля по НАШИМ правилам.
#
# Присланный файл не трогаем (правило 35) — правки идут в копию. Что чиним: код (внешний код +
# аббревиатура бренда из наименования, если бренд распознан), группа и Code128 — как у родни по
# внешнему коду, константы карточки (НДС 22, шт, Китай, гарантия 365), «Название WB» по формуле,
# пустой вес добираем у родни, «Связь» — если она проставлена у родни.
# НАИМЕНОВАНИЕ ПОСТАВЩИКА НЕ ПРАВИМ (решение Сергея 19.08.2026), даже когда в нём ошибочный ресурс:
# это его товар и его название, наши правила касаются полей карточки.
import sys, argparse, collections, json
sys.path.insert(0, "/opt/mp-analytics")
from dotenv import load_dotenv; load_dotenv("/opt/mp-analytics/.env")
import openpyxl
from core import ms_api
from prices.ms_import import abbr_by_name, wb_name

ap = argparse.ArgumentParser()
ap.add_argument("src")
ap.add_argument("--new-ec", default="", help="внешний код для строки, где он пуст")
ap.add_argument("--set", default="{}", help='JSON: {"код": {"колонка": "значение"}} — поверх всего')
a = ap.parse_args()
SET = json.loads(a.set)
dst = a.src.replace(".xlsx", "_fixed.xlsx")

wb = openpyxl.load_workbook(a.src)
ws = wb.active
col = {name: i + 1 for i, name in enumerate(c.value for c in ws[1])}
kin_cache, folder_cache, log = {}, {}, []

def kin(ec):
    if ec not in kin_cache:
        kin_cache[ec] = ms_api.get("/entity/product", {"filter": f"externalCode={ec}", "limit": 50}).get("rows", [])
    return kin_cache[ec]

def folder_path(fid):
    if fid not in folder_cache:
        f = ms_api.get(f"/entity/productfolder/{fid}")
        folder_cache[fid] = (f.get("pathName") + "/" if f.get("pathName") else "") + f.get("name")
    return folder_cache[fid]

def put(r, name, val, ch, why):
    if ws.cell(r, col[name]).value != val:
        ch.append(f"{why}: «{ws.cell(r, col[name]).value}» → «{val}»")
        ws.cell(r, col[name]).value = val

for r in range(2, ws.max_row + 1):
    art = ws.cell(r, col["Артикул"]).value
    if not art:
        continue
    ch, name = [], ws.cell(r, col["Наименование"]).value or ""
    ec = str(ws.cell(r, col["Внешний код"]).value or "").strip() or a.new_ec
    assert ec, f"строка {r}: внешний код пуст, задайте --new-ec"
    put(r, "Внешний код", ec, ch, "внешний код")
    family = kin(ec)

    ab = abbr_by_name(name)
    if ab:
        put(r, "Код", f"{ec}{ab}", ch, "код")
    else:
        ch.append(f"код «{ws.cell(r, col['Код']).value}» оставлен: бренда в наименовании нет")

    paths = collections.Counter(fid for p in family if (fid := ms_api.meta_id(p, "productFolder")))
    if paths:
        put(r, "Группы", folder_path(paths.most_common(1)[0][0]), ch, "группа")
    for cname, val in (("Единица измерения", "шт"), ("Страна", "Китай"), ("НДС", 22),
                       ("Доп. поле: Гарантия/ Срок службы", 365)):
        put(r, cname, val, ch, cname)

    nm, note = wb_name(ws.cell(r, col["Доп. поле: Название WB"]).value)
    put(r, "Доп. поле: Название WB", nm, ch, "название WB")
    if note:
        ch.append(f"название WB — замечание: {note}")

    bc = sorted({b["code128"] for p in family for b in p.get("barcodes") or [] if b.get("code128")})
    put(r, "Штрихкод Code128", bc[0] if bc else f"DSQSR000{ec}", ch,
        "code128" + ("" if bc else " (родни нет, маска семейства)"))

    if not ws.cell(r, col["Вес"]).value:
        w = [p.get("weight") for p in family if p.get("weight")]
        if w:
            put(r, "Вес", max(set(w), key=w.count), ch, "вес пуст, взят у родни")
        else:
            ch.append("вес пуст, у родни тоже нет")

    links = sorted({str(x["value"]).strip() for p in family for x in p.get("attributes") or []
                    if x.get("name") == "Связь" and x.get("value")})
    if links and not ws.cell(r, col["Доп. поле: Связь"]).value:
        put(r, "Доп. поле: Связь", links[0], ch, "связь у родни" + (f" (варианты {links})" if len(links) > 1 else ""))

    ws.cell(r, col["Артикул"]).value = str(art).strip()
    ws.cell(r, col["Доп. поле: Код поставщика"]).value = str(art).strip()
    for cname, val in SET.get(str(ws.cell(r, col["Код"]).value), {}).items():
        put(r, cname, val, ch, f"вручную ({cname})")
    log.append(f"{ws.cell(r, col['Код']).value:9} | " + ("; ".join(ch) if ch else "без правок"))

wb.save(dst)
print("\n".join(log))
print(f"\nстрок {len(log)}, с правками {sum(1 for l in log if 'без правок' not in l)} → {dst}")
