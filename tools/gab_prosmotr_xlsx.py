#!/usr/bin/env python3
# поток: gab
"""docs/PROSMOTR_gabarity.xlsx — файл для личного просмотра Сергея.

Листы: «Раздутые», «Занижённые», «ТОП-50» (по 25 самых дорогих из каждой группы),
«Без источника». Read-only по данным, на площадки ничего не шлёт.

ЖЁСТКОЕ ПРАВИЛО ФАЙЛА: размера без источника здесь быть не может. Источник — это конкретная
строка прайса конкретного поставщика (`supplier_dims`: поставщик · артикул · название · файл).
Строка, для которой строку прайса восстановить нельзя, уходит на лист «Без источника».

Восстановление источника: FINAL_gabarity хранит только имя поставщика («solutionsprint»,
«cactus(согл.2)»), поэтому строку прайса ищем обратным матчингом — в прайсе этого поставщика
берём строки с ТЕМИ ЖЕ сторонами короба и подтверждаем OEM-кодом из названия карточки.
Если сторон-близнецов в прайсе несколько и ни одна не подтверждается кодом — источник считается
невосстановленным (лучше пусто, чем наугад).
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict

sys.path.insert(0, "/opt/mp-analytics")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core import db

SRC = "/opt/mp-analytics/docs/FINAL_gabarity.xlsx"
OUT = "/opt/mp-analytics/docs/PROSMOTR_gabarity.xlsx"

# Ставка логистики ВБ: медиана 22.66 ₽ за литр за штуку по 700 строкам docs/dims_fixlist.csv
# (overpay_rub_mo / (Δобъём × qty)). Та же константа, что в tools/gab_waves.py.
TARIFF_RUB_PER_L = 22.66
SALES_FROM, SALES_TO = "2025-08-01", "2026-07-31"

# Имена в колонке «источник», за которыми НЕТ строки прайса поставщика.
NO_PRICE_ROW = {
    "набор": "короб набора рассчитан по составу, строки прайса на сам набор не существует",
    "веб-оригинал": "размер взят из веб-источника, а не из прайса поставщика",
    "наш-озон": "размер перенесён с нашей же карточки Ozon, прайса поставщика нет",
}

# Единичная упаковка картриджа физически не бывает больше ~25 л. Больше — строка группового
# короба поставщика (10-20 шт). Такой размер в файл для просмотра пускать нельзя: он даёт
# фантомные сотни тысяч рублей «эффекта» и садится в самый верх сортировки.
MAX_SINGLE_VOL_L = 25.0

# Прайс поставщика содержит и технику. Короб принтера под карточкой картриджа (и наоборот) —
# ошибка вида товара, а не источник размера.
HW_RX = re.compile(r"принтер|мфу|плоттер|сканер|копир|selphy|печат\w*\s+машин", re.I)

FAM = [
    r"\d{3}R\d{4,5}", r"MLT-?D\d{3}[A-Z]?", r"CLT-?[KCMY]\d{3}[A-Z]?", r"C-?EXV\d{1,3}",
    r"[CG]PR-?\d{1,3}", r"CRG-?\d{2,3}[A-Z]{0,2}", r"TN-?\d{3,4}[A-Z]{0,2}",
    r"DR-?\d{3,4}[A-Z]{0,2}", r"TK-?\d{3,4}[A-Z]?", r"DK-?\d{3,4}", r"C[EFB]\d{3}[AXYUD]?",
    r"W[12]\d{3}[AX]?", r"C[CN]\d{3}[A-Z]?", r"Q\d{4}[AX]?",
    r"C13[A-Z]\d{2}[A-Z0-9]\d{2}[A-Z]?", r"106R\d{5}|108R\d{5}|013R\d{5}|101R\d{5}",
]
RX = re.compile(r"(?<![A-Z0-9])(?:%s)(?![A-Z0-9])" % "|".join(FAM), re.I)
CONS_RX = re.compile(r"\s*\(согл\.\d+\)\s*$")

COLS = [
    ("артикул продавца", 18), ("nmID", 12), ("название", 52), ("тип", 8),
    ("размер сейчас на ВБ, Д×Ш×В", 20), ("наш размер, Д×Ш×В", 18),
    ("объём сейчас, л", 13), ("объём наш, л", 12), ("во сколько раз расходится", 14),
    ("ИСТОЧНИК: поставщик · строка прайса", 72),
    ("продажи 12 мес, шт", 13), ("продажи 12 мес, ₽", 15),
    ("эффект на логистику, ₽/год", 16),
]


def nrm(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def codes(text):
    return {nrm(m) for m in RX.findall(str(text or "").upper()) if len(nrm(m)) >= 5}


def fnum(x):
    try:
        v = float(str(x).replace(",", "."))
        return v if v > 0 else None
    except Exception:
        return None


def key3(l, w, h, nd=1):
    a = [fnum(l), fnum(w), fnum(h)]
    if not all(a):
        return None
    return tuple(round(v, nd) for v in sorted(a, reverse=True))


def fmt3(l, w, h):
    a = [fnum(l), fnum(w), fnum(h)]
    return "×".join(f"{v:g}" for v in a) if all(a) else "—"


# ---------------------------------------------------------------- входные данные
def load_cards():
    out = {}
    for r in db.query("select nm_id, vendor_code, title, account, length_cm, width_cm, "
                      "height_cm from wb_cards"):
        out[str(r["nm_id"])] = r
    return out


def load_sales():
    out = defaultdict(lambda: {"qty": 0.0, "rev": 0.0})
    for r in db.query(
            "select article, sum(qty) qty, sum(revenue_buyer) rev from sales "
            "where platform='wb' and period_from >= %s and period_to <= %s group by article",
            (SALES_FROM, SALES_TO)):
        d = out[str(r["article"])]
        d["qty"] += float(r["qty"] or 0)
        d["rev"] += float(r["rev"] or 0)
    return out


def load_price_index():
    """supplier -> округлённые стороны -> список строк прайса."""
    idx = defaultdict(lambda: defaultdict(list))
    n = 0
    for r in db.query("select supplier, article, title, src_file, length_cm, width_cm, "
                      "height_cm from supplier_dims where length_cm is not null "
                      "and width_cm is not null and height_cm is not null"):
        k = key3(r["length_cm"], r["width_cm"], r["height_cm"])
        if not k:
            continue
        idx[r["supplier"]][k].append(r)
        n += 1
    return idx, n


# ---------------------------------------------------------------- источник строки
INHERIT_RX = re.compile(r"наследование от\s+(\S+)", re.I)


def unwrap_inherit(src_raw, parent_src, depth=2):
    """«наследование от 0200 (поставщик)» → источник самого родителя 0200.

    Короб ребёнка физически тот же, что у родителя-картриджа, поэтому строка прайса родителя
    и есть источник размера ребёнка. Разворачиваем не глубже двух шагов, чтобы не зациклиться.
    """
    chain = []
    cur = str(src_raw or "")
    for _ in range(depth):
        m = INHERIT_RX.search(cur)
        if not m:
            break
        parent = m.group(1).strip("()")
        nxt = parent_src.get(parent)
        if not nxt or nxt == cur:
            return cur, chain
        chain.append(parent)
        cur = str(nxt)
    return cur, chain


def resolve_source(src_raw, our_key, card_title, card_vc, idx):
    """(строка источника, причина отказа). Ровно одно из двух не None."""
    base = CONS_RX.sub("", str(src_raw or "")).split(":")[0].strip().lower()
    if not base:
        return None, "в FINAL_gabarity источник не заполнен"
    if base in NO_PRICE_ROW:
        return None, NO_PRICE_ROW[base]
    if our_key is None:
        return None, "наш размер в FINAL_gabarity пуст или битый"

    suppliers = [base] if base in idx else (list(idx) if base == "поставщик" else [])
    if not suppliers:
        return None, f"поставщик «{base}» отсутствует в supplier_dims"

    vol_l = our_key[0] * our_key[1] * our_key[2] / 1000
    if vol_l > MAX_SINGLE_VOL_L:
        return None, (f"строка прайса — мастер-короб {vol_l:.0f} л: для единичного товара "
                      f"физически невозможна, размер бракуется")

    cand = []
    for sup in suppliers:
        cand += idx[sup].get(our_key, [])
    if not cand:
        return None, f"в прайсе «{base}» нет строки с коробом {fmt3(*our_key)}"

    # Отсекаем технику под расходником и расходник под техникой — это разный вид товара.
    card_hw = bool(HW_RX.search(str(card_title or "")))
    cand = [c for c in cand if bool(HW_RX.search(str(c["title"] or ""))) == card_hw]
    if not cand:
        return None, (f"строки прайса «{base}» с коробом {fmt3(*our_key)} — другой вид товара "
                      f"({'расходник под карточкой техники' if card_hw else 'техника под карточкой расходника'})")

    want = codes(card_title)
    hit = [c for c in cand
           if want & (codes(c["title"]) | codes(c["article"]))] if want else []
    conf = "OEM-код совпал"
    if not hit:
        # Код вне OEM-белого списка (Epson, Kyocera и пр.): подтверждаем тем, что артикул
        # поставщика буквально присутствует в названии нашей карточки или равен нашему артикулу.
        tn, vn = nrm(card_title), nrm(card_vc)
        hit = [c for c in cand
               if len(nrm(c["article"])) >= 5 and (nrm(c["article"]) in tn
                                                   or nrm(c["article"]) == vn)]
        if hit:
            conf = "артикул поставщика найден в названии карточки"

    if hit:
        pick = hit[0]
        extra = len(hit) - 1
    elif len(cand) == 1:
        pick, conf = cand[0], "единственная строка с таким коробом"
        extra = 0
    else:
        return None, (f"в прайсе «{base}» {len(cand)} строк с коробом {fmt3(*our_key)}, "
                      f"OEM-код карточки ни одну не подтверждает")

    tail = f" (+ ещё {extra} строк с тем же кодом)" if extra else ""
    return (f"{pick['supplier']} · арт. {pick['article']} · "
            f"{str(pick['title'] or '')[:70]} · прайс: {pick['src_file'] or '—'} "
            f"[{conf}]{tail}"), None


# ---------------------------------------------------------------- сборка строк
def build(group, ws_rows, cards, sales, idx, parent_src):
    ok, bad = [], []
    hdr = ws_rows[0]
    hi = {h: i for i, h in enumerate(hdr)}
    for r in ws_rows[1:]:
        nm = str(r[hi["nmID"]]) if r[hi["nmID"]] is not None else ""
        card = cards.get(nm, {})
        vc = r[hi["артикул"]] or card.get("vendor_code") or ""
        title = (r[hi["название"]] if "название" in hi else None) or card.get("title") or ""
        typ = {"ребёнок": "дочь"}.get(str(r[hi["тип"]]), str(r[hi["тип"]]))

        our = key3(r[hi["Д"]], r[hi["Ш"]], r[hi["В"]])
        cur = key3(card.get("length_cm"), card.get("width_cm"), card.get("height_cm"))
        v_our = (our[0] * our[1] * our[2] / 1000) if our else None
        v_cur = (cur[0] * cur[1] * cur[2] / 1000) if cur else None

        s = sales.get(nm, {"qty": 0.0, "rev": 0.0})
        eff = (abs(v_cur - v_our) * s["qty"] * TARIFF_RUB_PER_L
               if (v_cur and v_our) else 0.0)
        ratio = (max(v_cur, v_our) / min(v_cur, v_our)) if (v_cur and v_our) else None

        rec = [vc, nm, title, typ,
               fmt3(*cur) if cur else "—", fmt3(*our) if our else "—",
               round(v_cur, 2) if v_cur else None, round(v_our, 2) if v_our else None,
               round(ratio, 2) if ratio else None, None,
               round(s["qty"], 0), round(s["rev"], 0), round(eff, 0)]

        eff_src, chain = unwrap_inherit(r[hi["источник"]], parent_src)
        # У унаследованной строки код в названии — дочерний («для принтера …»), а строку прайса
        # подбирал родитель-картридж, поэтому сверяем по названию и артикулу РОДИТЕЛЯ.
        m_title, m_vc = (parent_src.get("title:" + chain[-1]) or title, chain[-1]) if chain \
            else (title, vc)
        src, why = resolve_source(eff_src, our, m_title, m_vc, idx)
        if src:
            rec[9] = (f"унаследовано от родителя {chain[-1]} → " if chain else "") + src
            ok.append(rec)
        else:
            bad.append(rec[:9] + [f"источник в FINAL: {r[hi['источник']]}"] + rec[10:] + [why, group])
    ok.sort(key=lambda x: -(x[12] or 0))
    bad.sort(key=lambda x: -(x[12] or 0))
    return ok, bad


# ---------------------------------------------------------------- запись
HEAD_FILL = PatternFill("solid", fgColor="1F3B52")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)


def sheet(wb, name, cols, rows, money_cols, extra_note=None):
    ws = wb.create_sheet(name)
    ws.append([c[0] for c in cols])
    for r in rows:
        ws.append(r)
    for i, (label, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for c in ws[1]:
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
    for ci in money_cols:
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            row[0].number_format = "# ##0"
    if extra_note:
        ws.oddFooter.left.text = extra_note
    return ws


def main():
    wbin = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    cards, sales = load_cards(), load_sales()
    idx, npr = load_price_index()

    # Карта «артикул родителя -> его собственный источник» из листа «К загрузке»: нужна, чтобы
    # развернуть строки вида «наследование от 0200 (поставщик)» до настоящей строки прайса.
    parent_src = {}
    it = wbin["К загрузке"].iter_rows(values_only=True)
    ph = {h: i for i, h in enumerate(next(it))}
    for r in it:
        vc = str(r[ph["артикул"]] or "")
        if vc and vc not in parent_src:
            parent_src[vc] = r[ph["источник"]]
    for c in cards.values():
        vc = str(c.get("vendor_code") or "")
        if vc:
            parent_src.setdefault("title:" + vc, c.get("title"))

    razd = [list(r) for r in wbin["Раздутые на ВБ"].iter_rows(values_only=True)]
    zani = [list(r) for r in wbin["Занижённые на ВБ"].iter_rows(values_only=True)]

    ok_r, bad_r = build("Раздутые", razd, cards, sales, idx, parent_src)
    ok_z, bad_z = build("Занижённые", zani, cards, sales, idx, parent_src)

    top = sorted(
        [["Раздутые"] + r for r in ok_r[:25]] + [["Занижённые"] + r for r in ok_z[:25]],
        key=lambda x: -(x[13] or 0))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    money = [7, 8, 11, 12, 13]
    sheet(wb, "Раздутые", COLS, ok_r, money)
    sheet(wb, "Занижённые", COLS, ok_z, money)
    sheet(wb, "ТОП-50", [("группа", 12)] + COLS, top, [m + 1 for m in money])
    sheet(wb, "Без источника",
          COLS + [("почему источник не восстановлен", 60), ("группа", 12)],
          bad_r + bad_z, money)
    wb.save(OUT)

    print(f"строк прайса в индексе: {npr}")
    print(f"Раздутые       : {len(ok_r):>5}  эффект {sum(r[12] or 0 for r in ok_r):>14,.0f} ₽/год")
    print(f"Занижённые     : {len(ok_z):>5}  эффект {sum(r[12] or 0 for r in ok_z):>14,.0f} ₽/год")
    print(f"ТОП-50         : {len(top):>5}  эффект {sum(r[13] or 0 for r in top):>14,.0f} ₽/год")
    print(f"Без источника  : {len(bad_r) + len(bad_z):>5}  "
          f"эффект {sum(r[12] or 0 for r in bad_r + bad_z):>14,.0f} ₽/год")
    print(f"ИТОГО в основных листах: {sum(r[12] or 0 for r in ok_r + ok_z):,.0f} ₽/год")
    print("файл:", OUT)


if __name__ == "__main__":
    main()
