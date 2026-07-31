#!/usr/bin/env python3
# поток: gab
"""Чистка двух дефектов docs/FINAL_gabarity.xlsx.

Дефект 1 — мастер-короб: строке присвоен короб объёмом > 25 л. Для расходника это физически
   невозможно: это короб оптовой упаковки из прайса поставщика, а не короб единицы товара.
   Для карточек техники (принтер/МФУ/плоттер) такой объём законен — они не бракуются.
Дефект 2 — не тот вид товара: размер расходника взят со строки прайса, которая относится
   к технике (случай «Принтер Canon SELPHY CP810» под карточкой картриджа Canon i-SENSYS),
   или наоборот. Проверка обратным мэтчингом: если ВСЕ строки прайса этого поставщика с таким
   коробом — техника, а карточка — расходник (или наоборот), размер бракуется.

Брак снимается со всех листов с размерами и выносится на новый лист «Брак_размера».
Дети, унаследовавшие размер от забракованной матери, снимаются вместе с ней.

Без --apply — только отчёт. На маркетплейсы ничего не отправляет.
"""
from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict

sys.path.insert(0, "/opt/mp-analytics")

import openpyxl

from core import db

XLSX = "/opt/mp-analytics/docs/FINAL_gabarity.xlsx"
MAX_SINGLE_VOL_L = 25.0
HW_RX = re.compile(r"принтер|мфу|плоттер|сканер|копир|selphy|печат\w*\s+машин", re.I)
INHERIT_RX = re.compile(r"наследование от\s+(\S+)", re.I)
DIM_SHEETS = {
    "К загрузке": ("артикул", "Д", "Ш", "В", "источник"),
    "Раздутые на ВБ": ("артикул", "Д", "Ш", "В", "источник"),
    "Занижённые на ВБ": ("артикул", "Д", "Ш", "В", "источник"),
    "Волны": ("артикул", "Д", "Ш", "В", "источник"),
}


def nrm(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def fnum(x):
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return None


def key3(l, w, h):
    a = [fnum(l), fnum(w), fnum(h)]
    if not all(a):
        return None
    return tuple(round(v, 1) for v in sorted(a, reverse=True))


def main(apply_it):
    cards = {str(r["vendor_code"]): r for r in
             db.query("select vendor_code, nm_id, title from wb_cards")}

    idx = defaultdict(lambda: defaultdict(list))
    for r in db.query("select supplier, article, title, length_cm, width_cm, height_cm "
                      "from supplier_dims where length_cm is not null "
                      "and width_cm is not null and height_cm is not null"):
        k = key3(r["length_cm"], r["width_cm"], r["height_cm"])
        if k:
            idx[str(r["supplier"]).lower()][k].append(r)

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["К загрузке"]
    hdr = [c.value for c in ws[1]]
    hi = {x: i for i, x in enumerate(hdr)}
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]

    src_by_vc = {}
    for r in rows:
        vc = str(r[hi["артикул"]] or "")
        if vc:
            src_by_vc[vc] = str(r[hi["источник"]] or "")

    def root_source(vc, depth=3):
        s = src_by_vc.get(vc, "")
        while depth > 0:
            m = INHERIT_RX.search(s)
            if not m:
                return s, vc
            par = m.group(1).strip("()")
            if par not in src_by_vc or par == vc:
                return s, par
            vc, s, depth = par, src_by_vc[par], depth - 1
        return s, vc

    bad = {}
    for r in rows:
        vc = str(r[hi["артикул"]] or "")
        if not vc:
            continue
        v = key3(r[hi["Д"]], r[hi["Ш"]], r[hi["В"]])
        if not v:
            continue
        vol_l = (v[0] * v[1] * v[2]) / 1000.0
        card = cards.get(vc, {})
        title = str(card.get("title") or "")
        card_hw = bool(HW_RX.search(title))

        if vol_l > MAX_SINGLE_VOL_L and not card_hw:
            bad[vc] = ("мастер-короб",
                       f"короб {vol_l:.1f} л при пороге {MAX_SINGLE_VOL_L:.0f} л — "
                       f"оптовая упаковка, не единица товара")
            continue

        src, owner = root_source(vc)
        sup = re.split(r"[:(]", src)[0].strip().lower()
        cand = idx.get(sup, {}).get(v, [])
        if cand and title:
            hw_rows = [c for c in cand if HW_RX.search(str(c["title"] or ""))]
            tail = "" if owner == vc else f" (унаследовано от {owner})"
            if not card_hw and len(hw_rows) == len(cand):
                bad[vc] = ("не тот вид товара",
                           f"карточка — расходник, а единственный источник короба в прайсе "
                           f"«{sup}» — техника: «{str(cand[0]['title'])[:60]}»{tail}")
            elif card_hw and not hw_rows:
                bad[vc] = ("не тот вид товара",
                           f"карточка — техника, а все строки прайса «{sup}» с таким коробом — "
                           f"расходники: «{str(cand[0]['title'])[:60]}»{tail}")

    # дети забракованных матерей
    added = 1
    while added:
        added = 0
        for r in rows:
            vc = str(r[hi["артикул"]] or "")
            if not vc or vc in bad:
                continue
            m = INHERIT_RX.search(src_by_vc.get(vc, ""))
            if m:
                par = m.group(1).strip("()")
                if par in bad:
                    bad[vc] = (bad[par][0], f"размер унаследован от забракованной матери {par}")
                    added += 1

    stat = Counter(v[0] for v in bad.values())
    removed = Counter()
    detail = []
    for name, (col, dl, dw, dh, csrc) in DIM_SHEETS.items():
        if name not in wb.sheetnames:
            continue
        s = wb[name]
        h2 = {c.value: i for i, c in enumerate(s[1])}
        keep, drop = [], 0
        for r in s.iter_rows(min_row=2):
            vals = [c.value for c in r]
            vc = str(vals[h2[col]] or "")
            if vc in bad:
                drop += 1
                if name == "К загрузке":
                    detail.append([vc, vals[h2.get("nmID", 0)],
                                   str(cards.get(vc, {}).get("title") or "")[:90],
                                   f"{vals[h2[dl]]}x{vals[h2[dw]]}x{vals[h2[dh]]}",
                                   str(vals[h2[csrc]] or "")[:80], bad[vc][0], bad[vc][1]])
            else:
                keep.append(vals)
        removed[name] = drop
        if apply_it and drop:
            s.delete_rows(2, s.max_row)
            for v in keep:
                s.append(v)

    if apply_it:
        if "Брак_размера" in wb.sheetnames:
            del wb["Брак_размера"]
        bs = wb.create_sheet("Брак_размера")
        bs.append(["артикул", "nmID", "название", "наш размер", "источник", "дефект", "почему"])
        for d in sorted(detail, key=lambda x: x[5]):
            bs.append(d)
        wb.save(XLSX)

    print(f"дефектных артикулов: {len(bad)}")
    for k, v in stat.most_common():
        print(f"  {v:>5}  {k}")
    print("снято строк по листам:")
    for k, v in removed.most_common():
        print(f"  {v:>5}  {k}")
    print("ЗАПИСАНО" if apply_it else "DRY-RUN (для записи: --apply)")


if __name__ == "__main__":
    main("--apply" in sys.argv)
