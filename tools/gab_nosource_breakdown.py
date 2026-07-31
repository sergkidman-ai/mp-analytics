#!/usr/bin/env python3
# поток: gab
"""Разбор листа «Без источника» из docs/PROSMOTR_gabarity.xlsx по группам.

Отвечает на вопрос: сколько строк получили размер способом, который НЕ является прямым замером
поставщика и НЕ является прямым наследованием от матери, то есть выведен по аналогии.

Read-only. Пишет docs/nosource_breakdown.csv и сводку в чат.
"""
from __future__ import annotations

import csv
import re
import sys
from collections import Counter, defaultdict

sys.path.insert(0, "/opt/mp-analytics")

import openpyxl

from core import db

PROSMOTR = "/opt/mp-analytics/docs/PROSMOTR_gabarity.xlsx"
COV = "/opt/mp-analytics/docs/coverage.xlsx"
FINAL = "/opt/mp-analytics/docs/FINAL_gabarity.xlsx"
AUDIT = "/opt/mp-analytics/docs/web_audit_v2"
OUT = "/opt/mp-analytics/docs/nosource_breakdown.csv"

COMP_RX = re.compile(r"([A-Z0-9\-]+)\s*=\s*([\d.]+)x([\d.]+)x([\d.]+)\s*\(([^)]*)\)", re.I)
CONS_RX = re.compile(r"\s*\(согл\.\d+\)\s*$")


def nrm(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def key3(l, w, h):
    try:
        a = sorted((float(l), float(w), float(h)), reverse=True)
    except Exception:
        return None
    return tuple(round(v, 1) for v in a) if all(a) else None


def classify(reason):
    r = str(reason)
    if "короб набора" in r:
        return "наборы"
    if "веб-источник" in r:
        return "веб-находки"
    if "мастер-короб" in r:
        return "мастер-короб"
    if "OEM-код карточки ни одну не подтверждает" in r:
        return "короб-близнец"
    if "нет строки с коробом" in r:
        return "короба нет в прайсе"
    if "другой вид товара" in r:
        return "вид товара не совпал"
    if "отсутствует в supplier_dims" in r:
        return "поставщик не в supplier_dims"
    if "наш-озон" in r or "карточки Ozon" in r:
        return "перенос с нашей карточки Ozon"
    return "прочее"


def main():
    # ---------- лист «Без источника» ----------
    wb = openpyxl.load_workbook(PROSMOTR, read_only=True, data_only=True)
    ws = wb["Без источника"]
    it = ws.iter_rows(values_only=True)
    h = {x: i for i, x in enumerate(next(it))}
    rows = []
    for r in it:
        rows.append({
            "vc": str(r[h["артикул продавца"]] or ""),
            "nm": str(r[h["nmID"]] or ""),
            "title": str(r[h["название"]] or ""),
            "тип": str(r[h["тип"]] or ""),
            "наш": str(r[h["наш размер, Д×Ш×В"]] or ""),
            "эффект": r[h["эффект на логистику, ₽/год"]] or 0,
            "почему": str(r[h["почему источник не восстановлен"]] or ""),
            "группа": str(r[h["группа"]] or ""),
        })
    print(f"строк на листе «Без источника»: {len(rows)}")

    # ---------- как размер был назначен изначально (coverage.xlsx) ----------
    cv = openpyxl.load_workbook(COV, read_only=True, data_only=True)
    ws = cv["Покрыто"]
    it = ws.iter_rows(values_only=True)
    hc = {x: i for i, x in enumerate(next(it))}
    sposob, src_by_vc = {}, {}
    for r in it:
        vc = str(r[hc["vendorCode"]] or "")
        if vc:
            sposob[vc] = str(r[hc["способ"]] or "")
            src_by_vc[vc] = str(r[hc["источник"]] or "")

    # ---------- наборы: есть ли источник у КАЖДОГО компонента ----------
    idx = defaultdict(lambda: defaultdict(list))
    for r in db.query("select supplier, article, title, length_cm, width_cm, height_cm "
                      "from supplier_dims where length_cm is not null and width_cm is not null "
                      "and height_cm is not null"):
        k = key3(r["length_cm"], r["width_cm"], r["height_cm"])
        if k:
            idx[r["supplier"]][k].append(r)

    sets_comp = {}
    ws = cv["Наборы_расчёт"]
    it = ws.iter_rows(values_only=True)
    hs = {x: i for i, x in enumerate(next(it))}
    for r in it:
        vc = str(r[hs["vendorCode"]] or "")
        comps = COMP_RX.findall(str(r[hs["компоненты"]] or ""))
        if not vc or not comps:
            continue
        good = 0
        for oem, l, w, hh, sup in comps:
            base = CONS_RX.sub("", sup).split(":")[0].strip().lower()
            k = key3(l, w, hh)
            cand = idx.get(base, {}).get(k, []) if k else []
            code = nrm(oem)
            if cand and (any(code and code in (nrm(c["title"]) + nrm(c["article"])) for c in cand)
                         or len(cand) == 1):
                good += 1
        sets_comp[vc] = (good, len(comps), str(r[hs["статус"]] or ""))

    # ---------- аудит веб-находок ----------
    audit = {}
    for name in ("verified", "rejected", "inaccessible"):
        try:
            with open(f"{AUDIT}/{name}.csv", encoding="utf-8-sig", newline="") as f:
                for r in csv.DictReader(f, delimiter=";"):
                    audit[str(r["vendorCode"])] = r.get("status") or name.upper()
        except FileNotFoundError:
            pass

    # ---------- раскладка ----------
    grp = Counter()
    money = defaultdict(float)
    for r in rows:
        g = classify(r["почему"])
        r["класс"] = g
        grp[g] += 1
        money[g] += float(r["эффект"] or 0)

    set_rows = [r for r in rows if r["класс"] == "наборы"]
    full, part, unknown = 0, 0, 0
    for r in set_rows:
        sc = sets_comp.get(r["vc"])
        if not sc:
            unknown += 1
        elif sc[0] == sc[1]:
            full += 1
        else:
            part += 1

    web_rows = [r for r in rows if r["класс"] == "веб-находки"]
    wa = Counter()
    for r in web_rows:
        vc, par = r["vc"], r["vc"][:4]
        if vc in audit:
            wa[f"карточка проверена аудитом: {audit[vc]}"] += 1
        elif par in audit:
            wa[f"мать проверена аудитом: {audit[par]} (размер унаследован от неё)"] += 1
        else:
            wa["не проверялась: ни карточки, ни матери в аудите"] += 1

    # ---------- запасной ключ: колонка «источник» листа «К загрузке» FINAL ----------
    fw = openpyxl.load_workbook(FINAL, read_only=True, data_only=True)
    ws = fw["К загрузке"]
    it = ws.iter_rows(values_only=True)
    hf = {x: i for i, x in enumerate(next(it))}
    fsrc = {}
    for r in it:
        vc = str(r[hf["артикул"]] or "")
        if vc:
            fsrc[vc] = str(r[hf["источник"]] or "")

    # ---------- способ назначения: замер / наследование / вывод ----------
    kind = Counter()
    for r in rows:
        sp = sposob.get(r["vc"], "") or fsrc.get(r["vc"], "")
        if r["класс"] == "мастер-короб":
            k = "брак: мастер-короб"
        elif r["класс"] == "наборы":
            k = "расчёт по составу набора"
        elif r["класс"] == "веб-находки":
            k = "веб-источник"
        elif "наследование" in sp:
            k = "наследование от матери"
        else:
            k = "прямой мэтчинг к строке прайса (связь не подтверждается)"
        r["способ"] = sp or "—"
        r["класс_способа"] = k
        kind[k] += 1

    with open(OUT, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, ["vc", "nm", "title", "тип", "наш", "эффект", "класс",
                               "класс_способа", "способ", "группа", "почему"],
                           delimiter=";", extrasaction="ignore")
        w.writeheader()
        for r in sorted(rows, key=lambda x: -float(x["эффект"] or 0)):
            w.writerow(r)

    print("\n=== ГРУППЫ ===")
    for k, v in grp.most_common():
        print(f"  {v:>5}  {k:<32} эффект {money[k]:>12,.0f} ₽/год")
    print(f"\n=== НАБОРЫ ({len(set_rows)}) ===")
    print(f"  все компоненты с восстановимым источником : {full}")
    print(f"  часть компонентов без источника           : {part}")
    print(f"  состав набора не найден в coverage        : {unknown}")
    print(f"\n=== ВЕБ-НАХОДКИ ({len(web_rows)}) ===")
    for k, v in wa.most_common():
        print(f"  {v:>5}  {k}")
    print("\n=== СПОСОБ ПОЛУЧЕНИЯ РАЗМЕРА ===")
    for k, v in kind.most_common():
        print(f"  {v:>5}  {k}")
    not_direct = sum(v for k, v in kind.items() if k not in
                     ("наследование от матери", "брак: мастер-короб"))
    print(f"\nНЕ замер и НЕ наследование: {not_direct}")
    print("подробно:", OUT)


if __name__ == "__main__":
    main()
