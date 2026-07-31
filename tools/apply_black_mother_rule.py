#!/usr/bin/env python3
# поток: gab
"""Правило чёрной матери (CLAUDE.md, «Габариты», п.4) для листа «К загрузке».

п.4: дочерняя карточка наследует габариты материнской; для ЦВЕТНЫХ серий принтерные
карточки ВСЕГДА получают габариты ЧЁРНОГО картриджа.

Сейчас в FINAL_gabarity.xlsx дети наследуют от матери своего 4-значного префикса —
в т.ч. от голубой/пурпурной/жёлтой. Скрипт находит таких детей, подбирает чёрную
мать той же серии (по списку принтеров в названии) и переносит её габариты.

Ничего не вычисляет и не оценивает (п.1): габариты только копируются с реальной
чёрной матери. Если чёрной матери нет или её габарит неизвестен — строка уходит в
ручной список, а НЕ остаётся на цветном коробе.

Read-only по БД. Пишет: docs/black_mother/*.csv + docs/FINAL_gabarity.xlsx (лист
«К загрузке» пересобирается) только при --apply. На маркетплейсы не пишет (п.7).

Запуск:  venv/bin/python tools/apply_black_mother_rule.py [--apply]
"""
from __future__ import annotations
import sys; sys.path.insert(0, "/opt/mp-analytics")
import csv, io, os, re
from collections import defaultdict, Counter

import openpyxl
from openpyxl import Workbook

from core import db

DOCS = "/opt/mp-analytics/docs"
OUTDIR = os.path.join(DOCS, "black_mother")
XLSX = os.path.join(DOCS, "FINAL_gabarity.xlsx")

APPLY = "--apply" in sys.argv

# ---------------------------------------------------------------- цвет / серия
COLOR_WORDS = [
    ("чёрн", "BK"), ("черн", "BK"),
    ("голуб", "C"), ("циан", "C"), ("син", "C"),
    ("пурпур", "M"), ("маджент", "M"), ("малинов", "M"),
    ("жёлт", "Y"), ("желт", "Y"),
]
# светлые/спец-цвета широкоформатных — НЕ подпадают под «BK/C/M/Y» тай-брейк
SPECIAL = ("светло-", "матов", "фото", "серый", "серая")


def color_of(title: str):
    t = (title or "").lower()
    if any(s in t for s in SPECIAL):
        # светло-голубой / матовый чёрный и т.п. — отдельная позиция серии
        for k, v in COLOR_WORDS:
            if k in t:
                return "SPEC:" + v
        return "SPEC"
    for k, v in COLOR_WORDS:
        if k in t:
            return v
    return None


TYPE_WORDS = ("тонер-картридж", "картридж", "фотобарабан", "драм-картридж",
              "тонер", "лента", "чернила", "печатающая головка", "блок")


def parse_title(title: str):
    """-> (тип, код, принтеры, цвет). Принтеры = всё между 'для' и цветом."""
    t = (title or "").strip()
    low = t.lower()
    typ = next((w for w in TYPE_WORDS if low.startswith(w)), "")
    rest = t[len(typ):].strip() if typ else t
    m = re.split(r"\s+для\s+", rest, maxsplit=1, flags=re.I)
    code = m[0].strip() if m else ""
    printers = m[1].strip() if len(m) > 1 else ""
    col = color_of(t)
    # срезать цветовое слово с хвоста списка принтеров
    printers = re.sub(
        r"[\s,]*(светло-|тёмно-|темно-)?(чёрный|черный|голубой|циан[а-я]*|синий|пурпурный|"
        r"маджента|малиновый|жёлтый|желтый|матовый|фото|серый|серая)\s*$",
        "", printers, flags=re.I).strip(" ,")
    return typ, code, printers, col


def series_key(title: str):
    typ, code, printers, _ = parse_title(title)
    p = re.sub(r"[^a-zа-я0-9]+", "", (printers or "").lower())
    return (typ.lower(), p) if p else None


def fnum(x):
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return None


def vol(l, w, h):
    l, w, h = fnum(l), fnum(w), fnum(h)
    return l * w * h if (l and w and h) else None


def close_enough(a, b):
    """Тай-брейк п.9: согласие = ≤10% ИЛИ ≤1.5 см по каждой стороне (сорт. ДхШхВ)."""
    A = sorted([x for x in (fnum(a[0]), fnum(a[1]), fnum(a[2])) if x], reverse=True)
    B = sorted([x for x in (fnum(b[0]), fnum(b[1]), fnum(b[2])) if x], reverse=True)
    if len(A) != 3 or len(B) != 3:
        return False
    return all(abs(x - y) <= 1.5 or abs(x - y) <= 0.10 * max(x, y) for x, y in zip(A, B))


def code_affinity(color_code: str, black_code: str) -> int:
    """Насколько код чёрного «родной» цветному: длина общего префикса + бонус за
    одинаковую длину и одинаковый хвостовой суффикс (CF531X -> CF530X)."""
    a, b = (color_code or "").upper(), (black_code or "").upper()
    n = 0
    for x, y in zip(a, b):
        if x != y:
            break
        n += 1
    score = n * 2
    if len(a) == len(b):
        score += 3
    if a[-1:] == b[-1:]:
        score += 2
    return score


# ---------------------------------------------------------------- данные
wc = db.query("select account, nm_id, vendor_code, title, length_cm, width_cm, "
              "height_cm, weight_kg from wb_cards")
by_nm = {r["nm_id"]: r for r in wc}
mothers = {}          # vc(4) -> row
for r in wc:
    vc = str(r["vendor_code"] or "")
    if re.fullmatch(r"\d{4}", vc):
        mothers.setdefault(vc, r)

# серия -> список матерей
series = defaultdict(list)
for vc, r in mothers.items():
    k = series_key(r["title"])
    if k:
        series[k].append(vc)

wb = openpyxl.load_workbook(XLSX)
ws = wb["К загрузке"]
head = [c.value for c in ws[1]]
hi = {v: i for i, v in enumerate(head)}
data = [[c.value for c in row] for row in ws.iter_rows(min_row=2) if row[0].value is not None]

# габариты, доступные в листе загрузки, по артикулу
dims_in_sheet = {}
for r in data:
    dims_in_sheet.setdefault(str(r[hi["артикул"]]), (r[hi["Д"]], r[hi["Ш"]], r[hi["В"]],
                                                     r[hi["источник"]], r[hi["ссылка/файл"]]))

stats = Counter()
fixed, need_measure, no_black, unchanged = [], [], [], []

for r in data:
    if str(r[hi["тип"]]) != "ребёнок":
        continue
    vc = str(r[hi["артикул"]])
    pref = vc[:4]
    m = mothers.get(pref)
    if not m:
        stats["ребёнок без матери"] += 1
        continue
    mcol = color_of(m["title"])
    if mcol in (None, "BK") or (mcol or "").startswith("SPEC"):
        stats["мать чёрная/без цвета — правило не нужно"] += 1
        continue

    own = color_of(by_nm.get(r[hi["nmID"]], {}).get("title", "") if r[hi["nmID"]] in by_nm else "")
    if own is not None:
        # у самой карточки есть цвет — это НЕ принтерная карточка, а цветная позиция
        stats["ребёнок сам цветной — оставляем свой короб"] += 1
        continue

    stats["принтерная карточка на цветной матери"] += 1
    k = series_key(m["title"])
    cands = [v for v in series.get(k, []) if color_of(mothers[v]["title"]) == "BK"] if k else []
    if not cands:
        no_black.append({"артикул": vc, "nmID": r[hi["nmID"]], "мать": pref,
                         "мать_title": m["title"], "цвет_матери": mcol,
                         "причина": "чёрная мать серии не найдена"})
        stats["→ чёрной матери нет"] += 1
        continue

    _, ccode, _, _ = parse_title(m["title"])
    cands.sort(key=lambda v: -code_affinity(ccode, parse_title(mothers[v]["title"])[1]))
    black = cands[0]
    bd = dims_in_sheet.get(black)
    if not bd or not vol(bd[0], bd[1], bd[2]):
        need_measure.append({"артикул": vc, "nmID": r[hi["nmID"]], "мать": pref,
                             "чёрная_мать": black,
                             "чёрная_title": mothers[black]["title"],
                             "причина": "габарит чёрной матери неизвестен"})
        stats["→ чёрная мать без габарита"] += 1
        continue

    old = (r[hi["Д"]], r[hi["Ш"]], r[hi["В"]])
    new = (bd[0], bd[1], bd[2])
    if close_enough(old, new):
        unchanged.append({"артикул": vc, "nmID": r[hi["nmID"]], "чёрная_мать": black,
                          "было": f"{old[0]}x{old[1]}x{old[2]}",
                          "стало": f"{new[0]}x{new[1]}x{new[2]}"})
        stats["→ короб совпал (в допуске) — правка не нужна"] += 1
    else:
        fixed.append({"артикул": vc, "nmID": r[hi["nmID"]], "мать": pref,
                      "цвет_матери": mcol, "чёрная_мать": black,
                      "чёрная_title": mothers[black]["title"],
                      "было": f"{old[0]}x{old[1]}x{old[2]}",
                      "стало": f"{new[0]}x{new[1]}x{new[2]}",
                      "объём_было_л": round((vol(*old) or 0) / 1000, 2),
                      "объём_стало_л": round((vol(*new) or 0) / 1000, 2),
                      "источник": bd[3]})
        stats["→ ПЕРЕНОС габарита чёрной матери"] += 1
    if APPLY:
        r[hi["Д"]], r[hi["Ш"]], r[hi["В"]] = new
        r[hi["источник"]] = f"чёрная мать {black} ({bd[3]})"
        r[hi["ссылка/файл"]] = bd[4]

# ---------------------------------------------------------------- вывод
os.makedirs(OUTDIR, exist_ok=True)


def dump(name, rowset):
    p = os.path.join(OUTDIR, name)
    if not rowset:
        return p
    with io.open(p, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rowset[0].keys()), delimiter=";")
        w.writeheader()
        w.writerows(rowset)
    return p


dump("fixed.csv", fixed)
dump("need_black_measure.csv", need_measure)
dump("no_black_mother.csv", no_black)
dump("unchanged.csv", unchanged)

for k, v in stats.most_common():
    print(f"{v:6d}  {k}")
print()
print(f"перенос габарита : {len(fixed)}")
print(f"совпало в допуске: {len(unchanged)}")
print(f"чёрная без разм. : {len(need_measure)}")
print(f"чёрной нет вовсе : {len(no_black)}")
print(f"CSV → {OUTDIR}/")

if APPLY:
    removed = {x["артикул"] for x in need_measure} | {x["артикул"] for x in no_black}
    keep = [r for r in data if str(r[hi["артикул"]]) not in removed
            or str(r[hi["тип"]]) != "ребёнок"]
    del wb["К загрузке"]
    nws = wb.create_sheet("К загрузке", 1)
    nws.append(head)
    for r in keep:
        nws.append(r)
    wb.save(XLSX)
    print(f"\nFINAL_gabarity.xlsx обновлён: «К загрузке» {len(data)} → {len(keep)} строк "
          f"(снято {len(data) - len(keep)} детей без чёрного источника)")
else:
    print("\nDRY-RUN. Для записи в FINAL_gabarity.xlsx — запустить с --apply")
