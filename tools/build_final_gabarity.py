#!/usr/bin/env python3
# поток: gab
"""Сборка итоговой FINAL_gabarity.xlsx из всех источников покрытия.
Read-only по данным; пишет только docs/FINAL_gabarity.xlsx. На ВБ ничего не шлёт.

Листы: «К загрузке», «Расхождения», «Оставить как есть», «Раздутые на ВБ», «Сводка».
Источники: coverage.xlsx (Покрыто/Наборы/Ручной), coverage_additions.csv (поставщик/наш-озон),
web_search_v2 master_confirmed (веб-оригинал, только группа без поставщика), wb_cards (nmID+текущий
размер+вес), dims_fixlist.csv (overpay для раздутых).
"""
from __future__ import annotations
import sys; sys.path.insert(0, "/opt/mp-analytics")
import csv, io, json, re
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from core import db

def rows(p):
    with io.open(p, encoding="utf-8-sig", newline="") as f: return list(csv.DictReader(f, delimiter=";"))
def fnum(x):
    try: return float(str(x).replace(",", "."))
    except Exception: return None
def vol(l,w,h):
    l,w,h=fnum(l),fnum(w),fnum(h)
    return (l*w*h) if (l and w and h) else None

# ---------- wb_cards: мост nmID + текущий размер/вес + дети по префиксу ----------
wc=db.query("select account,nm_id,vendor_code,title,length_cm,width_cm,height_cm,weight_kg from wb_cards")
by_nm={}; by_prefix=defaultdict(list)
for r in wc:
    by_nm[r['nm_id']]=r
    vc=str(r['vendor_code'] or "")
    m=re.match(r"^(\d{4})", vc)
    if m: by_prefix[m.group(1)].append(r)
def cur_dims(nm):
    r=by_nm.get(nm)
    if not r: return (None,None,None,None)
    return (r['length_cm'],r['width_cm'],r['height_cm'],r['weight_kg'])

zag=[]   # К загрузке
seen_nm=set()
def add_row(vc,nm,ctype,L,W,H,weight,src,link):
    if nm is not None:
        if nm in seen_nm: return
        seen_nm.add(nm)
    zag.append({"артикул":vc,"nmID":nm,"тип":ctype,"Д":L,"Ш":W,"В":H,"вес":weight,
                "источник":src,"ссылка/файл":link})

# 1) Покрыто (9888) — уже с наследованием
cov=openpyxl.load_workbook("docs/coverage.xlsx", read_only=True)
ws=cov["Покрыто"]; it=ws.iter_rows(values_only=True); h=next(it); hi={x:i for i,x in enumerate(h)}
for r in it:
    if r[hi['vendorCode']] is None: continue
    nm=r[hi['nmID']]; _,_,_,wt=cur_dims(nm)
    grp=r[hi['группа']] or ""
    ctype = "мать" if grp=="мать" else ("ребёнок" if "наслед" in str(r[hi['способ']]).lower() or grp!="мать" else grp)
    add_row(str(r[hi['vendorCode']]),nm,ctype,r[hi['L']],r[hi['W']],r[hi['H']],wt,
            r[hi['источник']], r[hi['способ']])

# 2) Наборы_расчёт (790) — короб набора
ws=cov["Наборы_расчёт"]; it=ws.iter_rows(values_only=True); h=next(it); hi={x:i for i,x in enumerate(h)}
naborov=0
for r in it:
    if r[hi['vendorCode']] is None: continue
    box=str(r[hi['короб_LxWxH']] or ""); mm=re.findall(r"[\d.]+",box)
    if len(mm)<3: continue
    nm=r[hi['nmID']]; _,_,_,wt=cur_dims(nm)
    add_row(str(r[hi['vendorCode']]),nm,"набор",mm[0],mm[1],mm[2],wt,
            f"набор:{r[hi['статус']]}", f"компоненты:{r[hi['компоненты']]}")
    naborov+=1

# 3) Добавления: поставщик + наш-озон (мать + дети по префиксу)
add_src_counts=defaultdict(int)
def expand(vc,L,W,H,src_type,link,weight=None):
    # мать
    mrow=[x for x in by_prefix.get(vc[:4],[]) if str(x['vendor_code'])==vc]
    mnm=mrow[0]['nm_id'] if mrow else None
    _,_,_,wt=cur_dims(mnm) if mnm else (None,None,None,weight)
    add_row(vc,mnm,"мать",L,W,H,weight or wt,src_type,link); add_src_counts[src_type]+=1
    # дети
    for c in by_prefix.get(vc[:4],[]):
        cvc=str(c['vendor_code'])
        if cvc==vc: continue
        if c['nm_id'] in seen_nm: continue
        add_row(cvc,c['nm_id'],"ребёнок",L,W,H,c['weight_kg'],
                f"наследование от {vc} ({src_type})",link); add_src_counts[src_type+"/дети"]+=1

for a in rows("docs/queue_v2/coverage_additions.csv"):
    expand(a['vendorCode'],a['L'],a['W'],a['H'],a['source_type'],a['evidence'])

# 4) веб-оригинал: 154 матери без поставщика (группа B queue_v2) с размером из master_confirmed
gB={vc for vc,st in json.load(open("/tmp/claude-0/-opt-mp-analytics/3fdb4371-d788-41b1-9936-c159c94fe1e4/scratchpad/queue_groups.json"))["B"]}
mconf={r['vendorCode']:r for r in rows("docs/web_search_v2/final_summary/master_confirmed.csv")}
web_cnt=0
for vc in gB:
    r=mconf.get(vc)
    if not r: continue
    mm=re.findall(r"[\d.]+", r['dimensions_cm'])
    if len(mm)<3: continue
    expand(vc,mm[0],mm[1],mm[2],"веб-оригинал", r['url']); web_cnt+=1

# ---------- Раздутые на ВБ: текущий объём >> новый ----------
fix={r['nm_id']:r for r in rows("docs/dims_fixlist.csv")}
razd=[]
for z in zag:
    nm=z['nmID'];
    if nm is None: continue
    cl,cw,ch,_=cur_dims(nm); cvv=vol(cl,cw,ch); nvv=vol(z['Д'],z['Ш'],z['В'])
    if cvv and nvv and cvv>=1.5*nvv:
        ov=fix.get(nm,{}).get('overpay_rub_mo','')
        razd.append({**z,"тек_ВБ_ДxШxВ":f"{cl}x{cw}x{ch}","тек_объём_л":round(cvv/1000,2),
                     "нов_объём_л":round(nvv/1000,2),"раздутость":round(cvv/nvv,2),"переплата_₽/мес":ov})
razd.sort(key=lambda x:-x["раздутость"])

# ---------- Расхождения (нужно решение) ----------
rashod=[]
for sheet in ["Ручной_расхождения","Наборы_неясные","Ручной разбор"]:
    if sheet not in cov.sheetnames: continue
    ws=cov[sheet]; it=ws.iter_rows(values_only=True); h=next(it); hi={x:i for i,x in enumerate(h)}
    for r in it:
        if r[0] is None: continue
        rashod.append({"vendorCode":r[hi.get('vendorCode',0)],"nmID":r[hi['nmID']] if 'nmID' in hi else '',
                       "title":r[hi['title']] if 'title' in hi else '',
                       "причина":r[hi['причина']] if 'причина' in hi else '',"лист":sheet})

# ---------- Оставить как есть: остаток без источника ----------
clean=rows("docs/queue_v2/search_queue_clean.csv")
covered_add={a['vendorCode'] for a in rows("docs/queue_v2/coverage_additions.csv")}
leave=[r for r in clean if r['vendorCode'] not in covered_add and r['vendorCode'] not in gB]

# ---------- запись xlsx ----------
out=Workbook()
def sheet(name,cols,data):
    ws=out.create_sheet(name); ws.append(cols)
    for d in data: ws.append([d.get(c,"") for c in cols])
out.remove(out.active)
sheet("К загрузке",["артикул","nmID","тип","Д","Ш","В","вес","источник","ссылка/файл"],zag)
sheet("Раздутые на ВБ",["артикул","nmID","тип","Д","Ш","В","тек_ВБ_ДxШxВ","тек_объём_л","нов_объём_л","раздутость","переплата_₽/мес","источник"],razd)
sheet("Расхождения",["vendorCode","nmID","title","причина","лист"],rashod)
sheet("Оставить как есть",["vendorCode","model","manufacturer","oem","prior_status","children_dependent"],leave)

# Сводка
from collections import Counter
src_counts=Counter(z['источник'].split('(')[0].split(':')[0].strip() for z in zag)
def src_fam(s):
    s=s.lower()
    if s.startswith("наслед"): return "наследование"
    if "веб-оригинал" in s: return "веб-оригинал"
    if s.startswith("наш-озон") or "ozon" in s: return "наш-озон"
    if s.startswith("набор"): return "наборы"
    if s.startswith("поставщик"): return "поставщик(доб.)"
    return "поставщик(база)"
fam=Counter(src_fam(z['источник']) for z in zag)
total_wb=len(by_nm)
sw=out.create_sheet("Сводка")
sw.append(["Показатель","Значение"])
sw.append(["Всего карточек в wb_cards (обе acc)",total_wb])
sw.append(["К ЗАГРУЗКЕ — всего строк",len(zag)])
for k,v in fam.most_common(): sw.append([f"  из них: {k}",v])
sw.append(["Раздутые на ВБ (тек ≥1.5× нового)",len(razd)])
sw.append(["Расхождения (решение Сергея)",len(rashod)])
sw.append(["Оставить как есть (без источника)",len(leave)])
sw.append(["Наборов в загрузке",naborov])
sw.append(["веб-оригинал матерей",web_cnt])
# порядок листов: Сводка первым
out.move_sheet("Сводка", -(len(out.sheetnames)-1))
out.save("docs/FINAL_gabarity.xlsx")

print("FINAL_gabarity.xlsx записан")
print("К загрузке строк:",len(zag))
for k,v in fam.most_common(): print(f"   {k}: {v}")
print("Раздутые:",len(razd),"| Расхождения:",len(rashod),"| Оставить как есть:",len(leave),"| наборов:",naborov,"| веб-оригинал:",web_cnt)
PY_END = True
