#!/usr/bin/env python3
"""Read-only impact analysis for independently confirmed WB mother dimensions."""

from __future__ import annotations

import ast
import csv
import hashlib
import json
import math
import re
import tempfile
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from core import db
CONFIRMED = ROOT / "docs/web_search_v2/final_summary/master_confirmed.csv"
MASTER_QUEUE = ROOT / "docs/web_search_v2/master_queue.csv"
WB_XLSX = ROOT / "docs/wb_cards.xlsx"
COVERAGE_XLSX = ROOT / "docs/coverage.xlsx"
OUT = ROOT / "docs/web_search_v2/impact_dryrun"

COMPARISON_STATUSES = {
    "EXACT_MATCH",
    "SAME_DIMENSIONS_DIFFERENT_ORDER",
    "SMALL_DIFFERENCE",
    "MATERIAL_CONFLICT",
    "NO_EXISTING_DIMENSIONS",
}
COLOR_WORDS = re.compile(
    r"\b(cyan|magenta|yellow|голуб(?:ой|ая)|пурпурн(?:ый|ая)|ж[её]лт(?:ый|ая)|"
    r"синий|красный|photo\s*cyan|photo\s*magenta)\b",
    re.I,
)
BLACK_WORDS = re.compile(r"\b(black|черн(?:ый|ая)|bk)\b", re.I)
SYNTHETIC_WORDS = re.compile(
    r"\b(synthetic|derived|estimated?|median|coefficient|assum|"
    r"синтет|расч[её]т|оценк|медиан|коэффициент|предполож)\w*",
    re.I,
)


def read_csv(path: Path) -> list[dict[str, str]]:
    sample = path.read_text(encoding="utf-8-sig")[:8192]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def code(value: Any) -> str:
    value = str(value or "").strip()
    return value.zfill(4) if value.isdigit() else value


def number_list(value: Any) -> list[float]:
    text = str(value or "").strip()
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = json.loads(text)
            return [float(x) for x in parsed]
        except (ValueError, TypeError, json.JSONDecodeError):
            pass
    return [
        float(x.replace(",", "."))
        for x in re.findall(r"\d+(?:[.,]\d+)?", text)
    ]


def dims_text(values: tuple[float, float, float] | list[float] | None) -> str:
    if not values:
        return ""
    return " x ".join(f"{float(x):g}" for x in values)


def compare_dims(
    confirmed: tuple[float, float, float],
    existing: tuple[float, float, float] | None,
) -> str:
    if not existing or any(x <= 0 for x in existing):
        return "NO_EXISTING_DIMENSIONS"
    if all(math.isclose(a, b, abs_tol=0.01) for a, b in zip(confirmed, existing)):
        return "EXACT_MATCH"
    c_sorted = sorted(confirmed)
    e_sorted = sorted(existing)
    if all(math.isclose(a, b, abs_tol=0.01) for a, b in zip(c_sorted, e_sorted)):
        return "SAME_DIMENSIONS_DIFFERENT_ORDER"
    if all(abs(a - b) <= max(1.5, 0.10 * a) for a, b in zip(c_sorted, e_sorted)):
        return "SMALL_DIFFERENCE"
    return "MATERIAL_CONFLICT"


def normalized_article(value: Any) -> str:
    return re.sub(r"[^A-ZА-Я0-9]", "", str(value or "").upper())


def xlsx_rows(path: Path, sheet: str) -> list[dict[str, Any]]:
    ws = load_workbook(path, read_only=True, data_only=True)[sheet]
    iterator = ws.iter_rows(values_only=True)
    headers = list(next(iterator))
    return [dict(zip(headers, row)) for row in iterator]


def source_conversion_ok(row: dict[str, str], cm: tuple[float, float, float]) -> bool:
    raw = number_list(row["original_dimensions"])
    if len(raw) != 3:
        return False
    unit = row["original_unit"].strip().lower()
    factor = {"mm": 0.1, "cm": 1.0, "m": 100.0, "in": 2.54}.get(unit)
    if factor is None:
        return False
    converted = sorted(x * factor for x in raw)
    return all(math.isclose(a, b, abs_tol=0.02) for a, b in zip(converted, sorted(cm)))


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=fields, delimiter=";", extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(rows)


def supplier_candidates(codes: set[str]) -> dict[str, list[dict[str, Any]]]:
    products = db.query(
        """
        SELECT external_code, article
        FROM products
        WHERE external_code = ANY(%s) AND article IS NOT NULL
        """,
        (list(codes),),
    )
    ext_articles: dict[str, set[str]] = defaultdict(set)
    wanted_articles: set[str] = set()
    for row in products:
        article = normalized_article(row["article"])
        if article:
            ext_articles[code(row["external_code"])].add(article)
            wanted_articles.add(article)
    if not wanted_articles:
        return {}
    supplier_rows = db.query(
        """
        SELECT supplier, article, length_cm, width_cm, height_cm, weight_kg,
               title, src_file, loaded_at
        FROM supplier_dims
        WHERE length_cm IS NOT NULL
          AND width_cm IS NOT NULL
          AND height_cm IS NOT NULL
        """
    )
    by_article: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in supplier_rows:
        article = normalized_article(row["article"])
        if article in wanted_articles:
            by_article[article].append(row)
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for external_code, articles in ext_articles.items():
        seen: set[tuple[Any, ...]] = set()
        for article in articles:
            for row in by_article.get(article, []):
                key = (
                    row["supplier"],
                    normalized_article(row["article"]),
                    float(row["length_cm"]),
                    float(row["width_cm"]),
                    float(row["height_cm"]),
                )
                if key in seen:
                    continue
                seen.add(key)
                result[external_code].append(dict(row))
    return result


def set_components() -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    for row in db.query(
        "SELECT external_code, components FROM set_cost WHERE components IS NOT NULL"
    ):
        values = row["components"]
        if isinstance(values, str):
            try:
                values = ast.literal_eval(values)
            except (ValueError, SyntaxError):
                values = []
        result[code(row["external_code"])] = [code(x) for x in (values or [])]
    return result


def main() -> None:
    confirmed = read_csv(CONFIRMED)
    if len(confirmed) != 241 or len({code(r["vendorCode"]) for r in confirmed}) != 241:
        raise RuntimeError("master_confirmed.csv is not 241 unique vendorCode")
    master = {code(r["vendorCode"]): r for r in read_csv(MASTER_QUEUE)}
    mothers = {code(r["vendorCode"]): r for r in xlsx_rows(WB_XLSX, "Материнские")}
    children = xlsx_rows(WB_XLSX, "Дочерние")
    sets = xlsx_rows(WB_XLSX, "Наборы_подтв") + xlsx_rows(
        WB_XLSX, "Наборы_кандидаты"
    )
    covered = {
        int(r["nmID"])
        for r in xlsx_rows(COVERAGE_XLSX, "Покрыто")
        if r.get("nmID") is not None
    }
    confirmed_codes = {code(r["vendorCode"]) for r in confirmed}
    suppliers = supplier_candidates(confirmed_codes)
    components_by_set = set_components()

    supplier_detail: list[dict[str, Any]] = []
    mother_output: list[dict[str, Any]] = []
    confirmed_dims: dict[str, tuple[float, float, float]] = {}
    mother_color_policy: dict[str, str] = {}
    data_issues: Counter[str] = Counter()

    for source in sorted(confirmed, key=lambda r: code(r["vendorCode"])):
        vc = code(source["vendorCode"])
        cm_values = number_list(source["dimensions_cm"])
        if len(cm_values) != 3 or any(x <= 0 for x in cm_values):
            raise RuntimeError(f"Invalid confirmed dimensions for {vc}")
        cm = tuple(cm_values)
        confirmed_dims[vc] = cm
        wb = mothers.get(vc, {})
        wb_dims = None
        if all(wb.get(k) is not None for k in ("L_cm", "W_cm", "H_cm")):
            wb_dims = tuple(float(wb[k]) for k in ("L_cm", "W_cm", "H_cm"))
        wb_status = compare_dims(cm, wb_dims)

        manufacturer = source["manufacturer"].strip()
        manufacturer_origin = "master_confirmed"
        if not manufacturer:
            manufacturer = str(master.get(vc, {}).get("manufacturer") or "").strip()
            manufacturer_origin = "master_queue" if manufacturer else "MISSING"
        if not manufacturer:
            data_issues["manufacturer_missing"] += 1

        conversion_ok = source_conversion_ok(source, cm)
        if not conversion_ok:
            data_issues["conversion_mismatch"] += 1
        synthetic_marker = bool(
            SYNTHETIC_WORDS.search(
                " ".join(
                    source.get(k, "")
                    for k in (
                        "original_dimensions",
                        "dimension_type",
                        "source_title",
                        "evidence_quote",
                    )
                )
            )
        )
        if synthetic_marker:
            data_issues["synthetic_marker"] += 1

        supplier_statuses: list[str] = []
        supplier_comparisons: list[str] = []
        for candidate in suppliers.get(vc, []):
            sdims = tuple(
                float(candidate[k]) for k in ("length_cm", "width_cm", "height_cm")
            )
            status = compare_dims(cm, sdims)
            supplier_statuses.append(status)
            supplier_comparisons.append(
                f"{candidate['supplier']}|{candidate['article']}|"
                f"{dims_text(sdims)}|{status}"
            )
            supplier_detail.append(
                {
                    "vendorCode": vc,
                    "model": source["model"],
                    "supplier": candidate["supplier"],
                    "supplier_article": candidate["article"],
                    "confirmed_LxWxH_cm": dims_text(cm),
                    "supplier_LxWxH_cm": dims_text(sdims),
                    "comparison_status": status,
                    "supplier_title": candidate.get("title") or "",
                    "supplier_src_file": candidate.get("src_file") or "",
                    "supplier_loaded_at": str(candidate.get("loaded_at") or ""),
                    "confirmed_url": source["url"],
                    "risk": "YES" if status == "MATERIAL_CONFLICT" else "NO",
                }
            )
        if not supplier_statuses:
            supplier_status = "NO_EXISTING_DIMENSIONS"
        elif "MATERIAL_CONFLICT" in supplier_statuses:
            supplier_status = "MATERIAL_CONFLICT"
        elif "SMALL_DIFFERENCE" in supplier_statuses:
            supplier_status = "SMALL_DIFFERENCE"
        elif "SAME_DIMENSIONS_DIFFERENT_ORDER" in supplier_statuses:
            supplier_status = "SAME_DIMENSIONS_DIFFERENT_ORDER"
        else:
            supplier_status = "EXACT_MATCH"

        title_and_model = f"{wb.get('title', '')} {source['model']}"
        is_explicit_black = bool(BLACK_WORDS.search(title_and_model))
        color_policy = (
            "BLACK_MOTHER_ELIGIBLE"
            if is_explicit_black
            else "COLOR_OR_UNCLEAR_MOTHER_LINK_ONLY"
        )
        mother_color_policy[vc] = color_policy
        mother_output.append(
            {
                "vendorCode": vc,
                "nmID": wb.get("nmID", ""),
                "wb_title": wb.get("title", ""),
                "model": source["model"],
                "manufacturer": manufacturer,
                "manufacturer_origin": manufacturer_origin,
                "original_dimensions": source["original_dimensions"],
                "original_unit": source["original_unit"],
                "confirmed_LxWxH_cm": dims_text(cm),
                "dimension_type": source["dimension_type"],
                "current_WB_LxWxH_cm": dims_text(wb_dims),
                "wb_comparison_status": wb_status,
                "supplier_match_count": len(supplier_statuses),
                "supplier_comparison_status": supplier_status,
                "supplier_comparisons": " || ".join(supplier_comparisons),
                "other_independent_confirmed_sources": max(
                    0, len(source["all_confirmation_sources"].split(" | ")) - 1
                ),
                "source_conversion_ok": conversion_ok,
                "synthetic_marker_found": synthetic_marker,
                "current_coverage": "NOT_COVERED",
                "color_policy": color_policy,
                "url": source["url"],
                "source_title": source["source_title"],
                "evidence_quote": source["evidence_quote"],
                "wave": source["wave"],
                "validation_file": source["validation_file"],
            }
        )

    affected_children: list[dict[str, Any]] = []
    for child in children:
        parent = code(child.get("parent4"))
        if parent not in confirmed_codes:
            continue
        dims = confirmed_dims[parent]
        policy = mother_color_policy[parent]
        eligible = policy == "BLACK_MOTHER_ELIGIBLE"
        affected_children.append(
            {
                "child_vendorCode": code(child["vendorCode"]),
                "child_nmID": child["nmID"],
                "account": child["acc"],
                "child_title": child["title"],
                "parent_vendorCode": parent,
                "confirmed_parent_LxWxH_cm": dims_text(dims),
                "current_child_LxWxH_cm": dims_text(
                    (
                        float(child["L_cm"]),
                        float(child["W_cm"]),
                        float(child["H_cm"]),
                    )
                    if all(child.get(k) is not None for k in ("L_cm", "W_cm", "H_cm"))
                    else None
                ),
                "current_coverage": "COVERED" if int(child["nmID"]) in covered else "NOT_COVERED",
                "potential_relation": (
                    "DIRECT_PARENT_INHERITANCE"
                    if eligible
                    else "COLOR_CHILD_REQUIRES_CONFIRMED_BLACK_MOTHER_LINK"
                ),
                "forecast_eligible": "YES" if eligible else "NO",
                "dry_run_only": "YES",
            }
        )

    affected_sets: list[dict[str, Any]] = []
    for item in sets:
        set_code = code(item["vendorCode"])
        components = components_by_set.get(set_code, [])
        hits = sorted(set(components) & confirmed_codes)
        if not hits:
            continue
        affected_sets.append(
            {
                "set_vendorCode": set_code,
                "set_nmID": item["nmID"],
                "account": item["acc"],
                "set_group": item["group"],
                "set_title": item["title"],
                "confirmed_component_vendorCodes": ",".join(hits),
                "confirmed_component_count": len(hits),
                "all_known_components": ",".join(components),
                "current_set_LxWxH_cm": dims_text(
                    (
                        float(item["L_cm"]),
                        float(item["W_cm"]),
                        float(item["H_cm"]),
                    )
                    if all(item.get(k) is not None for k in ("L_cm", "W_cm", "H_cm"))
                    else None
                ),
                "action": "POTENTIALLY_AFFECTED_NO_CALCULATION",
                "dry_run_only": "YES",
            }
        )

    mother_fields = list(mother_output[0])
    supplier_fields = [
        "vendorCode",
        "model",
        "supplier",
        "supplier_article",
        "confirmed_LxWxH_cm",
        "supplier_LxWxH_cm",
        "comparison_status",
        "supplier_title",
        "supplier_src_file",
        "supplier_loaded_at",
        "confirmed_url",
        "risk",
    ]
    child_fields = list(affected_children[0]) if affected_children else ["child_vendorCode"]
    set_fields = list(affected_sets[0]) if affected_sets else ["set_vendorCode"]
    OUT.mkdir(parents=True, exist_ok=True)
    write_csv(OUT / "mother_comparison.csv", mother_output, mother_fields)
    write_csv(
        OUT / "supplier_conflicts.csv",
        [r for r in supplier_detail if r["risk"] == "YES"],
        supplier_fields,
    )
    write_csv(OUT / "affected_children.csv", affected_children, child_fields)
    write_csv(OUT / "affected_sets.csv", affected_sets, set_fields)

    wb_counts = Counter(r["wb_comparison_status"] for r in mother_output)
    supplier_mothers = [r for r in mother_output if r["supplier_match_count"]]
    supplier_match = sum(
        r["supplier_comparison_status"]
        in {
            "EXACT_MATCH",
            "SAME_DIMENSIONS_DIFFERENT_ORDER",
            "SMALL_DIFFERENCE",
        }
        for r in supplier_mothers
    )
    supplier_conflict = sum(
        r["supplier_comparison_status"] == "MATERIAL_CONFLICT"
        for r in supplier_mothers
    )
    supplier_status_counts = Counter(
        r["supplier_comparison_status"] for r in supplier_mothers
    )
    auto_children = [
        r
        for r in affected_children
        if r["forecast_eligible"] == "YES" and r["current_coverage"] == "NOT_COVERED"
    ]
    color_children = [
        r
        for r in affected_children
        if r["forecast_eligible"] == "NO"
    ]
    forecast_new = 241 + len(auto_children)
    current_covered = len(covered)
    total_cards = len(xlsx_rows(WB_XLSX, "Все"))

    summary = [
        {"metric": "confirmed_mothers_input", "value": 241},
        {"metric": "confirmed_unique_vendorCode", "value": 241},
        {"metric": "confirmed_missing_current_coverage", "value": 241},
        {"metric": "supplier_matched_mothers", "value": len(supplier_mothers)},
        {"metric": "supplier_agree_or_small_difference", "value": supplier_match},
        {"metric": "supplier_material_conflicts", "value": supplier_conflict},
        {"metric": "potential_children_total", "value": len(affected_children)},
        {"metric": "potential_auto_children_new", "value": len(auto_children)},
        {"metric": "color_children_link_only", "value": len(color_children)},
        {"metric": "potential_sets_no_calculation", "value": len(affected_sets)},
        {"metric": "current_covered_cards", "value": current_covered},
        {"metric": "forecast_newly_covered_cards", "value": forecast_new},
        {"metric": "forecast_covered_cards", "value": current_covered + forecast_new},
        {
            "metric": "forecast_coverage_percent",
            "value": round(100 * (current_covered + forecast_new) / total_cards, 2),
        },
        {"metric": "total_wb_cards_snapshot", "value": total_cards},
        {"metric": "manufacturer_missing_after_enrichment", "value": data_issues["manufacturer_missing"]},
        {"metric": "source_conversion_mismatch", "value": data_issues["conversion_mismatch"]},
        {"metric": "synthetic_marker_found", "value": data_issues["synthetic_marker"]},
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows, fields in [
        ("Summary", summary, ["metric", "value"]),
        ("Mothers", mother_output, mother_fields),
        ("Supplier comparison", supplier_detail, supplier_fields),
        ("Affected children", affected_children, child_fields),
        ("Affected sets", affected_sets, set_fields),
    ]:
        ws = workbook.create_sheet(name)
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            ws.column_dimensions[column[0].column_letter].width = min(70, max(10, width))
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=OUT, delete=False) as tmp:
        temp_path = Path(tmp.name)
    workbook.save(temp_path)
    temp_path.replace(OUT / "impact_summary.xlsx")

    risky = [
        r for r in mother_output if r["wb_comparison_status"] == "MATERIAL_CONFLICT"
    ]
    risky_wb_codes = ", ".join(r["vendorCode"] for r in risky)
    risky_supplier_codes = ", ".join(
        sorted(
            r["vendorCode"]
            for r in supplier_mothers
            if r["supplier_comparison_status"] == "MATERIAL_CONFLICT"
        )
    )
    report = f"""# Dry-run влияния 241 web-confirmed матерей на покрытие WB

Дата: {datetime.now(timezone.utc).date().isoformat()}. Анализ локальный и read-only.
`coverage.xlsx`, `wb_cards.xlsx` и Wildberries не изменялись.

## Контроль входа

- Строк: **241**, уникальных vendorCode: **241**.
- Модель/OEM, исходные размеры, единицы, нормализованные размеры, URL и цитата заполнены у всех.
- Производитель после дополнения только из `master_queue.csv`: отсутствует у **{data_issues['manufacturer_missing']}**.
- Ошибок пересчёта исходных единиц в сантиметры: **{data_issues['conversion_mismatch']}**.
- Маркеров synthetic/derived/estimate/median/coefficient: **{data_issues['synthetic_marker']}**.
- Других независимо подтверждённых строк для тех же vendorCode: **0**; конфликтов web-confirmed между источниками: **0**.

## Сравнение с текущими размерами WB

| Статус | Матерей |
|---|---:|
| EXACT_MATCH | {wb_counts['EXACT_MATCH']} |
| SAME_DIMENSIONS_DIFFERENT_ORDER | {wb_counts['SAME_DIMENSIONS_DIFFERENT_ORDER']} |
| SMALL_DIFFERENCE | {wb_counts['SMALL_DIFFERENCE']} |
| MATERIAL_CONFLICT | {wb_counts['MATERIAL_CONFLICT']} |
| NO_EXISTING_DIMENSIONS | {wb_counts['NO_EXISTING_DIMENSIONS']} |

Стороны для сравнения сортировались, но исходный порядок значений сохранён в CSV.
`SMALL_DIFFERENCE`: каждая отсортированная сторона отличается не более чем на 10% или 1,5 см.

## Поставщики

- Матерей с точным поставщицким матчем и полными Д×Ш×В: **{len(supplier_mothers)}**.
- Совпадение, перестановка сторон или малая разница: **{supplier_match}**.
- Материальный конфликт хотя бы с одним поставщиком: **{supplier_conflict}**.
- Детализация всех поставщицких сравнений сохранена в `mother_comparison.csv` и листе Excel
  `Supplier comparison`; в `supplier_conflicts.csv` вынесены только материальные конфликты.
  Объём-only и приблизительные OEM-совпадения исключены.

| Статус сравнения матери с поставщиками | Матерей |
|---|---:|
| EXACT_MATCH | {supplier_status_counts['EXACT_MATCH']} |
| SAME_DIMENSIONS_DIFFERENT_ORDER | {supplier_status_counts['SAME_DIMENSIONS_DIFFERENT_ORDER']} |
| SMALL_DIFFERENCE | {supplier_status_counts['SMALL_DIFFERENCE']} |
| MATERIAL_CONFLICT | {supplier_status_counts['MATERIAL_CONFLICT']} |
| NO_EXISTING_DIMENSIONS | {241 - len(supplier_mothers)} |

## Потенциальное влияние

- Новые подтверждённые матери вне текущего покрытия: **241**.
- Связанных дочерних WB-карточек: **{len(affected_children)}**.
- Из них потенциально добавляемых наследованием только от явно чёрной матери и ещё не покрытых: **{len(auto_children)}**.
- Цветных или неопределённых дочерних связей только для дальнейшей проверки чёрной матери: **{len(color_children)}**; в прогноз не включены.
- Наборов, содержащих хотя бы одну подтверждённую компоненту: **{len(affected_sets)}**; размеры наборов не рассчитывались.

Текущее покрытие: **{current_covered} из {total_cards}** карточек. Консервативный прогноз после
добавления 241 матерей и только разрешённых новых дочерних связей: **{current_covered + forecast_new}
из {total_cards} ({100 * (current_covered + forecast_new) / total_cards:.2f}%)**,
то есть **+{forecast_new}** карточек. Цветные связи и наборы в прирост не включены.

## Риски

- Материальных расхождений web-confirmed с текущим WB: **{len(risky)}** — это ожидаемые кандидаты на исправление, но перед записью требуют отдельного утверждения.
- Материальных расхождений с точными поставщицкими коробами: **{supplier_conflict}**; перечислены в `supplier_conflicts.csv` с `risk=YES`.
- Ни одна строка этого dry-run не является командой на изменение WB.

**Рискованные vendorCode по сравнению с WB:** {risky_wb_codes}

**Рискованные vendorCode по сравнению с поставщиками:** {risky_supplier_codes}
"""
    (OUT / "report.md").write_text(report, encoding="utf-8")

    checksum_names = [
        "mother_comparison.csv",
        "supplier_conflicts.csv",
        "affected_children.csv",
        "affected_sets.csv",
        "impact_summary.xlsx",
        "report.md",
    ]
    lines = [
        f"{hashlib.sha256((OUT / name).read_bytes()).hexdigest()}  {name}"
        for name in checksum_names
    ]
    (OUT / "SHA256SUMS.txt").write_text("\n".join(lines) + "\n", encoding="ascii")
    print(
        json.dumps(
            {
                "mothers": len(mother_output),
                "wb_statuses": dict(wb_counts),
                "supplier_matched_mothers": len(supplier_mothers),
                "supplier_agree": supplier_match,
                "supplier_conflict": supplier_conflict,
                "children": len(affected_children),
                "auto_children_new": len(auto_children),
                "color_link_only": len(color_children),
                "sets": len(affected_sets),
                "forecast_new": forecast_new,
                "forecast_covered": current_covered + forecast_new,
                "forecast_percent": round(
                    100 * (current_covered + forecast_new) / total_cards, 2
                ),
                "data_issues": dict(data_issues),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
