# поток: gab
"""Шаг 1: построчный разбор исходных файлов поставщиков.

Один парсер на файл, колонки резолвятся по имени шапки, а не по номеру.
Ни один парсер не обращается к supplier_dims и не ходит в сеть.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
import xlrd

from common import (INCOMING, Row, ST_EMPTY, ST_LOADED, ST_LOADED_NOUNIT, ST_NOTFOUND,
                    ST_SERVICE, canon, file_meta, num, set_dims, volume_check)

SEP = re.compile(r'[*xх×]', re.I)


# --- чтение листов ----------------------------------------------------------
def read_sheet(path: Path, sheet: str, header_row: int = 1):
    """Один проход по листу. Возвращает (шапка, [(номер строки, значения)], хвост).

    Пустые строки ВНУТРИ данных возвращаются как есть; сплошной пустой хвост
    Excel считается отдельно и в CSV не пишется.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    declared_max = ws.max_row or 0
    header, seen, last_nonempty = None, {}, header_row
    for i, raw in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [canon(v) for v in raw]
        if i < header_row:
            continue
        if i == header_row:
            header = vals
            continue
        if any(vals):
            last_nonempty = i
            seen[i] = vals
        elif i not in seen:
            seen[i] = vals
    wb.close()
    # материализуем непрерывный диапазон, чтобы арифметика сходилась даже если
    # openpyxl не отдал пустые строки (их нет в XML листа)
    rows = [(i, seen.get(i, [])) for i in range(header_row + 1, last_nonempty + 1)]
    tail = max(0, declared_max - last_nonempty)
    return header, rows, tail


def col_index(header, name: str) -> int:
    for i, h in enumerate(header):
        if h.strip() == name:
            return i
    raise KeyError(f'колонка «{name}» не найдена в шапке: {header[:40]}')


def get(vals, idx: int) -> str:
    return vals[idx] if idx < len(vals) else ''


def _dim_reason(triples) -> str:
    filled = [r for _, r in triples if r]
    if not filled:
        return 'пустая ячейка'
    if len(filled) < len(triples):
        return 'заполнены не все три стороны'
    return 'нечитаемый формат'


# --- cactus -----------------------------------------------------------------
def parse_cactus():
    path = INCOMING / 'Вся расходка Китай РФ Cactus GG PR.xlsx'
    meta = file_meta(path)
    sheet = 'Лист1'
    header, rows, tail = read_sheet(path, sheet)
    ci = {n: col_index(header, n) for n in (
        'PartNo', 'Brand', 'Наименование товара', 'Объем единицы', 'Вес брутто шт. кг',
        'Штрих-код EAN',
        'Кол-во ед._в_упаковке', 'Длина(ИУ) мм', 'Ширина(ИУ) мм', 'Высота(ИУ) мм',
        'Длина(МК) мм', 'Ширина(МК) мм', 'Высота(МК) мм')}
    out = []
    for rn, vals in rows:
        r = Row(source='cactus', src_file=meta['file'], src_sheet=sheet,
                src_locator=f'{sheet}!r{rn}', src_sha256=meta['sha256'])
        if not any(vals):
            r.status, r.reason = ST_EMPTY, 'пустая строка внутри данных'
            out.append(r)
            continue
        r.supplier_code, r.code_field = get(vals, ci['PartNo']), 'PartNo'
        r.barcode_raw, r.barcode_field = get(vals, ci['Штрих-код EAN']), 'Штрих-код EAN'
        r.brand = get(vals, ci['Brand'])
        r.name = get(vals, ci['Наименование товара'])
        r.weight_field, r.weight_raw = 'Вес брутто шт. кг', get(vals, ci['Вес брутто шт. кг'])
        r.volume_field, r.volume_raw = 'Объем единицы', get(vals, ci['Объем единицы'])
        r.qty_in_pack = get(vals, ci['Кол-во ед._в_упаковке'])
        iu = [(n, get(vals, ci[n])) for n in ('Длина(ИУ) мм', 'Ширина(ИУ) мм', 'Высота(ИУ) мм')]
        mk = [(n, get(vals, ci[n])) for n in ('Длина(МК) мм', 'Ширина(МК) мм', 'Высота(МК) мм')]
        ok_iu = set_dims(r, 'iu', iu, 'мм')
        ok_mk = set_dims(r, 'mk', mk, 'мм')
        if ok_iu:
            r.status = ST_LOADED
            volume_check(r, 'iu', r.volume_raw)
        else:
            r.status, r.reason = ST_NOTFOUND, _dim_reason(iu)
        # диагностика: сходится ли «Объем единицы» с МК, делённым на кратность
        qty, dec = num(r.qty_in_pack), num(r.volume_raw)
        if ok_mk and qty and dec:
            mkv = (num(r.mk_l_mm) * num(r.mk_w_mm) * num(r.mk_h_mm)) / 1e9 / qty
            r.extra = f'мк_на_шт_м3={mkv:.6f};сходится_с_объемом={"1" if abs(mkv - dec) <= 0.05 * dec else "0"}'
        out.append(r)
    return out, meta, {'sheet': sheet, 'max_row': len(rows) + 1 + tail, 'tail': tail,
                       'header_rows': 1}


# --- изи --------------------------------------------------------------------
def parse_izi():
    path = INCOMING / 'Изи.xlsx'
    meta = file_meta(path)
    sheet = 'Данные по картриджам'
    header, rows, tail = read_sheet(path, sheet)
    ci = {n: col_index(header, n) for n in (
        'артикул', 'аналог', 'бренд', 'Наименование по образцу', 'Вес', 'Размеры мм',
        'ШТ в упаковке', 'Штрихкод')}
    out = []
    for rn, vals in rows:
        r = Row(source='izi', src_file=meta['file'], src_sheet=sheet,
                src_locator=f'{sheet}!r{rn}', src_sha256=meta['sha256'])
        if not any(vals):
            r.status, r.reason = ST_EMPTY, 'пустая строка внутри данных'
            out.append(r)
            continue
        r.supplier_code, r.code_field = get(vals, ci['артикул']), 'артикул'
        r.oem_code, r.oem_field = get(vals, ci['аналог']), 'аналог'
        r.barcode_raw, r.barcode_field = get(vals, ci['Штрихкод']), 'Штрихкод'
        r.brand = get(vals, ci['бренд'])
        r.name = get(vals, ci['Наименование по образцу'])
        r.weight_field, r.weight_raw = 'Вес', get(vals, ci['Вес'])
        r.qty_in_pack = get(vals, ci['ШТ в упаковке'])
        raw = get(vals, ci['Размеры мм'])
        parts = [p.strip() for p in SEP.split(raw) if p.strip()] if raw else []
        if len(parts) == 3 and all(num(p) for p in parts):
            set_dims(r, 'iu', list(zip(['Размеры мм'] * 3, parts)), 'мм')
            r.iu_fields, r.iu_raw = 'Размеры мм', raw  # указатель — одна ячейка
            r.status = ST_LOADED
        else:
            r.iu_fields, r.iu_raw = 'Размеры мм', raw
            r.iu_unit = 'мм'
            r.status, r.reason = ST_NOTFOUND, ('пустая ячейка' if not raw else 'нечитаемый формат')
        out.append(r)
    return out, meta, {'sheet': sheet, 'max_row': len(rows) + 1 + tail, 'tail': tail,
                       'header_rows': 1}


# --- sakura -----------------------------------------------------------------
SAKURA_FILE = 'САКУРА АБДУЛ 05.06.xlsx'


def parse_sakura_laser():
    path = INCOMING / SAKURA_FILE
    meta = file_meta(path)
    sheet = 'Лазерная Sakura'
    header, rows, tail = read_sheet(path, sheet)
    ci = {n: col_index(header, n) for n in (
        'Артикул', 'Номер оригинального картриджа', 'Бренд', 'Наименование для печати',
        'Длина м', 'Ширина м', 'Высота м', 'Объем м3', 'Вес кг', 'штук в МК', 'ШК шт.')}
    out = []
    for rn, vals in rows:
        r = Row(source='sakura_laser', src_file=meta['file'], src_sheet=sheet,
                src_locator=f'{sheet}!r{rn}', src_sha256=meta['sha256'])
        if not any(vals):
            r.status, r.reason = ST_EMPTY, 'пустая строка внутри данных'
            out.append(r)
            continue
        r.supplier_code, r.code_field = get(vals, ci['Артикул']), 'Артикул'
        r.oem_code, r.oem_field = (get(vals, ci['Номер оригинального картриджа']),
                                   'Номер оригинального картриджа')
        r.barcode_raw, r.barcode_field = get(vals, ci['ШК шт.']), 'ШК шт.'
        r.brand = get(vals, ci['Бренд'])
        r.name = get(vals, ci['Наименование для печати'])
        r.weight_field, r.weight_raw = 'Вес кг', get(vals, ci['Вес кг'])
        r.volume_field, r.volume_raw = 'Объем м3', get(vals, ci['Объем м3'])
        r.qty_in_pack = get(vals, ci['штук в МК'])
        iu = [(n, get(vals, ci[n])) for n in ('Длина м', 'Ширина м', 'Высота м')]
        if set_dims(r, 'iu', iu, 'м'):
            r.status = ST_LOADED
            volume_check(r, 'iu', r.volume_raw)
        else:
            r.status, r.reason = ST_NOTFOUND, _dim_reason(iu)
        out.append(r)
    return out, meta, {'sheet': sheet, 'max_row': len(rows) + 1 + tail, 'tail': tail,
                       'header_rows': 1}


def parse_sakura_inkjet():
    path = INCOMING / SAKURA_FILE
    meta = file_meta(path)
    sheet = 'Струйная Sakura'
    header, rows, tail = read_sheet(path, sheet)
    ci = {n: col_index(header, n) for n in (
        'Код производителя Sakura', 'Артикул оригинального картриджа', 'Бренд',
        'Наименование полное', 'Вес коробки кг', 'Кол-во в упаковке', 'Обьем в м3', 'GTIN',
        'Длина коробки мм', 'Ширина коробки мм', 'Высота коробки мм',
        'Длина коробки см', 'Ширина коробки см', 'Высота коробки см',
        'Длина МК мм', 'Ширина МК мм', 'Высота МК мм')}
    out = []
    for rn, vals in rows:
        r = Row(source='sakura_inkjet', src_file=meta['file'], src_sheet=sheet,
                src_locator=f'{sheet}!r{rn}', src_sha256=meta['sha256'])
        if not any(vals):
            r.status, r.reason = ST_EMPTY, 'пустая строка внутри данных'
            out.append(r)
            continue
        r.supplier_code, r.code_field = (get(vals, ci['Код производителя Sakura']),
                                         'Код производителя Sakura')
        r.oem_code, r.oem_field = (get(vals, ci['Артикул оригинального картриджа']),
                                   'Артикул оригинального картриджа')
        r.barcode_raw, r.barcode_field = get(vals, ci['GTIN']), 'GTIN'
        r.brand = get(vals, ci['Бренд'])
        r.name = get(vals, ci['Наименование полное'])
        r.weight_field, r.weight_raw = 'Вес коробки кг', get(vals, ci['Вес коробки кг'])
        r.volume_field, r.volume_raw = 'Обьем в м3', get(vals, ci['Обьем в м3'])
        r.qty_in_pack = get(vals, ci['Кол-во в упаковке'])
        iu = [(n, get(vals, ci[n])) for n in ('Длина коробки мм', 'Ширина коробки мм', 'Высота коробки мм')]
        alt = [(n, get(vals, ci[n])) for n in ('Длина коробки см', 'Ширина коробки см', 'Высота коробки см')]
        mk = [(n, get(vals, ci[n])) for n in ('Длина МК мм', 'Ширина МК мм', 'Высота МК мм')]
        ok_iu = set_dims(r, 'iu', iu, 'мм')
        ok_alt = set_dims(r, 'alt', alt, 'см')
        set_dims(r, 'mk', mk, 'мм')
        if ok_iu:
            r.status = ST_LOADED
            volume_check(r, 'iu', r.volume_raw)
        else:
            r.status, r.reason = ST_NOTFOUND, _dim_reason(iu)
        if ok_iu and ok_alt:
            pairs = [(num(getattr(r, f'iu_{k}_mm')), num(getattr(r, f'alt_{k}_mm'))) for k in ('l', 'w', 'h')]
            r.alt_agrees = '1' if all(abs(a - b) <= 0.05 * max(a, b) for a, b in pairs) else '0'
        out.append(r)
    return out, meta, {'sheet': sheet, 'max_row': len(rows) + 1 + tail, 'tail': tail,
                       'header_rows': 1}


# --- профилайн (только справочник, размеров в файле нет) --------------------
def parse_profiline():
    path = INCOMING / 'Профилайн.xls'
    meta = file_meta(path)
    sheet = 'TDSheet'
    sh = xlrd.open_workbook(path).sheet_by_index(0)
    header = [canon(v) for v in sh.row_values(9)]  # шапка в строке 10
    ci = {n: col_index(header, n) for n in ('GUID', 'Артикул', 'Полное наименование',
                                            'Производитель', 'Запас', 'Вес', 'Объем')}
    out = []
    for i in range(sh.nrows):
        rn = i + 1
        vals = [canon(v) for v in sh.row_values(i)]
        r = Row(source='profiline', src_file=meta['file'], src_sheet=sheet,
                src_locator=f'{sheet}!r{rn}', src_sha256=meta['sha256'])
        if not any(vals):
            r.status = ST_EMPTY
            r.reason = 'пустая строка преамбулы' if rn <= 10 else 'пустая строка внутри данных'
            out.append(r)
            continue
        if rn <= 10:
            r.status, r.reason = ST_SERVICE, 'преамбула прайса / шапка'
            out.append(r)
            continue
        if not get(vals, ci['GUID']) or not get(vals, ci['Артикул']):
            r.status, r.reason = ST_SERVICE, 'строка-группировка (без GUID/артикула)'
            r.extra = get(vals, 0)[:60]
            out.append(r)
            continue
        r.supplier_code, r.code_field = get(vals, ci['Артикул']), 'Артикул'
        r.name = get(vals, ci['Полное наименование'])
        r.brand = get(vals, ci['Производитель'])
        r.weight_field, r.weight_raw = 'Вес', get(vals, ci['Вес'])
        r.volume_field, r.volume_raw = 'Объем', get(vals, ci['Объем'])
        r.extra = f'GUID={get(vals, ci["GUID"])};Запас={get(vals, ci["Запас"])}'
        # линейных размеров в файле нет — ни одна карточка не может их отсюда получить
        r.status = ST_NOTFOUND
        r.reason = 'в файле нет колонок линейных размеров'
        out.append(r)
    # header_rows=0: у профилайна В CSV пишутся ВСЕ строки, включая преамбулу,
    # поэтому вычитать шапку из арифметики нельзя.
    return out, meta, {'sheet': sheet, 'max_row': sh.nrows, 'tail': 0, 'header_rows': 0,
                       'header_row_index': 10}


# --- rapid (сохранённый ответ API) ------------------------------------------
RAPID_DIMS = ('GabarityDlina', 'GabarityShirina', 'GabarityVisota')
RAPID_UNIT_MIN_SHARE = 0.90   # порог задан Сергеем 2026-07-31
RAPID_UNIT_TOL = 0.02         # Д×Ш×В против Volume, ±2%


def rapid_dims_filled(rec) -> list[str]:
    """Заполненная сторона. Ноль в выгрузке рапида = НЕ НАЙДЕНО, а не ноль."""
    vals = [canon(rec.get(n)) for n in RAPID_DIMS]
    return [v for v in vals if v and (num(v) is None or num(v) > 0)]


def rapid_unit_probe(recs) -> dict:
    """Гипотеза «размеры рапида в метрах» проверяется объёмом из той же записи.

    Единица в файле не объявлена, поэтому проставить её можно только если
    объём подтверждает подавляющее большинство строк. Ниже порога — не ставим
    вообще (решение Сергея: «подтвердится меньше 90% — единицу не проставляй»).
    """
    per_row, ok, bad, novol = {}, 0, 0, 0
    for i, rec in enumerate(recs):
        if len(rapid_dims_filled(rec)) != 3:
            continue
        d = [num(canon(rec.get(n))) for n in RAPID_DIMS]
        vol = num(canon(rec.get('Volume')))
        if vol is None or vol <= 0:
            per_row[i], novol = 'нет_Volume', novol + 1
            continue
        dev = abs(d[0] * d[1] * d[2] - vol) / vol
        if dev <= RAPID_UNIT_TOL:
            per_row[i], ok = 'подтверждено', ok + 1
        else:
            per_row[i], bad = f'не_сходится({dev * 100:.0f}%)', bad + 1
    total = ok + bad + novol
    share = ok / total if total else 0.0
    return {'per_row': per_row, 'с_тремя_размерами': total, 'подтверждено': ok,
            'не_сходится': bad, 'без_Volume': novol, 'доля_подтверждённых': round(share, 4),
            'порог': RAPID_UNIT_MIN_SHARE, 'единица_проставлена': share >= RAPID_UNIT_MIN_SHARE,
            'единица_гипотеза': 'м'}


def parse_rapid():
    path = INCOMING / 'rapid_2026-07-18.json'
    meta = file_meta(path)
    doc = json.loads(path.read_text(encoding='utf-8'))
    recs = doc['price']
    probe = rapid_unit_probe(recs)
    unit = probe['единица_гипотеза'] if probe['единица_проставлена'] else ''
    out = []
    for i, rec in enumerate(recs):
        loc = f'price[{i}]#ID_1C2={rec.get("ID_1C2", "")}'
        r = Row(source='rapid', src_file=meta['file'], src_sheet='price',
                src_locator=loc, src_sha256=meta['sha256'])
        r.supplier_code, r.code_field = canon(rec.get('CodeID')), 'CodeID'
        r.oem_code, r.oem_field = canon(rec.get('CodeShort')), 'CodeShort'
        r.barcode_raw, r.barcode_field = canon(rec.get('Ean')), 'Ean'
        r.brand = canon(rec.get('Vendor'))
        r.name = canon(rec.get('Name'))
        r.weight_field, r.weight_raw = 'Weight', canon(rec.get('Weight'))
        r.volume_field, r.volume_raw = 'Volume', canon(rec.get('Volume'))
        r.qty_in_pack = canon(rec.get('ShtVUpak'))
        dims = [(n, canon(rec.get(n))) for n in RAPID_DIMS]
        set_dims(r, 'iu', dims, unit)
        # в выгрузке рапида пустое значение записано нулём-строкой '0'
        filled = rapid_dims_filled(rec)
        if len(filled) == 3:
            # принадлежность ИУ/МК в файле не объявлена → карантин при любой единице
            r.quarantine = '1'
            r.status = ST_LOADED if unit else ST_LOADED_NOUNIT
            r.reason = ('принадлежность ИУ/МК не объявлена в файле — не источник размера карточки'
                        if unit else
                        'единица не подтверждена объёмом; принадлежность ИУ/МК не объявлена')
            r.extra = f'объём_проверка={probe["per_row"].get(i, "")}'
        elif filled:
            r.status, r.reason = ST_NOTFOUND, 'заполнены не все три поля Gabarity'
        else:
            r.status, r.reason = ST_NOTFOUND, 'поля Gabarity пустые (записаны нулём)'
        out.append(r)
    info = {'sheet': 'price', 'max_row': len(recs), 'tail': 0, 'header_rows': 0,
            'total_field': doc.get('total'), 'date_field': doc.get('date')}
    info['проверка_единицы'] = {k: v for k, v in probe.items() if k != 'per_row'}
    return out, meta, info


PARSERS = {
    'cactus': parse_cactus,
    'izi': parse_izi,
    'sakura_laser': parse_sakura_laser,
    'sakura_inkjet': parse_sakura_inkjet,
    'profiline': parse_profiline,
    'rapid': parse_rapid,
}
