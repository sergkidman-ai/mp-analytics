#!/usr/bin/env python3
# поток: gab
"""Разбивка листа «К загрузке» на ВОЛНЫ + отдельный учёт ЗАНИЖЕННЫХ карточек.

Зачем занижённые: если реальный короб БОЛЬШЕ того, что сейчас объявлено на ВБ, площадка при
перемере ловит расхождение и штрафует. Это зеркальный риск к «раздутым» (там мы просто
переплачиваем логистику). Обе группы должны идти в первой волне.

Также проверяет два правила наследования:
  • дочерняя карточка наследует короб ТОЛЬКО от одиночного картриджа-матери;
  • НАБОР/комплект наследования не даёт (короб набора нельзя переносить на дочерние).

Read-only по БД. Пишет листы в docs/FINAL_gabarity.xlsx и CSV-выгрузки в docs/waves/.
На маркетплейсы ничего не отправляет.

Запуск:  ./venv/bin/python tools/gab_waves.py [--apply]
    без --apply  — только считает и печатает, файл не трогает
    с  --apply   — дописывает листы «Занижённые на ВБ» и «Волны» в FINAL_gabarity.xlsx
"""
from __future__ import annotations

import argparse
import csv
import io
import os
import re
import sys
from collections import Counter, defaultdict

sys.path.insert(0, "/opt/mp-analytics")

import openpyxl  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from core import db  # noqa: E402

XLSX = "docs/FINAL_gabarity.xlsx"
OUTDIR = "docs/waves"

MATERIAL = 1.5          # во сколько раз объём должен разойтись, чтобы считать расхождение материальным
SIDE_TOL_CM = 1.5       # допуск по стороне (правило п.9)

# Гейт на мастер-картон. Порог выведен ИЗ ДАННЫХ, не назначен волевым решением: среди
# одиночных карточек, где ≥2 независимых поставщика дали согласующийся короб, p99 объёма
# = 19.2 л, максимум 35 л. У позиций с ЕДИНСТВЕННЫМ источником хвост уходит до 267 л —
# это строки групповой упаковки поставщика (короб на 10-20 шт), а не наш товар. Правило 12:
# физически невозможный короб = брак строки прайса; размер отбраковываем, код не трогаем.
MASTER_CARTON_L = 20.0
# Набор, где один компонент во столько раз крупнее медианы остальных, посчитан по битой
# строке прайса — короб набора раздувается в разы (укладка идёт по крупнейшему, п.10).
SET_OUTLIER_K = 3.0
# Ставка логистики ВБ: медиана 22.66 ₽ за литр за штуку по 700 строкам docs/dims_fixlist.csv
# (overpay_rub_mo / (Δобъём × qty)). Используется ТОЛЬКО для приоритета внутри волны.
TARIFF_RUB_PER_L = 22.66
CLEAN_SUP = ("sakura", "cactus", "rapid", "изи")
SET_COMP_RX = re.compile(r"=\s*([\d.]+)x([\d.]+)x([\d.]+)")

SET_RX = re.compile(r"комплект|набор", re.I)
# мн. число носителя = набор из нескольких картриджей. «Чернила» — тоже мн. число, но это
# один товар (бутылка), не набор: их сюда включать нельзя, иначе ложные срабатывания.
PLURAL_RX = re.compile(r"^\s*(картриджи|фотобарабаны|тонер-картриджи|ленты)\b", re.I)
CONSENSUS_RX = re.compile(r"\(согл\.\d")

# Расходник-заправка: наша карточка — это ПОРОШОК/ЧЕРНИЛА (бутылка, пакет), а не картридж.
# Код модели в названии ссылается на картридж, под который заправка, поэтому мэтчинг по коду
# честно находит строку прайса КАРТРИДЖА и переносит его короб на бутылку. Это ошибка вида товара.
REFILL_RX = re.compile(r"^\s*(чернила|тонер(?![\s-]*картридж)|заправочный|заправка|тонер-порошок)", re.I)


def fnum(x):
    try:
        v = float(str(x).replace(",", "."))
        return v if v > 0 else None
    except Exception:
        return None


def vol(l, w, h):
    a, b, c = fnum(l), fnum(w), fnum(h)
    return a * b * c / 1000.0 if (a and b and c) else None


def side_excess(new, cur):
    """Максимальное превышение новой стороны над текущей, см (сортированные стороны)."""
    n = sorted(x for x in (fnum(new[0]), fnum(new[1]), fnum(new[2])) if x)
    c = sorted(x for x in (fnum(cur[0]), fnum(cur[1]), fnum(cur[2])) if x)
    if len(n) < 3 or len(c) < 3:
        return None
    return max(a - b for a, b in zip(n, c))


def is_set_title(t):
    t = t or ""
    return bool(SET_RX.search(t)) or bool(PLURAL_RX.match(t))


def load_cards():
    rows = db.query(
        "select account, nm_id, vendor_code, title, length_cm, width_cm, height_cm, weight_kg from wb_cards"
    )
    return {r["nm_id"]: dict(r) for r in rows}


def load_zag():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb["К загрузке"]
    it = ws.iter_rows(values_only=True)
    h = next(it)
    hi = {x: i for i, x in enumerate(h)}
    out = []
    for r in it:
        if r[hi["артикул"]] is None:
            continue
        out.append({k: r[hi[k]] for k in h})
    return out


def src_family(src: str) -> str:
    s = str(src or "")
    if s.startswith("наследование"):
        return "наследование"
    if s.startswith("набор"):
        return "набор"
    if "веб-оригинал" in s:
        return "веб-оригинал"
    if s.startswith("наш-озон"):
        return "наш-озон"
    return re.sub(r"\(.*", "", s).strip()


def inherit_parent(src: str) -> str | None:
    m = re.match(r"наследование от (\S+)", str(src or ""))
    return m.group(1) if m else None


def load_sales() -> dict:
    """{nm_id: продажи шт/мес} по трём последним ПОЛНЫМ месяцам margin_by_sku (WB).

    Нужны только для приоритета внутри волны: сначала то, что реально продаётся.
    """
    rows = db.query(
        """select article, sum(qty)::float / 3 as per_month
             from margin_by_sku
            where platform = 'wb'
              and period_from >= date_trunc('month', current_date) - interval '3 months'
              and period_from <  date_trunc('month', current_date)
            group by article"""
    )
    out = {}
    for r in rows:
        try:
            out[int(r["article"])] = float(r["per_month"] or 0)
        except (TypeError, ValueError):
            continue
    return out


def set_outliers() -> dict:
    """{vendorCode: пояснение} — наборы, чей короб раздут компонентом-выбросом."""
    path = "docs/coverage.xlsx"
    try:
        wb = load_workbook(path, read_only=True)
    except Exception:
        return {}
    if "Наборы_расчёт" not in wb.sheetnames:
        return {}
    ws = wb["Наборы_расчёт"]
    it = ws.iter_rows(values_only=True)
    h = next(it)
    hi = {x: i for i, x in enumerate(h)}
    bad = {}
    for r in it:
        vc = r[hi["vendorCode"]]
        if vc is None:
            continue
        comps = SET_COMP_RX.findall(str(r[hi["компоненты"]] or ""))
        if len(comps) < 2:
            continue
        vols = sorted(float(a) * float(b) * float(c) / 1000.0 for a, b, c in comps)
        rest = vols[:-1]
        med = rest[len(rest) // 2]
        if med > 0 and vols[-1] > SET_OUTLIER_K * med:
            bad[str(vc)] = (f"компонент {vols[-1]:.1f} л против {med:.1f} л у остальных — "
                            f"строка прайса похожа на групповую упаковку")
    return bad


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    cards = load_cards()
    zag = load_zag()
    sales = load_sales()
    bad_sets = set_outliers()
    print(f"К загрузке: {len(zag)} строк | карточек в wb_cards: {len(cards)}")
    print(f"продажи известны по {len(sales)} nmID (3 полных мес., margin_by_sku) | "
          f"наборов с компонентом-выбросом: {len(bad_sets)}")

    # ---------- правила наследования ----------
    by_vc_title = {}
    for c in cards.values():
        by_vc_title.setdefault(str(c["vendor_code"]), c["title"])
    bad_inherit = []
    inh = 0
    for z in zag:
        p = inherit_parent(z["источник"])
        if not p:
            continue
        inh += 1
        pt = by_vc_title.get(p)
        if pt and is_set_title(pt):
            bad_inherit.append({"ребёнок": z["артикул"], "nmID": z["nmID"],
                                "мать": p, "название_матери": pt,
                                "короб": f"{z['Д']}x{z['Ш']}x{z['В']}"})
    # наборы как источник для детей
    set_nm = {str(z["артикул"]) for z in zag if z["тип"] == "набор"}
    from_set = [z for z in zag if inherit_parent(z["источник"]) in set_nm]
    print(f"\n=== ПРАВИЛО НАСЛЕДОВАНИЯ")
    print(f"  строк-наследований: {inh}")
    print(f"  наследование от матери-НАБОРА (запрещено): {len(bad_inherit)}")
    print(f"  наследование от карточки, помеченной как набор в «К загрузке»: {len(from_set)}")

    # ---------- раздутые / занижённые ----------
    tiers = Counter()
    under, over = [], []
    for z in zag:
        nm = z["nmID"]
        c = cards.get(nm)
        if not c:
            tiers["нет карточки в выгрузке ВБ"] += 1
            continue
        cur = (c["length_cm"], c["width_cm"], c["height_cm"])
        new = (z["Д"], z["Ш"], z["В"])
        cv, nv = vol(*cur), vol(*new)
        if not cv or not nv:
            tiers["нет текущего размера на ВБ"] += 1
            continue
        exc = side_excess(new, cur)
        qty = sales.get(nm, 0.0)
        # vol() уже возвращает ЛИТРЫ (см³/1000) — второй раз на 1000 делить нельзя.
        row = {**z, "тек_ВБ_ДхШхВ": f"{c['length_cm']}x{c['width_cm']}x{c['height_cm']}",
               "тек_объём_л": round(cv, 3), "нов_объём_л": round(nv, 3),
               "кратность": round(nv / cv, 2) if cv else "",
               "макс_превышение_стороны_см": round(exc, 1) if exc is not None else "",
               "название": c["title"], "кабинет": c["account"],
               "продажи_шт_мес": round(qty, 1),
               "деньги_₽_мес": round(abs(cv - nv) * qty * TARIFF_RUB_PER_L) if qty else 0}
        if nv >= MATERIAL * cv:
            tiers["ЗАНИЖЕНО на ВБ (новый ≥1.5× текущего)"] += 1
            under.append(row)
        elif cv >= MATERIAL * nv:
            tiers["раздуто на ВБ (текущий ≥1.5× нового)"] += 1
            over.append(row)
        elif exc is not None and exc > SIDE_TOL_CM:
            tiers["занижена ОТДЕЛЬНАЯ сторона (>1.5 см), объём близок"] += 1
            row["_side_only"] = True
            under.append(row)
        else:
            tiers["совпадает в допуске"] += 1

    print(f"\n=== ТЕКУЩИЙ РАЗМЕР НА ВБ vs НОВЫЙ")
    for k, v in tiers.most_common():
        print(f"  {v:6d}  {k}")

    under.sort(key=lambda x: -(x["кратность"] if isinstance(x["кратность"], float) else 0))
    over.sort(key=lambda x: -(x["тек_объём_л"] / x["нов_объём_л"] if x["нов_объём_л"] else 0))

    vol_under = [u for u in under if not u.get("_side_only")]
    side_under = [u for u in under if u.get("_side_only")]
    print(f"\n  ЗАНИЖЕННЫЕ всего: {len(under)} = по объёму {len(vol_under)} + только по стороне {len(side_under)}")
    print(f"  из них кратность ≥3×: {sum(1 for u in vol_under if isinstance(u['кратность'], float) and u['кратность'] >= 3)}")

    # ---------- волны ----------
    def reliable(src):
        s = str(src or "")
        base = src_family(s)
        if base in ("набор", "веб-оригинал"):
            return False
        if CONSENSUS_RX.search(s):
            return True
        # наследование наследует надёжность родителя — считаем по тексту источника
        if "согл." in s:
            return True
        return False

    # карантин: строки, которые нельзя грузить без ручного решения
    bad_inh_nm = {b["nmID"] for b in bad_inherit}

    def quarantine(z, c):
        """Причина карантина или '' — почему строку нельзя грузить как есть."""
        t = str((c or {}).get("title") or "")
        if REFILL_RX.match(t):
            return "расходник-заправка: короб взят от картриджа, а товар — порошок/чернила"
        if z["nmID"] in bad_inh_nm:
            return "наследование от матери-НАБОРА (правило запрещает)"
        nv = vol(z["Д"], z["Ш"], z["В"])
        if z["тип"] == "набор":
            if str(z["артикул"]) in bad_sets:
                return "набор с компонентом-выбросом: " + bad_sets[str(z["артикул"])]
        elif nv and nv > MASTER_CARTON_L and not reliable(z["источник"]):
            return (f"подозрение на мастер-картон: короб {nv:.1f} л на одиночную карточку "
                    f"при единственном источнике (порог {MASTER_CARTON_L:.0f} л = p99 "
                    f"по позициям с согласием ≥2 поставщиков)")
        return ""

    waves = defaultdict(list)
    for z in zag:
        nm = z["nmID"]
        c = cards.get(nm)
        cv = vol(c["length_cm"], c["width_cm"], c["height_cm"]) if c else None
        nv = vol(z["Д"], z["Ш"], z["В"])
        fam = src_family(z["источник"])
        material = bool(cv and nv and (nv >= MATERIAL * cv or cv >= MATERIAL * nv))
        exc = side_excess((z["Д"], z["Ш"], z["В"]),
                          (c["length_cm"], c["width_cm"], c["height_cm"])) if (c and cv) else None
        risky_side = exc is not None and exc > SIDE_TOL_CM
        q = quarantine(z, c)
        if q:
            w = "Волна 0 — КАРАНТИН, не грузить без решения"
        elif fam in ("набор",):
            w = "Волна 4 — наборы (короб посчитан из компонентов)"
        elif fam == "веб-оригинал":
            w = "Волна 4 — веб-оригинал (упаковка оригинала, не нашей совместимки)"
        elif material or risky_side:
            w = ("Волна 1 — материальное расхождение, источник с консенсусом"
                 if reliable(z["источник"]) else
                 "Волна 2 — материальное расхождение, одиночный источник")
        else:
            w = "Волна 3 — расхождение в допуске (косметика)"
        waves[w].append({**z, "кабинет": c["account"] if c else "",
                         "название": c["title"] if c else "",
                         "тек_ВБ": f"{c['length_cm']}x{c['width_cm']}x{c['height_cm']}" if c else "",
                         "кратность_нов/тек": round(nv / cv, 2) if (cv and nv) else "",
                         "занижено": "да" if (cv and nv and nv >= MATERIAL * cv) or risky_side else "нет",
                         "занижена_сторона_см": round(exc, 1) if (exc and exc > SIDE_TOL_CM) else "",
                         "продажи_шт_мес": round(sales.get(nm, 0.0), 1),
                         # Знак осмысленный: ПЛЮС — столько маркетплейс доначислит за логистику,
                         # когда перемерит и увидит короб больше заявленного; МИНУС — столько мы
                         # экономим, сняв раздутый размер. Тариф 22.66 ₽/л/шт выведен из
                         # docs/dims_fixlist.csv (медиана по 700 строкам), не назначен.
                         # У строк с занижённой СТОРОНОЙ объём может падать — там деньги
                         # отрицательные, а риск всё равно есть: товар не влезает в заявленный габарит.
                         "деньги_₽_мес": (round((nv - cv) * sales.get(nm, 0.0) * TARIFF_RUB_PER_L)
                                          if (cv and nv and sales.get(nm)) else 0),
                         "причина_карантина": q})

    print(f"\n=== ВОЛНЫ ЗАГРУЗКИ")
    for w in sorted(waves):
        z1 = waves[w]
        u = sum(1 for x in z1 if x["занижено"] == "да")
        print(f"  {len(z1):6d}  {w}   (из них занижённых {u})")
    qcnt = Counter(x["причина_карантина"] for x in waves["Волна 0 — КАРАНТИН, не грузить без решения"])
    if qcnt:
        print("  расшифровка карантина:")
        for k, v in qcnt.most_common():
            print(f"     {v:6d}  {k}")

    # пометить занижённые/раздутые, попавшие в карантин
    qmap = {x["nmID"]: x["причина_карантина"]
            for x in waves["Волна 0 — КАРАНТИН, не грузить без решения"]}
    for lst in (under, over):
        for x in lst:
            x["карантин"] = qmap.get(x["nmID"], "")
    print(f"  из ЗАНИЖЕННЫХ в карантине: {sum(1 for x in under if x['карантин'])}"
          f" | из РАЗДУТЫХ в карантине: {sum(1 for x in over if x['карантин'])}")

    os.makedirs(OUTDIR, exist_ok=True)

    def dump(path, data, cols):
        with io.open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, delimiter=";", extrasaction="ignore")
            w.writeheader()
            w.writerows(data)

    ucols = ["артикул", "nmID", "кабинет", "название", "тип", "Д", "Ш", "В", "тек_ВБ_ДхШхВ",
             "тек_объём_л", "нов_объём_л", "кратность", "макс_превышение_стороны_см",
             "продажи_шт_мес", "деньги_₽_мес", "источник", "карантин"]
    dump(f"{OUTDIR}/zanizhennye.csv", under, ucols)
    dump(f"{OUTDIR}/razdutye.csv", over, ucols)
    wcols = ["волна", "артикул", "nmID", "кабинет", "название", "тип", "Д", "Ш", "В",
             "тек_ВБ", "кратность_нов/тек", "занижено", "занижена_сторона_см", "продажи_шт_мес", "деньги_₽_мес",
             "источник", "причина_карантина"]
    allw = []
    for w in sorted(waves):
        for x in waves[w]:
            allw.append({**x, "волна": w})
    dump(f"{OUTDIR}/waves.csv", allw, wcols)
    if bad_inherit:
        dump(f"{OUTDIR}/inherit_from_set.csv", bad_inherit,
             ["ребёнок", "nmID", "мать", "название_матери", "короб"])
    print(f"\nCSV: {OUTDIR}/zanizhennye.csv, razdutye.csv, waves.csv")

    print("\n=== ТОП-20 ВОЛНЫ 1 (по денежному/штрафному эффекту)")
    w1 = [x for x in allw if x["волна"].startswith("Волна 1")]
    # сортировка по деньгам: продажи известны не у всех, поэтому вторым ключом — кратность,
    # чтобы позиции без статистики продаж не выпадали из выборки вовсе.
    def money(x):
        return x.get("деньги_₽_мес") or 0

    w1u = sorted([x for x in w1 if x["занижено"] == "да"],
                 key=lambda x: (-money(x), -(x["кратность_нов/тек"] or 0)))
    print("  — занижённые (+₽ = доначислят при перемере; у «сторона» объём может падать):")
    for x in w1u[:10]:
        print(f"    vc={x['артикул']} nm={x['nmID']} {str(x['название'])[:50]}")
        print(f"       ВБ сейчас {x['тек_ВБ']} → реальный {x['Д']}x{x['Ш']}x{x['В']} "
              f"(×{x['кратность_нов/тек']}) | {x['продажи_шт_мес']} шт/мес "
              f"| {money(x)} ₽/мес | {x['источник']}")
    w1o = sorted([x for x in w1 if x["занижено"] == "нет"],
                 key=lambda x: (money(x), (x["кратность_нов/тек"] or 9)))
    print("  — раздутые (−₽ = экономия логистики в месяц):")
    for x in w1o[:10]:
        print(f"    vc={x['артикул']} nm={x['nmID']} {str(x['название'])[:50]}")
        print(f"       ВБ сейчас {x['тек_ВБ']} → реальный {x['Д']}x{x['Ш']}x{x['В']} "
              f"(×{x['кратность_нов/тек']}) | {x['продажи_шт_мес']} шт/мес "
              f"| {money(x)} ₽/мес | {x['источник']}")
    print(f"\n  деньги по волне 1: экономия на раздутых "
          f"{-sum(money(x) for x in w1o if money(x) < 0):,.0f} ₽/мес | доначисление по "
          f"занижённым {sum(money(x) for x in w1u if money(x) > 0):,.0f} ₽/мес "
          f"(продажи известны у {sum(1 for x in w1 if money(x))} из {len(w1)} карточек)")

    if args.apply:
        wb = load_workbook(XLSX)
        for name in ("Занижённые на ВБ", "Волны"):
            if name in wb.sheetnames:
                del wb[name]
        ws = wb.create_sheet("Занижённые на ВБ")
        ws.append(ucols)
        for x in under:
            ws.append([x.get(c, "") for c in ucols])
        ws = wb.create_sheet("Волны")
        ws.append(wcols)
        for x in allw:
            ws.append([x.get(c, "") for c in wcols])
        wb.save(XLSX)
        print(f"\nЛисты «Занижённые на ВБ» ({len(under)}) и «Волны» ({len(allw)}) записаны в {XLSX}")
    else:
        print("\n(dry-run: файл не изменён; для записи листов — запустить с --apply)")


if __name__ == "__main__":
    main()
